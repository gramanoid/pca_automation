"""
START/END Marker Validation Component
Critical component for ensuring data extraction boundaries are properly defined
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from pathlib import Path
import tempfile
import re
from typing import Dict, List, Tuple, Optional
import os
import numpy as np

class MarkerValidationComponent:
    """Component for validating and placing START/END markers in Excel files"""
    
    def __init__(self):
        # Expected sheets and table counts for different file types
        self.file_requirements = {
            'PLANNED': {
                'sheets': ['DV360', 'META', 'TIKTOK'],
                'tables_per_sheet': 1
            },
            'DELIVERED': {
                'sheet_patterns': ['DV360', 'META', 'TIKTOK'],  # Patterns to look for in sheet names
                'tables_per_sheet': 2  # R&F table and Media table
            }
        }
    
    def validate_cell_reference(self, cell_ref: str) -> bool:
        """Validate if a string is a valid Excel cell reference"""
        pattern = r'^[A-Z]+[0-9]+$'
        return bool(re.match(pattern, cell_ref.upper()))
    
    def auto_detect_table_boundaries(self, sheet_name: str, file_path: str, table_num: int = 1) -> Optional[Tuple[str, str]]:
        """
        Auto-detect table boundaries by analyzing data patterns
        Returns: Tuple of (top_left, bottom_right) or None if not detected
        """
        try:
            # Read the sheet with pandas
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Skip if empty
            if df.empty:
                return None
            
            # Find first non-empty row (potential header)
            first_data_row = None
            for idx in range(min(50, len(df))):  # Check first 50 rows
                row_data = df.iloc[idx]
                non_empty_count = row_data.notna().sum()
                if non_empty_count >= 3:  # At least 3 non-empty cells
                    first_data_row = idx
                    break
            
            if first_data_row is None:
                return None
            
            # For DELIVERED files with 2 tables, try to find the second table
            if table_num == 2:
                # Look for a gap in data (empty rows between tables)
                gap_start = None
                for idx in range(first_data_row + 2, min(len(df), 100)):
                    row_data = df.iloc[idx]
                    if row_data.notna().sum() == 0:  # Empty row
                        if gap_start is None:
                            gap_start = idx
                    elif gap_start is not None and row_data.notna().sum() >= 3:
                        # Found start of second table
                        first_data_row = idx
                        break
            
            # Find the data boundaries from the detected start
            data_subset = df.iloc[first_data_row:]
            
            # Find columns with data
            cols_with_data = []
            for col in range(len(df.columns)):
                if data_subset.iloc[:, col].notna().any():
                    cols_with_data.append(col)
            
            if not cols_with_data:
                return None
            
            first_col = min(cols_with_data)
            last_col = max(cols_with_data)
            
            # Find last row with data
            last_data_row = first_data_row
            for idx in range(len(data_subset) - 1, -1, -1):
                row_data = data_subset.iloc[idx]
                if row_data.notna().sum() > 0:
                    last_data_row = first_data_row + idx
                    break
            
            # Convert to Excel coordinates
            top_left = f"{get_column_letter(first_col + 1)}{first_data_row + 1}"
            bottom_right = f"{get_column_letter(last_col + 1)}{last_data_row + 1}"
            
            return (top_left, bottom_right)
            
        except Exception as e:
            st.warning(f"Auto-detection failed: {str(e)}")
            return None
    
    def scan_for_markers(self, file_path: str, file_type: str) -> Dict[str, Dict[str, bool]]:
        """
        Scan Excel file for START/END markers
        Returns: Dict with sheet names as keys and marker status as values
        """
        results = {}
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            # Determine which sheets to check based on file type
            if file_type == 'PLANNED':
                sheets_to_check = self.file_requirements['PLANNED']['sheets']
            else:  # DELIVERED
                # Find sheets that match the patterns
                sheets_to_check = []
                for sheet_name in wb.sheetnames:
                    for pattern in self.file_requirements['DELIVERED']['sheet_patterns']:
                        if pattern.upper() in sheet_name.upper():
                            sheets_to_check.append(sheet_name)
                            break
            
            # Check each sheet for markers
            for sheet_name in sheets_to_check:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet_results = {
                        'has_start': False,
                        'has_end': False,
                        'start_locations': [],
                        'end_locations': []
                    }
                    
                    # Scan for START and END markers
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                        for col_idx, cell in enumerate(row, start=1):
                            if cell.value and isinstance(cell.value, str):
                                if 'START' in str(cell.value).upper():
                                    sheet_results['has_start'] = True
                                    sheet_results['start_locations'].append(cell.coordinate)
                                elif 'END' in str(cell.value).upper():
                                    sheet_results['has_end'] = True
                                    sheet_results['end_locations'].append(cell.coordinate)
                    
                    # Determine if markers are valid
                    expected_tables = self.file_requirements[file_type]['tables_per_sheet']
                    
                    # For PLANNED files, markers can form a rectangle (multiple START/END)
                    # or be single markers. Check if we have at least the expected tables
                    if file_type == 'PLANNED':
                        # Valid if we have at least one START and one END marker
                        sheet_results['valid'] = (
                            len(sheet_results['start_locations']) >= 1 and
                            len(sheet_results['end_locations']) >= 1
                        )
                    else:
                        # For DELIVERED, we expect exactly 2 pairs (one for R&F, one for Media)
                        # But allow for multiple markers forming rectangles
                        sheet_results['valid'] = (
                            len(sheet_results['start_locations']) >= expected_tables and
                            len(sheet_results['end_locations']) >= expected_tables
                        )
                    
                    results[sheet_name] = sheet_results
                else:
                    results[sheet_name] = {
                        'error': f'Sheet {sheet_name} not found in file'
                    }
            
            wb.close()
            return results
            
        except Exception as e:
            return {'error': str(e)}
    
    def display_marker_validation_ui(self, file_path: str, file_type: str) -> Optional[Dict[str, List[Dict]]]:
        """
        Display UI for user-guided marker placement
        Returns: Dict with sheet names as keys and list of table boundaries as values
        """
        st.markdown(f'<div class="warning-message">⚠️ Action Required: START/END markers are missing or misplaced in your <b>{file_type}</b> file. Let\'s define your data tables so we can add them for you.</div>', unsafe_allow_html=True)
        
        # Show visual guide
        with st.expander("📖 Visual Guide - Understanding Table Boundaries", expanded=True):
            if file_type == 'PLANNED':
                st.markdown("""
                <div class="info-box">
                <h4>PLANNED File - Single Table per Sheet:</h4>
                <p>Each platform sheet (DV360, META, TIKTOK) contains <b>one data table</b> with campaign information.</p>
                <ul>
                    <li><b>Top-Left Cell:</b> The first cell of your header row (usually contains column names like 'Campaign', 'Market', etc.)</li>
                    <li><b>Bottom-Right Cell:</b> The last cell of your last data row</li>
                </ul>
                <p><b>Example:</b> If your data starts at A5 (header) and ends at G50, then:</p>
                <ul>
                    <li>Top-Left Cell = A5</li>
                    <li>Bottom-Right Cell = G50</li>
                </ul>
                <p><b>Note:</b> START marker will be placed one row above your top-left cell, END marker one row below your bottom row.</p>
                </div>
                """, unsafe_allow_html=True)
            else:  # DELIVERED
                st.markdown("""
                <div class="info-box">
                <h4>DELIVERED File - Two Tables per Sheet:</h4>
                <p>Each platform sheet contains <b>two separate data tables</b>:</p>
                <ol>
                    <li><b>R&F (Reach & Frequency) Table:</b> Usually the first table, contains reach and frequency metrics</li>
                    <li><b>Media Table:</b> Usually the second table, contains impressions, clicks, spend, etc.</li>
                </ol>
                <p><b>For each table, identify:</b></p>
                <ul>
                    <li><b>Top-Left Cell:</b> The first cell of the header row for that specific table</li>
                    <li><b>Bottom-Right Cell:</b> The last cell of the last data row for that specific table</li>
                </ul>
                <p><b>Example:</b></p>
                <ul>
                    <li>R&F Table: A5:G20 (Top-Left = A5, Bottom-Right = G20)</li>
                    <li>Media Table: A25:H50 (Top-Left = A25, Bottom-Right = H50)</li>
                </ul>
                <p><b>Note:</b> START/END markers will bracket each table separately.</p>
                </div>
                """, unsafe_allow_html=True)
        
        # Get sheet information
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            st.error(f"Error reading file: {str(e)}")
            return None
        
        # Determine required sheets based on file type
        if file_type == 'PLANNED':
            required_sheets = self.file_requirements['PLANNED']['sheets']
            tables_per_sheet = self.file_requirements['PLANNED']['tables_per_sheet']
        else:  # DELIVERED
            # Find sheets matching patterns
            required_sheets = []
            for sheet_name in sheet_names:
                for pattern in self.file_requirements['DELIVERED']['sheet_patterns']:
                    if pattern.upper() in sheet_name.upper():
                        required_sheets.append(sheet_name)
                        break
            tables_per_sheet = self.file_requirements['DELIVERED']['tables_per_sheet']
        
        # Initialize session state for boundaries if not exists
        if 'marker_boundaries' not in st.session_state:
            st.session_state.marker_boundaries = {}
        
        # Collect boundary information for each sheet and table
        all_valid = True
        boundaries = {}
        
        for sheet_name in required_sheets:
            if sheet_name not in sheet_names:
                st.error(f"Required sheet '{sheet_name}' not found in file!")
                all_valid = False
                continue
            
            st.markdown(f"### 📋 {file_type} File - Sheet: {sheet_name}")
            
            sheet_boundaries = []
            
            for table_num in range(1, tables_per_sheet + 1):
                with st.container():
                    st.markdown(f'<div class="marker-input-container">', unsafe_allow_html=True)
                    
                    if tables_per_sheet > 1:
                        st.markdown(f"#### Table {table_num}")
                        if file_type == 'DELIVERED':
                            table_hint = "R&F Data Table" if table_num == 1 else "Media Data Table"
                            st.caption(f"({table_hint})")
                    
                    # Auto-detect button
                    if st.button(f"🔍 Auto-Detect Boundaries", key=f"{sheet_name}_table{table_num}_auto"):
                        detected = self.auto_detect_table_boundaries(sheet_name, file_path, table_num)
                        if detected:
                            st.session_state[f"{sheet_name}_table{table_num}_top_left"] = detected[0]
                            st.session_state[f"{sheet_name}_table{table_num}_bottom_right"] = detected[1]
                            st.success(f"✅ Auto-detected: {detected[0]} to {detected[1]}")
                            st.rerun()
                        else:
                            st.warning("Could not auto-detect boundaries. Please enter manually.")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        top_left = st.text_input(
                            "Top-Left Cell (e.g., A5):",
                            key=f"{sheet_name}_table{table_num}_top_left",
                            value=st.session_state.get(f"{sheet_name}_table{table_num}_top_left", ""),
                            help="First cell of your header row"
                        ).upper()
                    
                    with col2:
                        bottom_right = st.text_input(
                            "Bottom-Right Cell (e.g., G50):",
                            key=f"{sheet_name}_table{table_num}_bottom_right",
                            value=st.session_state.get(f"{sheet_name}_table{table_num}_bottom_right", ""),
                            help="Last cell of your last data row"
                        ).upper()
                    
                    # Validate inputs
                    if top_left and bottom_right:
                        if not self.validate_cell_reference(top_left):
                            st.error(f"Invalid cell reference: {top_left}")
                            all_valid = False
                        elif not self.validate_cell_reference(bottom_right):
                            st.error(f"Invalid cell reference: {bottom_right}")
                            all_valid = False
                        else:
                            # Additional validation: bottom-right should be after top-left
                            try:
                                tl_col, tl_row = coordinate_to_tuple(top_left)
                                br_col, br_row = coordinate_to_tuple(bottom_right)
                                
                                if tl_row >= br_row or tl_col > br_col:
                                    st.error("Bottom-right cell must be after top-left cell!")
                                    all_valid = False
                                else:
                                    st.success(f"✅ Valid range: {top_left}:{bottom_right}")
                                    sheet_boundaries.append({
                                        'table_num': table_num,
                                        'top_left': top_left,
                                        'bottom_right': bottom_right
                                    })
                            except Exception as e:
                                st.error(f"Error validating cell references: {str(e)}")
                                all_valid = False
                    elif table_num == 1 or (top_left or bottom_right):  # Required for first table
                        if not top_left:
                            st.warning("Please enter top-left cell")
                        if not bottom_right:
                            st.warning("Please enter bottom-right cell")
                        all_valid = False
                    
                    st.markdown('</div>', unsafe_allow_html=True)
            
            if sheet_boundaries:
                boundaries[sheet_name] = sheet_boundaries
        
        # Show process button only if all inputs are valid
        if all_valid and boundaries:
            if st.button("🔧 Process and Add Markers", use_container_width=True, type="primary"):
                return boundaries
        elif not boundaries:
            st.info("Please define boundaries for at least one table to proceed.")
        
        return None
    
    def add_markers_to_file(self, file_path: str, boundaries: Dict[str, List[Dict]], output_path: str) -> bool:
        """
        Add START/END markers to Excel file based on user-defined boundaries
        """
        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            
            for sheet_name, table_boundaries in boundaries.items():
                if sheet_name not in wb.sheetnames:
                    continue
                
                sheet = wb[sheet_name]
                
                for table_info in table_boundaries:
                    top_left = table_info['top_left']
                    bottom_right = table_info['bottom_right']
                    
                    # Parse cell references
                    tl_col, tl_row = coordinate_to_tuple(top_left)
                    br_col, br_row = coordinate_to_tuple(bottom_right)
                    
                    # Place START marker (one row above top-left, same column)
                    start_row = tl_row - 1
                    if start_row < 1:
                        start_row = 1
                    start_cell = sheet.cell(row=start_row, column=tl_col)
                    start_cell.value = "START"
                    
                    # Place END marker (one row below bottom-right, same column as top-left)
                    end_row = br_row + 1
                    end_cell = sheet.cell(row=end_row, column=tl_col)
                    end_cell.value = "END"
                    
                    # Optional: Add some formatting to make markers stand out
                    from openpyxl.styles import Font, PatternFill
                    marker_font = Font(bold=True, color="FF0000")
                    marker_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
                    start_cell.font = marker_font
                    start_cell.fill = marker_fill
                    end_cell.font = marker_font
                    end_cell.fill = marker_fill
            
            # Save the modified workbook
            wb.save(output_path)
            wb.close()
            return True
            
        except Exception as e:
            st.error(f"Error adding markers: {str(e)}")
            return False
    
    def display_validation_results(self, scan_results: Dict[str, Dict[str, bool]], file_type: str):
        """Display the results of marker scanning"""
        if 'error' in scan_results:
            st.error(f"Error scanning file: {scan_results['error']}")
            return False, scan_results
        
        all_valid = True
        
        # Create a summary
        valid_sheets = []
        invalid_sheets = []
        
        for sheet_name, results in scan_results.items():
            if 'error' in results:
                invalid_sheets.append(f"{sheet_name}: {results['error']}")
                all_valid = False
            elif results.get('valid', False):
                valid_sheets.append(sheet_name)
            else:
                issues = []
                if not results.get('has_start', False):
                    issues.append("missing START marker(s)")
                if not results.get('has_end', False):
                    issues.append("missing END marker(s)")
                if issues:
                    invalid_sheets.append(f"{sheet_name}: {', '.join(issues)}")
                else:
                    invalid_sheets.append(f"{sheet_name}: incorrect number of markers")
                all_valid = False
        
        # Display summary
        if all_valid:
            st.markdown(f'<div class="success-message">✅ Critical Check: START/END markers found and correctly placed in all relevant sheets of your {file_type} file.</div>', unsafe_allow_html=True)
            
            # Show details in expander
            with st.expander(f"📋 {file_type} File - Marker Validation Details", expanded=False):
                for sheet_name in valid_sheets:
                    sheet_data = scan_results.get(sheet_name, {})
                    start_count = len(sheet_data.get('start_locations', []))
                    end_count = len(sheet_data.get('end_locations', []))
                    
                    if file_type == 'PLANNED':
                        if start_count > 1 or end_count > 1:
                            st.success(f"✓ {sheet_name}: Data region defined with {start_count} START and {end_count} END markers")
                        else:
                            st.success(f"✓ {sheet_name}: Single table with valid START/END markers")
                    else:
                        st.success(f"✓ {sheet_name}: Found {start_count} START and {end_count} END markers for R&F and Media tables")
        else:
            st.markdown(f'<div class="error-message">❌ Critical Check Failed: START/END markers are missing or incorrectly placed in your {file_type} file.</div>', unsafe_allow_html=True)
            
            # Show details
            with st.expander(f"❌ {file_type} File - Validation Issues", expanded=True):
                if file_type == 'DELIVERED':
                    st.info("Note: DELIVERED files require 2 sets of START/END markers per sheet (R&F table and Media table)")
                for issue in invalid_sheets:
                    st.error(f"✗ {issue}")
                
                st.info("You'll need to define the data table boundaries so we can add the markers for you.")
        
        return all_valid, scan_results
    
    def run_validation(self, file_path: str, file_type: str, file_key: str) -> bool:
        """
        Main method to run the complete marker validation workflow
        Returns: True if validation passes or markers are successfully added
        """
        st.markdown(f"### 🔍 Critical START/END Marker Validation - {file_type} File")
        
        # Check if we've already validated this file successfully
        if f"{file_key}_markers_validated" in st.session_state and st.session_state[f"{file_key}_markers_validated"]:
            st.success("✅ START/END markers already validated for this file")
            return True
        
        # Scan for existing markers
        with st.spinner("Scanning for START/END markers..."):
            scan_results = self.scan_for_markers(file_path, file_type)
        
        # Display results
        is_valid, scan_results = self.display_validation_results(scan_results, file_type)
        
        if is_valid:
            st.session_state[f"{file_key}_markers_validated"] = True
            return True
        
        # If not valid, show UI for marker placement
        st.markdown("---")
        boundaries = self.display_marker_validation_ui(file_path, file_type)
        
        if boundaries:
            # Create a modified file with markers
            with st.spinner("Adding START/END markers to your file..."):
                # Create output path
                original_name = Path(file_path).stem
                extension = Path(file_path).suffix
                output_filename = f"{original_name}_markers_added{extension}"
                output_path = Path(st.session_state.temp_dir) / output_filename
                
                # Add markers
                success = self.add_markers_to_file(file_path, boundaries, str(output_path))
                
                if success:
                    st.success("✅ Markers added successfully!")
                    
                    # Provide download link
                    with open(output_path, 'rb') as f:
                        file_data = f.read()
                    
                    st.markdown("### 📥 Download Modified File")
                    st.markdown("""
                    <div class="info-box">
                    <b>Important:</b> Please download the modified file below and:
                    <ol>
                        <li>Open it in Excel to verify the START/END markers are correctly placed</li>
                        <li>If correct, save it and re-upload this modified version</li>
                        <li>If there's an issue, you can try defining the boundaries again</li>
                    </ol>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.download_button(
                        label="⬇️ Download File with Markers",
                        data=file_data,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.info("After verification, please re-upload the modified file to continue.")
                    return False
                else:
                    st.error("Failed to add markers. Please check your inputs and try again.")
                    return False
        
        return False