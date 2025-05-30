"""
Marker Preview Component - Visual feedback for marker placement
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from typing import Dict, List, Tuple, Optional
import json

class MarkerPreviewComponent:
    """Component for visualizing where markers will be placed"""
    
    def __init__(self):
        self.colors = {
            'data': '#E8F4F8',
            'marker': '#FF6B6B',
            'header': '#4ECDC4',
            'empty': '#F8F9FA',
            'boundary': '#2ECC71'
        }
    
    def create_grid_visualization(self, sheet_data: pd.DataFrame, 
                                boundaries: Dict[str, Tuple[str, str]], 
                                show_markers: bool = True) -> go.Figure:
        """
        Create a visual grid showing the data region and marker placement
        """
        # Limit visualization to a reasonable size
        max_rows = min(50, len(sheet_data))
        max_cols = min(20, len(sheet_data.columns))
        
        # Create cell text and colors
        cell_values = []
        cell_colors = []
        
        for row_idx in range(max_rows):
            row_values = []
            row_colors = []
            
            for col_idx in range(max_cols):
                # Get cell value
                try:
                    value = sheet_data.iloc[row_idx, col_idx]
                    if pd.isna(value):
                        cell_text = ""
                        color = self.colors['empty']
                    else:
                        cell_text = str(value)[:20]  # Truncate long values
                        color = self.colors['data']
                except:
                    cell_text = ""
                    color = self.colors['empty']
                
                # Check if this is a boundary cell
                for table_name, (top_left, bottom_right) in boundaries.items():
                    tl_col, tl_row = coordinate_to_tuple(top_left)
                    br_col, br_row = coordinate_to_tuple(bottom_right)
                    
                    # Adjust for 0-based indexing
                    tl_col -= 1
                    tl_row -= 1
                    br_col -= 1
                    br_row -= 1
                    
                    # Check if this is a header row
                    if row_idx == tl_row and tl_col <= col_idx <= br_col:
                        color = self.colors['header']
                    # Check if this is within data region
                    elif tl_row < row_idx <= br_row and tl_col <= col_idx <= br_col:
                        if color != self.colors['header']:
                            color = self.colors['data']
                    
                    # Mark where markers will be placed
                    if show_markers:
                        # START marker position
                        if row_idx == tl_row - 1 and col_idx == tl_col:
                            cell_text = "START"
                            color = self.colors['marker']
                        # END marker position
                        elif row_idx == br_row + 1 and col_idx == tl_col:
                            cell_text = "END"
                            color = self.colors['marker']
                
                row_values.append(cell_text)
                row_colors.append(color)
            
            cell_values.append(row_values)
            cell_colors.append(row_colors)
        
        # Create header labels
        header_values = [get_column_letter(i+1) for i in range(max_cols)]
        row_labels = [str(i+1) for i in range(max_rows)]
        
        # Create the figure
        fig = go.Figure(data=[go.Table(
            header=dict(
                values=[''] + header_values,
                fill_color='lightgray',
                align='center',
                font=dict(size=10)
            ),
            cells=dict(
                values=[row_labels] + list(zip(*cell_values)),
                fill_color=[['lightgray'] * max_rows] + list(zip(*cell_colors)),
                align='center',
                font=dict(size=9),
                height=20
            )
        )])
        
        fig.update_layout(
            title="Data Region Preview (Markers in Red)",
            height=400,
            margin=dict(l=0, r=0, t=30, b=0)
        )
        
        return fig
    
    def create_simple_preview(self, top_left: str, bottom_right: str) -> str:
        """
        Create a simple text-based preview of the selected range
        """
        try:
            tl_col, tl_row = coordinate_to_tuple(top_left)
            br_col, br_row = coordinate_to_tuple(bottom_right)
            
            preview = f"""
            📊 **Selected Range: {top_left}:{bottom_right}**
            
            ```
            START marker → Row {tl_row - 1}, Column {get_column_letter(tl_col)}
            ┌─────────────────────────┐
            │ Header Row (Row {tl_row})     │ ← Your data starts here
            │ Data rows...            │
            │ ...                     │
            │ Last row (Row {br_row})       │ ← Your data ends here
            └─────────────────────────┘
            END marker → Row {br_row + 1}, Column {get_column_letter(tl_col)}
            ```
            
            **Columns:** {get_column_letter(tl_col)} to {get_column_letter(br_col)} ({br_col - tl_col + 1} columns)
            **Rows:** {tl_row} to {br_row} ({br_row - tl_row + 1} rows)
            """
            return preview
        except:
            return "Invalid cell references"
    
    def validate_cell_content(self, file_path: str, sheet_name: str, cell_ref: str) -> Dict[str, any]:
        """
        Validate and preview content at a specific cell
        """
        try:
            col, row = coordinate_to_tuple(cell_ref)
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Check if cell is within bounds
            if row > len(df) or col > len(df.columns):
                return {
                    'valid': False,
                    'message': f"Cell {cell_ref} is outside the data range",
                    'content': None
                }
            
            # Get cell content (adjust for 0-based indexing)
            content = df.iloc[row-1, col-1]
            
            if pd.isna(content):
                return {
                    'valid': True,
                    'message': f"Cell {cell_ref} is empty",
                    'content': None,
                    'warning': True
                }
            else:
                return {
                    'valid': True,
                    'message': f"Cell {cell_ref}: {str(content)[:50]}",
                    'content': content,
                    'warning': False
                }
        except Exception as e:
            return {
                'valid': False,
                'message': f"Error accessing cell: {str(e)}",
                'content': None
            }