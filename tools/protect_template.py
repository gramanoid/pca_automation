#!/usr/bin/env python3
"""
Excel Template Protection Script
Protects the output template to prevent accidental modifications
"""

import os
import sys
from pathlib import Path
import argparse
from openpyxl import load_workbook
from openpyxl.worksheet.protection import SheetProtection
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def protect_excel_template(template_path: str, password: str = None, allow_formatting: bool = True):
    """
    Protect an Excel template file to prevent modifications
    
    Args:
        template_path: Path to the Excel template file
        password: Optional password for protection (if None, protection without password)
        allow_formatting: Whether to allow cell formatting (default: True)
    """
    try:
        logger.info(f"Loading template: {template_path}")
        wb = load_workbook(template_path)
        
        # Protect each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Protecting sheet: {sheet_name}")
            
            # Create protection object
            protection = SheetProtection()
            
            # Set protection settings
            protection.sheet = True  # Protect sheet structure
            protection.objects = True  # Protect objects
            protection.scenarios = True  # Protect scenarios
            
            # Allow certain operations if specified
            if allow_formatting:
                protection.formatCells = False  # Allow cell formatting
                protection.formatColumns = False  # Allow column formatting
                protection.formatRows = False  # Allow row formatting
            
            # These are typically allowed for data entry templates
            protection.selectLockedCells = False  # Allow selecting locked cells
            protection.selectUnlockedCells = False  # Allow selecting unlocked cells
            
            # Apply password if provided
            if password:
                protection.set_password(password)
                logger.info(f"  - Password protection applied")
            else:
                logger.info(f"  - Protection applied (no password)")
            
            # Apply protection to sheet
            ws.protection = protection
            
        # Also protect the workbook structure
        if password:
            wb.security.workbookPassword = password
            wb.security.lockStructure = True
            logger.info("Workbook structure protected with password")
        else:
            wb.security.lockStructure = True
            logger.info("Workbook structure protected (no password)")
        
        # Save the protected workbook
        output_path = template_path.replace('.xlsx', '_PROTECTED.xlsx')
        wb.save(output_path)
        logger.info(f"Protected template saved to: {output_path}")
        
        return output_path
        
    except Exception as e:
        logger.error(f"Error protecting template: {e}")
        raise


def protect_template_for_script_use(template_path: str, manual_entry_cells: list = None, password: str = None):
    """
    Protect template for use by the automation script
    - All cells are locked except specified manual entry cells
    - Users can select and copy from locked cells but not edit them
    - Only Census TA and TA Population cells are editable by default
    
    Args:
        template_path: Path to the Excel template
        manual_entry_cells: List of cell references that should remain editable
        password: Optional password for protection
    """
    try:
        logger.info(f"Loading template for script-use protection: {template_path}")
        wb = load_workbook(template_path)
        
        # Default manual entry cells if not specified
        if manual_entry_cells is None:
            manual_entry_cells = [
                # DV360 Census TA and TA Population (C16:D16 and C17:D17 are merged)
                'C16', 'C17',
                # META Census TA and TA Population (C54:D54 and C55:D55 are merged)  
                'C54', 'C55',
                # TIKTOK Census TA and TA Population (C92:D92 and C93:D93 are merged)
                'C92', 'C93'
            ]
        
        # Process the active worksheet (should be the main template)
        ws = wb.active
        logger.info(f"Protecting sheet: {ws.title}")
        
        # Import Protection class
        from openpyxl.styles import Protection
        
        # First, lock all cells by default
        logger.info("Locking all cells by default...")
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)
        
        # Then unlock only the manual entry cells
        logger.info("Unlocking manual entry cells...")
        for cell_ref in manual_entry_cells:
            try:
                cell = ws[cell_ref]
                cell.protection = Protection(locked=False)
                logger.info(f"  - Unlocked cell: {cell_ref}")
                
                # Also handle merged cells - unlock all cells in the merge
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        for row in ws[merged_range.coord]:
                            for merged_cell in row:
                                merged_cell.protection = Protection(locked=False)
                        logger.info(f"    (Part of merged range: {merged_range.coord})")
            except Exception as e:
                logger.warning(f"Could not unlock cell {cell_ref}: {e}")
        
        # Apply sheet protection with specific settings
        protection = SheetProtection()
        
        # Core protection settings
        protection.sheet = True  # Enable sheet protection
        protection.objects = True  # Protect objects
        protection.scenarios = True  # Protect scenarios
        
        # IMPORTANT: Allow users to select locked cells (for copying)
        protection.selectLockedCells = False  # False means users CAN select locked cells
        protection.selectUnlockedCells = False  # False means users CAN select unlocked cells
        
        # Prevent all modifications except on unlocked cells
        protection.formatCells = True  # Prevent formatting changes
        protection.formatColumns = False  # ALLOW column width changes
        protection.formatRows = False  # ALLOW row height changes
        protection.insertColumns = True  # Prevent inserting columns
        protection.insertRows = True  # Prevent inserting rows
        protection.deleteColumns = True  # Prevent deleting columns
        protection.deleteRows = True  # Prevent deleting rows
        protection.sort = True  # Prevent sorting
        protection.autoFilter = True  # Prevent auto filter changes
        protection.pivotTables = True  # Prevent pivot table changes
        
        # Apply password if provided
        if password:
            protection.set_password(password)
            logger.info("Password protection applied")
        
        # Apply protection to sheet
        ws.protection = protection
        logger.info("Sheet protection applied successfully")
        
        # Also protect the workbook structure to prevent sheet modifications
        try:
            if password:
                wb.security.workbookPassword = password
                wb.security.lockStructure = True
            else:
                wb.security.lockStructure = True
            logger.info("Workbook structure protected")
        except AttributeError:
            # Some Excel files don't have security attribute
            logger.info("Note: Workbook structure protection not available for this file format")
        
        # Save the protected workbook
        output_path = template_path.replace('.xlsx', '_PROTECTED_FOR_SCRIPT.xlsx')
        wb.save(output_path)
        logger.info(f"\nProtected template saved to: {output_path}")
        logger.info("\nProtection summary:")
        logger.info("- All cells locked except Census TA and TA Population fields")
        logger.info("- Users can select and copy from any cell")
        logger.info("- Only manual entry cells can be edited")
        logger.info("- Template structure is protected from changes")
        
        return output_path
        
    except Exception as e:
        logger.error(f"Error protecting template: {e}")
        raise


def unlock_cells_for_data_entry(template_path: str, data_ranges: list, password: str = None):
    """
    [DEPRECATED - Use protect_template_for_script_use instead]
    Protect template but unlock specific cells for data entry
    
    Args:
        template_path: Path to the Excel template
        data_ranges: List of cell ranges to keep unlocked (e.g., ['C18:P23', 'C25:P42'])
        password: Optional password for protection
    """
    try:
        logger.info(f"Loading template for selective protection: {template_path}")
        wb = load_workbook(template_path)
        
        # Import Protection class
        from openpyxl.styles import Protection
        
        # Process each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Processing sheet: {sheet_name}")
            
            # First, lock all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=True)
            
            # Then unlock specified ranges
            for cell_range in data_ranges:
                logger.info(f"  - Unlocking range: {cell_range}")
                for row in ws[cell_range]:
                    for cell in row:
                        cell.protection = Protection(locked=False)
            
            # Apply sheet protection
            protection = SheetProtection()
            protection.sheet = True
            
            # Allow users to select and edit unlocked cells
            protection.selectLockedCells = True
            protection.selectUnlockedCells = True
            
            if password:
                protection.set_password(password)
            
            ws.protection = protection
        
        # Save the protected workbook
        output_path = template_path.replace('.xlsx', '_PROTECTED_WITH_EDITABLE_CELLS.xlsx')
        wb.save(output_path)
        logger.info(f"Selectively protected template saved to: {output_path}")
        
        return output_path
        
    except Exception as e:
        logger.error(f"Error in selective protection: {e}")
        raise


def main():
    """CLI entry point"""
    parser = argparse.ArgumentParser(description='Protect Excel Template')
    parser.add_argument('--template', required=True, help='Path to Excel template file')
    parser.add_argument('--password', help='Password for protection (optional)')
    parser.add_argument('--mode', choices=['full', 'selective', 'script'], default='script',
                       help='Protection mode: full (lock everything), selective (custom ranges), or script (for automation use)')
    parser.add_argument('--allow-formatting', action='store_true',
                       help='Allow cell formatting in protected sheets')
    parser.add_argument('--manual-cells', nargs='*', 
                       help='Additional cells to keep unlocked for manual entry (e.g., C16 D54)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.template):
        logger.error(f"Template file not found: {args.template}")
        return 1
    
    try:
        if args.mode == 'full':
            # Full protection
            output = protect_excel_template(
                args.template, 
                args.password,
                args.allow_formatting
            )
        elif args.mode == 'script':
            # Protection for script use - only Census TA and TA Population editable
            manual_cells = args.manual_cells if args.manual_cells else None
            output = protect_template_for_script_use(
                args.template,
                manual_cells,
                args.password
            )
        else:
            # Selective protection - define which cells should remain editable
            # These are the data entry cells based on the new template structure
            editable_ranges = [
                # DV360 Campaign Level data (both planned and actuals)
                'C18:P23',
                # DV360 Awareness/Consideration/Purchase data  
                'C25:P42',
                
                # META Campaign Level data
                'C56:P61', 
                # META Awareness/Consideration/Purchase data
                'C63:P80',
                
                # TIKTOK Campaign Level data
                'C94:P99',
                # TIKTOK Awareness/Consideration/Purchase data
                'C101:P118',
                
                # Budget values in header
                'B3:B4',
                
                # Campaign name and dates
                'A1:A2', 'A5'
            ]
            
            from openpyxl.styles import Protection
            output = unlock_cells_for_data_entry(
                args.template,
                editable_ranges,
                args.password
            )
        
        print(f"\n✅ Template protection completed!")
        print(f"   Output: {output}")
        if args.password:
            print(f"   Password: {args.password}")
            print(f"   ⚠️  Keep this password safe - you'll need it to unprotect the template!")
        
        return 0
        
    except Exception as e:
        logger.error(f"Protection failed: {e}")
        return 1


if __name__ == '__main__':
    exit(main())