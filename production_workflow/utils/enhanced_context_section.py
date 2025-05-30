"""
Enhanced Additional Context Section for Simple LLM Mapper
This module provides an enhanced version of the _write_additional_context method
with improved aesthetics and content quality.
"""

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
from datetime import datetime


def _write_additional_context_enhanced(self, ws, combined_df):
    """Write enhanced additional context data with improved aesthetics and insights"""
    logger.info("Writing enhanced additional context data...")
    
    # Define colors for aesthetic enhancement
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue
    SUBHEADER_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")  # Light blue
    ALTERNATING_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
    METRIC_FILL = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")  # Very light blue
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    thick_bottom = Border(
        bottom=Side(style='thick', color='1F4E78')
    )
    
    # Start position
    summary_start_row = 125
    current_row = summary_start_row
    
    # Main header with enhanced styling
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "CAMPAIGN INSIGHTS & ADDITIONAL CONTEXT"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{current_row}'].fill = HEADER_FILL
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    
    # Add thick bottom border to header
    for col in range(1, 9):  # A to H
        ws.cell(row=current_row, column=col).border = thick_bottom
    
    current_row += 2
    
    # Section 1: Campaign Overview
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "📊 CAMPAIGN OVERVIEW"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{current_row}'].fill = SUBHEADER_FILL
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Calculate key metrics
    total_budget = combined_df['BUDGET_LOCAL'].sum() if 'BUDGET_LOCAL' in combined_df.columns else 0
    total_impressions = combined_df['IMPRESSIONS'].sum() if 'IMPRESSIONS' in combined_df.columns else 0
    unique_markets = len(combined_df['MARKET'].dropna().unique()) if 'MARKET' in combined_df.columns else 0
    
    # Campaign duration
    if 'START_DATE' in combined_df.columns and 'END_DATE' in combined_df.columns:
        start_date = pd.to_datetime(combined_df['START_DATE'].dropna()).min()
        end_date = pd.to_datetime(combined_df['END_DATE'].dropna()).max()
        campaign_days = (end_date - start_date).days if pd.notna(start_date) and pd.notna(end_date) else 0
    else:
        campaign_days = 0
    
    # Write overview metrics in a grid layout
    overview_metrics = [
        ("Total Investment", f"£{total_budget:,.0f}", "The complete media investment across all platforms"),
        ("Total Impressions", f"{total_impressions:,.0f}", "Combined reach across all campaigns"),
        ("Markets Covered", f"{unique_markets} markets", "Geographic spread of the campaign"),
        ("Campaign Duration", f"{campaign_days} days", "Total runtime of the campaign"),
        ("Data Completeness", f"{combined_df.notna().sum().sum() / combined_df.size * 100:.1f}%", "Percentage of non-empty data fields"),
        ("Unique Creatives", f"{len(combined_df['CREATIVE_NAME'].dropna().unique()) if 'CREATIVE_NAME' in combined_df.columns else 0}", "Number of different creative assets used")
    ]
    
    for i, (metric, value, description) in enumerate(overview_metrics):
        row = current_row + (i // 2)
        col_offset = (i % 2) * 4  # 0 or 4
        
        # Metric name
        ws[f'{get_column_letter(1 + col_offset)}{row}'] = metric
        ws[f'{get_column_letter(1 + col_offset)}{row}'].font = Font(bold=True, size=10)
        ws[f'{get_column_letter(1 + col_offset)}{row}'].fill = METRIC_FILL
        
        # Metric value
        ws[f'{get_column_letter(2 + col_offset)}{row}'] = value
        ws[f'{get_column_letter(2 + col_offset)}{row}'].font = Font(size=10, color="1F4E78")
        ws[f'{get_column_letter(2 + col_offset)}{row}'].alignment = Alignment(horizontal='right')
        
        # Description (merged across 2 cells)
        ws.merge_cells(f'{get_column_letter(3 + col_offset)}{row}:{get_column_letter(4 + col_offset)}{row}')
        ws[f'{get_column_letter(3 + col_offset)}{row}'] = description
        ws[f'{get_column_letter(3 + col_offset)}{row}'].font = Font(size=9, italic=True, color="666666")
        
        # Apply borders
        for col in range(1 + col_offset, 5 + col_offset):
            ws.cell(row=row, column=col).border = thin_border
    
    current_row += 4
    
    # Section 2: Platform Performance Summary
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "📈 PLATFORM PERFORMANCE SUMMARY"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{current_row}'].fill = SUBHEADER_FILL
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Platform breakdown
    if 'PLATFORM' in combined_df.columns and 'BUDGET_LOCAL' in combined_df.columns:
        platform_summary = combined_df.groupby('PLATFORM')['BUDGET_LOCAL'].agg(['sum', 'count'])
        platform_summary['percentage'] = (platform_summary['sum'] / platform_summary['sum'].sum() * 100).round(1)
        platform_summary = platform_summary.sort_values('sum', ascending=False)
        
        # Headers
        headers = ['Platform', 'Investment', '% of Total', 'Campaigns', 'Avg per Campaign']
        for i, header in enumerate(headers):
            ws[f'{get_column_letter(i+1)}{current_row}'] = header
            ws[f'{get_column_letter(i+1)}{current_row}'].font = Font(bold=True, size=10)
            ws[f'{get_column_letter(i+1)}{current_row}'].fill = METRIC_FILL
            ws[f'{get_column_letter(i+1)}{current_row}'].border = thin_border
        
        current_row += 1
        
        # Platform data
        for idx, (platform, data) in enumerate(platform_summary.iterrows()):
            # Apply alternating row colors
            fill = ALTERNATING_FILL if idx % 2 == 1 else None
            
            ws[f'A{current_row}'] = platform
            ws[f'B{current_row}'] = f"£{data['sum']:,.0f}"
            ws[f'C{current_row}'] = f"{data['percentage']:.1f}%"
            ws[f'D{current_row}'] = f"{int(data['count'])}"
            ws[f'E{current_row}'] = f"£{data['sum']/data['count']:,.0f}"
            
            # Apply styling
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if fill:
                    cell.fill = fill
                if col > 1:  # Right-align numbers
                    cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
    
    current_row += 1
    
    # Section 3: Campaign Elements
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "🎯 CAMPAIGN ELEMENTS & TARGETING"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{current_row}'].fill = SUBHEADER_FILL
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Enhanced context fields with better organization
    context_sections = [
        {
            'title': 'Targeting & Delivery',
            'fields': [
                ('DEVICE', 'Device Targeting', '📱'),
                ('PLACEMENT', 'Ad Placements', '📍'),
                ('BUYING_MODEL', 'Buying Models', '💰'),
                ('CEJ_OBJECTIVES', 'Campaign Objectives', '🎯')
            ]
        },
        {
            'title': 'Creative & Content',
            'fields': [
                ('FORMAT_TYPE', 'Ad Formats', '🎨'),
                ('CREATIVE_NAME', 'Creative Assets', '🖼️'),
                ('MEDIA_KPIS', 'Performance KPIs', '📊')
            ]
        },
        {
            'title': 'Financial & Operations',
            'fields': [
                ('PLATFORM_FEE_LOCAL', 'Platform Fees', '💳'),
                ('LOCAL_CURRENCY', 'Currencies Used', '💱'),
                ('COMMENTS', 'Additional Notes', '📝')
            ]
        }
    ]
    
    for section in context_sections:
        # Section title
        ws[f'A{current_row}'] = section['title']
        ws[f'A{current_row}'].font = Font(bold=True, size=10, color="1F4E78")
        current_row += 1
        
        for field, label, emoji in section['fields']:
            if field in combined_df.columns:
                unique_values = combined_df[field].dropna().unique()
                if len(unique_values) > 0:
                    # Label with emoji
                    ws[f'A{current_row}'] = f"{emoji} {label}:"
                    ws[f'A{current_row}'].font = Font(bold=True, size=9)
                    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', indent=1)
                    
                    # Format values based on field type
                    if field == 'PLATFORM_FEE_LOCAL':
                        # Sum up fees and show total
                        total_fees = combined_df[field].sum()
                        value_str = f"£{total_fees:,.2f} total across all platforms"
                    elif field == 'LOCAL_CURRENCY':
                        # Show unique currencies
                        value_str = ', '.join(sorted(set(str(v) for v in unique_values)))
                    elif field == 'CEJ_OBJECTIVES':
                        # Show objectives with counts
                        obj_counts = combined_df[field].value_counts()
                        value_str = ' | '.join([f"{obj}: {count}" for obj, count in obj_counts.items()[:5]])
                    else:
                        # Show unique values with intelligent truncation
                        if len(unique_values) <= 5:
                            value_str = ' | '.join(str(v) for v in unique_values)
                        else:
                            # Show top 5 most common
                            value_counts = combined_df[field].value_counts()
                            top_values = value_counts.head(5)
                            value_str = ' | '.join([f"{v} ({c}x)" for v, c in top_values.items()])
                            if len(unique_values) > 5:
                                value_str += f" ... +{len(unique_values) - 5} more"
                    
                    # Merge cells for value display
                    ws.merge_cells(f'B{current_row}:H{current_row}')
                    ws[f'B{current_row}'] = value_str
                    ws[f'B{current_row}'].font = Font(size=9, color="333333")
                    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
                    
                    # Add light border
                    for col in range(1, 9):
                        ws.cell(row=current_row, column=col).border = Border(
                            bottom=Side(style='dotted', color='CCCCCC')
                        )
                    
                    current_row += 1
        
        current_row += 1
    
    # Section 4: Data Quality Report
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "✅ DATA QUALITY & PROCESSING SUMMARY"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{current_row}'].fill = SUBHEADER_FILL
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Data quality metrics
    quality_metrics = []
    
    # Calculate completeness by key fields
    key_fields = ['BUDGET_LOCAL', 'IMPRESSIONS', 'PLATFORM', 'MARKET', 'CEJ_OBJECTIVES']
    for field in key_fields:
        if field in combined_df.columns:
            completeness = combined_df[field].notna().sum() / len(combined_df) * 100
            quality_metrics.append((field.replace('_', ' ').title(), f"{completeness:.1f}% complete"))
    
    # Processing summary
    quality_metrics.extend([
        ("Total Rows Processed", f"{len(combined_df):,}"),
        ("Data Sources", ', '.join(combined_df['Source_Type'].unique() if 'Source_Type' in combined_df.columns else [])),
        ("Processing Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Mapper Version", "V2.0 - 100% Coverage")
    ])
    
    # Write in two columns
    for i, (metric, value) in enumerate(quality_metrics):
        row = current_row + (i // 2)
        col_offset = (i % 2) * 4
        
        ws[f'{get_column_letter(1 + col_offset)}{row}'] = metric
        ws[f'{get_column_letter(1 + col_offset)}{row}'].font = Font(size=9)
        
        ws.merge_cells(f'{get_column_letter(2 + col_offset)}{row}:{get_column_letter(3 + col_offset)}{row}')
        ws[f'{get_column_letter(2 + col_offset)}{row}'] = value
        ws[f'{get_column_letter(2 + col_offset)}{row}'].font = Font(size=9, bold=True, color="1F4E78")
        ws[f'{get_column_letter(2 + col_offset)}{row}'].alignment = Alignment(horizontal='right')
    
    # Adjust column widths for better visibility
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    logger.info("Enhanced additional context data written successfully")