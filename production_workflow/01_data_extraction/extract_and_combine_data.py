"""
Excel Data Extractor - Core extraction logic for PLANNED and DELIVERED formats

For documentation and troubleshooting, see INDEX.md in project root
"""
from __future__ import annotations

import pandas as pd
import numpy as np
from pathlib import Path
import logging
import sys
from datetime import datetime
import re
import argparse
import os
import textwrap
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows
import traceback
from unidecode import unidecode
import zipfile
from typing import Dict, Tuple, Optional, Any, List
from openpyxl.worksheet.worksheet import Worksheet

# Import error handling utilities
try:
    from error_handler import (
        validate_file_path, validate_dataframe, handle_excel_errors,
        create_error_report, FileProcessingError, ConfigurationError,
        ExtractorError, DataValidationError
    )
except ImportError:
    # Fallback if error_handler is not available
    def validate_file_path(path): return os.path.exists(path)
    def validate_dataframe(df, **kwargs): return {'valid': True, 'errors': []}
    def handle_excel_errors(func): return func
    def create_error_report(errors, warnings): return "\n".join(errors + warnings)
    class FileProcessingError(Exception): pass
    class ConfigurationError(Exception): pass
    class ExtractorError(Exception): pass
    class DataValidationError(Exception): pass

import time
import json # Added import for JSON

# ---------------------------------------------------------------------------
# Optimised worksheet-bounds utility (formerly in the v0.6 wrapper)
# ---------------------------------------------------------------------------
def _determine_effective_bounds(sheet: Worksheet) -> tuple[int, int]:
    """Return (last_row, last_col) that still contain data in *sheet*."""
    max_row, max_col = sheet.max_row, sheet.max_column

    # last non-empty row
    last_row = 0
    for r in range(max_row, 0, -1):
        if any(sheet.cell(row=r, column=c).value not in (None, "") for c in range(1, max_col + 1)):
            last_row = r
            break

    # last non-empty column (only search the rows that have data)
    last_col = 0
    for c in range(max_col, 0, -1):
        if any(sheet.cell(row=r, column=c).value not in (None, "") for r in range(1, last_row + 1)):
            last_col = c
            break

    return last_row, last_col

# Configure logging level from environment variable
LOG_LEVEL = os.environ.get('EXCEL_EXTRACTOR_LOG_LEVEL', 'DEBUG').upper()
VALID_LOG_LEVELS = {'DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'}
if LOG_LEVEL not in VALID_LOG_LEVELS:
    LOG_LEVEL = 'INFO'

# --- Configuration --- 
LOG_FILE = Path("logs/excel_processor.log") # Define log file path globally

# Global CONFIG dictionary and load_config function
CONFIG = {}

def load_config(config_path="config.json", schema_path="config_schema.json"):
    global CONFIG
    
    # First, try to load from the new comprehensive config.json
    config_loaded = False
    
    try:
        # Check parent directory first (where config.json should be)
        parent_config = Path(__file__).parent.parent / config_path
        if parent_config.exists():
            config_path = str(parent_config)
        
        with open(config_path, 'r') as f:
            CONFIG = json.load(f)
        
        # Validate against schema if available
        try:
            parent_schema = Path(__file__).parent.parent / schema_path
            if parent_schema.exists():
                schema_path = str(parent_schema)
            
            if Path(schema_path).exists():
                import jsonschema
                with open(schema_path, 'r') as f:
                    schema = json.load(f)
                jsonschema.validate(CONFIG, schema)
                logging.info(f"Configuration validated against schema successfully")
        except ImportError:
            logging.warning("jsonschema not installed, skipping validation")
        except Exception as e:
            logging.warning(f"Schema validation failed: {e}, continuing with loaded config")
        
        # Map new config structure to expected format
        if 'extraction_config' in CONFIG:
            CONFIG['global_markers'] = CONFIG['extraction_config'].get('markers', {
                'start': ['START'], 'end': ['END']
            })
            # Convert to uppercase keys for backwards compatibility
            if 'start' in CONFIG['global_markers']:
                CONFIG['global_markers']['START'] = CONFIG['global_markers'].pop('start')
            if 'end' in CONFIG['global_markers']:
                CONFIG['global_markers']['END'] = CONFIG['global_markers'].pop('end')
        
        CONFIG.setdefault('sheet_specific_rules', {})
        CONFIG.setdefault('default_sheet_rules', {"ignore_markers_in_columns": [], "custom_markers": {}})
        CONFIG.setdefault('problematic_patterns', [])
        
        logging.info(f"Configuration loaded successfully from {config_path}")
        config_loaded = True
        
    except FileNotFoundError:
        logging.warning(f"Configuration file {config_path} not found. Using default values.")
    except json.JSONDecodeError as e:
        logging.error(f"Error decoding JSON from {config_path}: {e}. Using default values.")
    except Exception as e:
        logging.error(f"Unexpected error loading config: {e}. Using default values.")
    
    # If config not loaded or incomplete, set defaults
    if not config_loaded or 'global_markers' not in CONFIG:
        logging.info("Using default configuration values")
        CONFIG['global_markers'] = {
            'START': [
                'START', '[START]', '*START*', 'START:', 'START.',
                'TABLE START', 'DATA START', 'REPORT START',
                'CAMPAIGN START', 'MEDIA PLAN START'
            ],
            'END': [
                'END', '[END]', '*END*', 'END:', 'END.',
                'TABLE END', 'DATA END', 'REPORT END',
                'CAMPAIGN END', 'MEDIA PLAN END'
            ]
        }
        CONFIG['sheet_specific_rules'] = {}
        CONFIG['default_sheet_rules'] = {"ignore_markers_in_columns": [], "custom_markers": {}}
        CONFIG['problematic_patterns'] = []

    # --------------------------------------------------------------
    # Apply optional overrides from CONFIG to global constants
    # --------------------------------------------------------------
    overrides = {
        'sheet_mapping': 'SHEET_MAPPING',
        'platform_standardization': 'PLATFORM_STANDARDIZATION',
        'column_alternatives': 'COLUMN_ALTERNATIVES',
        'standardized_columns': 'STANDARDIZED_COLUMNS',
        'numeric_columns': 'NUMERIC_COLUMNS',
        'essential_validation_columns': 'ESSENTIAL_DATA_COLUMNS_FOR_VALIDATION',
        'pca_indicators': 'PCA_INDICATORS',
        'rf_metric_map': 'RF_METRIC_MAP',
        'rf_cej_map': 'RF_CEJ_MAP',
        'default_platform_config': 'DEFAULT_PLATFORM_CONFIG',
        'tiktok_market_mapping': 'TIKTOK_MARKET_MAPPING'
    }
    for json_key, global_name in overrides.items():
        if json_key in CONFIG and CONFIG[json_key]:  # only override with non-empty values
            globals()[global_name] = CONFIG[json_key]
            logging.info(f"Applied config override for {global_name} from '{json_key}' key.")

# Define which sheets should be processed
SHEETS_TO_PROCESS = ['DV360', 'META', 'TIKTOK']

# These are the columns we aim to have in the FINAL, COMBINED DataFrame.
# The script will try to map various input headers to these standard names.
# Order here will influence final output order.
FINAL_OUTPUT_COLUMNS = [
    "Source_File", "Source_Sheet",
    "MARKET",
    "BRAND",
    "CAMPAIGN",
    "PLATFORM",
    "CEJ_OBJECTIVES",
    "FORMAT_TYPE",
    "PLACEMENT",
    "AD_UNIT_TYPE",
    "DEVICE",
    "TARGET_AUDIENCE",
    "BUYING_MODEL",
    "START_DATE",
    "END_DATE",
    "WEEKS",
    "LOCAL_CURRENCY",
    "BUDGET_LOCAL",
    "IMPRESSIONS",
    "CLICKS_ACTIONS",
    "VIDEO_VIEWS",
    "FREQUENCY",
    "UNIQUES_REACH",
    "PERCENT_UNIQUES",
    "CPM_LOCAL",
    "CPC_LOCAL",
    "CPV_LOCAL",
    "CTR_PERCENT",
    "VTR_PERCENT",
    "PLATFORM_FEE_LOCAL",
    "PLATFORM_BUDGET_LOCAL",
    "TA_SIZE",
    "MEDIA_KPIS",
    "COMMENTS",
    "CREATIVE_NAME"
]

# Columns considered essential for a data row to be valid after processing.
# Used by the validate_data_types function. This list should be a SUBSET
# of the data-centric fields in FINAL_OUTPUT_COLUMNS.
ESSENTIAL_DATA_COLUMNS_FOR_VALIDATION = [
    "MARKET",
    "CAMPAIGN",
    "PLATFORM",
    "IMPRESSIONS",
    "BUDGET_LOCAL", # Added as essential
    "START_DATE",   # Added as essential
    "END_DATE"      # Added as essential
]

# Columns that should be treated as numeric after mapping.
# This list uses the FINAL_OUTPUT_COLUMN names.
NUMERIC_COLUMNS = [
    "BUDGET_LOCAL", "IMPRESSIONS", "CLICKS_ACTIONS", "VIDEO_VIEWS",
    "FREQUENCY", "UNIQUES_REACH", "PERCENT_UNIQUES", "CPM_LOCAL", "CPC_LOCAL", "CPV_LOCAL",
    "CTR_PERCENT", "VTR_PERCENT", "PLATFORM_FEE_LOCAL", "PLATFORM_BUDGET_LOCAL", "TA_SIZE", "WEEKS"
]

# Standardized column names - this might be redundant if FINAL_OUTPUT_COLUMNS is used for reordering
# For now, let's ensure it includes all potential mapped keys from COLUMN_ALTERNATIVES plus any other fixed ones.
# This will be used by ensure_all_standard_columns.
STANDARDIZED_COLUMNS = list(set([
    "MARKET", "BRAND", "CAMPAIGN", "PLATFORM", "CEJ_OBJECTIVES", "FORMAT_TYPE", 
    "PLACEMENT", "AD_UNIT_TYPE", "DEVICE", "TARGET_AUDIENCE", "BUYING_MODEL",
    "START_DATE", "END_DATE", "WEEKS", "LOCAL_CURRENCY", "BUDGET_LOCAL",
    "IMPRESSIONS", "CLICKS_ACTIONS", "VIDEO_VIEWS", "FREQUENCY", "UNIQUES_REACH", 
    "PERCENT_UNIQUES", "CPM_LOCAL", "CPC_LOCAL", "CPV_LOCAL", "CTR_PERCENT", "VTR_PERCENT",
    "PLATFORM_FEE_LOCAL", "PLATFORM_BUDGET_LOCAL", "TA_SIZE", "MEDIA_KPIS", "COMMENTS", "CREATIVE_NAME",
    # Include R&F specific metrics if they are to be distinct columns before a potential later pivot/merge
    "Campaign Reach (Absl)", "Campaign Freq.", "Awareness Reach Absl" # From R&F tables
]))

# Update the sheet mapping to handle both template types more comprehensively
SHEET_MAPPING = {
    'DV360': ['DV360', 'Display & Video 360', 'Summary - DV360', 'DV360 Summary', 'Display & Video', 'DISPLAY & VIDEO', 'YouTube', 'YOUTUBE'],
    'META': ['META', 'FACEBOOK', 'INSTAGRAM', 'Facebook', 'Instagram', 'Summary - META', 'META Summary', 'Meta', 'Meta Ads', 'FB', 'IG'],
    'TIKTOK': ['TIKTOK', 'TikTok', 'Summary - TIKTOK', 'TikTok Summary', 'Tiktok', 'TikTok Ads']
}

# Enhanced platform standardization mapping with more variations
PLATFORM_STANDARDIZATION = {
    # DV360 variations
    'DV360': 'DV360',
    'DISPLAY & VIDEO 360': 'DV360',
    'DISPLAY & VIDEO': 'DV360',
    'DISPLAY&VIDEO': 'DV360',
    'D&V360': 'DV360',
    'YOUTUBE': 'DV360',
    'YT': 'DV360',
    
    # META variations
    'META': 'META',
    'FACEBOOK': 'META',
    'FB': 'META',
    'INSTAGRAM': 'META',
    'IG': 'META',
    'META ADS': 'META',
    
    # TikTok variations
    'TIKTOK': 'TikTok',
    'TIK TOK': 'TikTok',
    'TIKTOKADS': 'TikTok',
    'TIK-TOK': 'TikTok'
}

# ------------------------------------------------------------------
# Default TikTok market name mapping (overridable via config.json)
# ------------------------------------------------------------------
TIKTOK_MARKET_MAPPING = {
    'UAE': 'UAE',
    'KSA': 'KSA',
    'QAT': 'Qatar',
    'BAH': 'Bahrain',
    'LEB': 'Lebanon',
    'EGY': 'Egypt',
    'KWT': 'Kuwait',
    'OMN': 'Oman'
}

# Default platform configuration - defines platform-specific settings
DEFAULT_PLATFORM_CONFIG = {
    'DV360': {
        'market_col_threshold': 0.3,  # Default threshold for R&F tables in DV360
        'rf_detection': 'metrics_market'  # Method to detect R&F tables
    },
    'META': {
        'market_col_threshold': 0.4,  # Lower threshold for META as it might have more metric columns
        'rf_detection': 'metrics_market'
    },
    'TIKTOK': {
        'market_col_threshold': 0.3,
        'rf_detection': 'metrics_market'
    },
    'default': {
        'market_col_threshold': 0.5,
        'rf_detection': 'metrics_market'
    }
}

def get_default_platform_config(platform=None):
    """
    Returns the default platform configuration settings.
    
    Args:
        platform: Optional platform to get specific configuration for
        
    Returns:
        Dict containing platform-specific settings
    """
    if platform and platform in DEFAULT_PLATFORM_CONFIG:
        return DEFAULT_PLATFORM_CONFIG[platform]
    elif platform:
        # If platform specified but not found, try to match a key
        for key in DEFAULT_PLATFORM_CONFIG:
            if key in platform or platform in key:
                return DEFAULT_PLATFORM_CONFIG[key]
    
    # Return the default configuration if platform is not specified or not found
    return DEFAULT_PLATFORM_CONFIG.get('DEFAULT', {'market_threshold': 0.5})

def standardize_platform_name(platform_value):
    """
    Standardize platform names to ensure consistent naming across different Excel files.
    
    Args:
        platform_value: The platform value to standardize
        
    Returns:
        Standardized platform name
    """
    if platform_value is None:
        return None
        
    # Convert to string for processing
    platform_str = str(platform_value).strip()
    
    # CRITICAL: Don't standardize R&F metrics - they should remain as-is
    if any(keyword in platform_str for keyword in ['Reach', 'Freq', 'reach', 'freq']):
        return platform_value
    
    # Convert to uppercase for comparison
    platform_str_upper = platform_str.upper()
    
    # Check for exact matches first
    if platform_str_upper in PLATFORM_STANDARDIZATION:
        return PLATFORM_STANDARDIZATION[platform_str_upper]
    
    # Check for partial matches
    for source, target in PLATFORM_STANDARDIZATION.items():
        if source in platform_str_upper or platform_str_upper in source:
            return target
            
    # Return original if no match found
    return platform_value

# ---------------------------------------------------------------------------
# Helper: do we *expect* Campaign Reach/Freq rows in this sheet?
# ---------------------------------------------------------------------------
def _rf_expected_for_sheet(sheet_df: pd.DataFrame,
                           file_format: str,
                           normalized_sheet_name: str) -> bool:
    """
    Return True only when Campaign Reach/Freq metrics are LEGITIMATELY
    expected in *this* DataFrame.

    Conditions:
      • workbook was detected as "delivered"
      • and PLATFORM col (if present) already shows at least one Reach/Freq row
        (for safety in mixed-format files).
    """
    if file_format.lower() != "delivered":
        return False
    if "PLATFORM" not in sheet_df.columns:
        return False
    return sheet_df["PLATFORM"].astype(str).str.contains(r"(Reach|Freq)",
                                                         case=False,
                                                         na=False).any()

def clean_text(text):
    """Clean and normalize text for comparison."""
    if text is None:
        return ""
    return str(text).strip().upper().replace(" ", "").replace("&", "").replace("-", "")

def get_normalized_sheet_name(sheet_name: str) -> Optional[str]:
    """
    Get the normalized sheet name based on the SHEET_MAPPING.
    Returns None if the sheet should not be processed.
    """
    if not sheet_name:
        return None
        
    # If the sheet name is exactly "SUMMARY", do not attempt to map it via partial matches
    # to "Summary - DV360" etc., unless "SUMMARY" itself is explicitly listed as a variation.
    # This prevents the main summary tab in planned files from being processed as a platform sheet.
    if sheet_name.strip().upper() == 'SUMMARY':
        # Check if "SUMMARY" itself is a direct variation for any target normalized name
        for normalized, variations in SHEET_MAPPING.items():
            if sheet_name.strip().upper() in [v.strip().upper() for v in variations]:
                return normalized
        # If "SUMMARY" is not an explicit variation for any target, it should not be processed.
        logging.info(f"Sheet '{sheet_name}' identified as a generic summary tab and will be skipped based on current mapping rules.")
        return None

    sheet_name_clean = clean_text(sheet_name)
    
    for normalized, variations in SHEET_MAPPING.items():
        # Try exact match first (case-insensitive)
        if sheet_name.strip().upper() in [v.strip().upper() for v in variations]:
            return normalized
            
        # Try cleaned variant match (more robust to spacing/special chars)
        if any(clean_text(var) == sheet_name_clean for var in variations):
            return normalized
            
        # Try partial match if needed - be careful here to avoid over-matching
        # This was the problematic part for "SUMMARY"
        # We'll keep it but the check above for exact "SUMMARY" should prevent the issue.
        for var in variations:
            cleaned_var = clean_text(var)
            # Ensure partial match is significant, e.g., the variation is a substring of the sheet name,
            # or the sheet name (if short) is a substring of a longer variation.
            if (
                (len(cleaned_var) > 2 and cleaned_var in sheet_name_clean) or \
                (len(sheet_name_clean) > 2 and sheet_name_clean in cleaned_var)
            ):
                # The check for exact "SUMMARY" at the beginning of the function handles the main issue.
                return normalized
            
    return None

# Define marker variations and table identifiers
MARKER_VARIATIONS = {
    'START': [
        'START', '[START]', '*START*', 'START:', 'START.',
        'TABLE START', 'DATA START', 'REPORT START',
        'CAMPAIGN START', 'MEDIA PLAN START'
    ],
    'END': [
        'END', '[END]', '*END*', 'END:', 'END.',
        'TABLE END', 'DATA END', 'REPORT END',
        'CAMPAIGN END', 'MEDIA PLAN END'
    ]
}

# Keywords that might identify a header row for a data table if markers are ambiguous or missing
TABLE_IDENTIFIER_KEYWORDS = [
    'CAMPAIGN', 'PLATFORM', 'IMPRESSIONS', 'CLICKS', 'SPEND', 'BUDGET', 'COST', 
    'START DATE', 'END DATE', 'METRICS', 'MARKET', 'VIEWS', 'OBJECTIVES', 'FORMAT', 
    'NTM (Spend)', 'Campaign Name', 'Metrics/ Market', 'MEDIA COST', 'INVESTMENT',
    'TOTAL SPEND', 'ACTUAL SPEND', 'DELIVERED', 'ACTUAL', 'PERFORMANCE',
    'TOTAL IMPRESSIONS', 'DELIVERED IMPRESSIONS', 'TOTAL CLICKS', 'LINK CLICKS',
    'VIDEO VIEWS', 'VIDEO COMPLETIONS', 'REACH', 'FREQUENCY', 'CPM', 'CPC', 'CPV'
]

# Mapping from various possible header names to standardized column names
COLUMN_ALTERNATIVES = {
    # Target Standard Name: [List of variations]
    "MARKET": ["Market", "Country", "Metrics/ Market", "Metrics / Markets", "Market/Region", "Region", "Market Name"],
    "BRAND": ["Brand", "BRAND", "Brand Name", "Product", "Product Name"],
    "CAMPAIGN": ["Campaign", "Campaign Name", "CAMPAIGN NAME", "Campaign name", "Ad set name", "Adset Name", "Ad Set Name", "Ad Name"],
    "PLATFORM": ["Platform", "Media Platform", "Media Owner", "Publisher", "Platform Name", "Media Placement"],
    "CEJ_OBJECTIVES": ["CEJ", "CEJ Objectives", "Objectives", "Business Objective", "CEJ Objective", "cej", "Objective", "Campaign Objective"],
    "FORMAT_TYPE": ["Format", "Format Type", "Media Format", "Type", "Creative Format", "Ad Format", "FORMAT", "Ad Type"],
    "PLACEMENT": ["Placement", "Platform Placement", "Ad Placement", "Media Placement", "Placement name", "Placement Name"],
    "AD_UNIT_TYPE": ["Ad Unit Type", "Ad Unit", "Unit Type", "Ad Type", "Unit", "Ad Format"],
    "DEVICE": ["Device", "Device Type", "Targeting: Device", "Device Platform", "Device breakdown", "Device Breakdown"],
    "TARGET_AUDIENCE": ["Target Audience", "TA", "Audience", "Target", "Targeting", "Target Group", "Target Demographics", "Target Demographic"],
    "BUYING_MODEL": ["Buying Model", "Buy Type", "Buying Type", "Campaign type", "Campaign Type", "Buying method"],
    "START_DATE": ["Start Date", "Start", "Campaign Start", "Flight Start", "Plan Start", "Start date", "Campaign start date", "Start Time"],
    "END_DATE": ["End Date", "End", "Campaign End", "Flight End", "Plan End", "End date", "Campaign end date", "End Time"],
    "WEEKS": ["# Weeks", "Weeks", "# WEEKS", "Number of Weeks", "Week Count", "Total Weeks", "Week count", "Duration (weeks)"],
    "LOCAL_CURRENCY": ["Local Currency", "Currency", "LOCAL CURRENCY", "Account Currency", "Reporting Currency"],
    "BUDGET_LOCAL": ["Budget [LOCAL]", "Budget Local", "Local Budget", "Media Cost (Local)", "Cost Local", "Budget (Local Currency)", 
                    "Local Budget", "NTM (Spend)", "Spend", "Amount spent", "Amount Spent", "Total Spend", "Cost", "Total Cost", 
                    "Media Cost", "Media Spend", "Total Media Spend", "Investment", "Budget", "Total investment", "Total Investment", 
                    "Actual spend", "Actual Spend", "Actual cost", "Actual Cost", "Delivered Spend", "Delivered Cost"],
    "IMPRESSIONS": ["Impressions", "Imps", "Total Impressions", "Delivered Impressions", "Impression", "Impressions delivered", 
                   "Impressions Delivered", "Imps Delivered", "Total Imps", "Impressions served", "Impressions Served", "Served Impressions"],
    "CLICKS_ACTIONS": ["Clicks / Actions", "Clicks", "Actions", "Total Clicks", "Clicks/Actions", "Total Actions", "Link clicks", 
                      "Link Clicks", "Clicks delivered", "Clicks Delivered", "Total engagements", "Total Engagements", "Engagement"],
    "VIDEO_VIEWS": ["Video Views", "Video Completions", "Views", "Total Video Views", "Total Views", "Video views", "Total video views", 
                   "Video completes", "Video Completes", "Complete Video Views", "Complete views", "2-second continuous video plays"],
    "FREQUENCY": ["Frequency", "Freq", "Avg Frequency", "Average Frequency", "Freq.", "Avg. Frequency", "Frequency (Impr./Reach)", "Average frequency"],
    "UNIQUES_REACH": ["Uniques / Reach", "Uniques Reach", "Reach", "Unique Users", "Unique Reach", "Total Reach", "People reached", 
                     "People Reached", "Reach (unique users)", "Reach (people)", "Post reach", "Total unique users", "Unique Viewers"],
    "PERCENT_UNIQUES": ["% UNIQUES", "% Uniques", "UNIQUES %", "Uniques %", "Reach %", "% REACH", "Percent Uniques", "% Unique Reach", 
                       "Reach Rate (%)", "% People Reached", "Percent reach"],
    "CPM_LOCAL": ["CPM", "CPM Local", "Cost Per Mille", "CPM Buying Rate", "CPM BUYING RATE", "CPM BUYING RATE\n[LOCAL]", 
                 "Cost per 1,000 impressions (CPM)", "CPM (cost per 1,000 impressions)", "Cost per thousand impressions", "Cost per 1000 impressions"],
    "CPC_LOCAL": ["CPC", "CPC Local", "Cost Per Click", "CPC Buying Rate", "CPC BUYING RATE", "CPC BUYING RATE\n[LOCAL]", 
                 "Cost per click (CPC)", "CPC (cost per click)", "Cost per link click", "Cost per Click"],
    "CPV_LOCAL": ["CPV", "CPV Local", "Cost Per View", "CPV Buying Rate", "CPV BUYING RATE", "CPV BUYING RATE\n[LOCAL]", 
                 "Cost per video view", "CPV (cost per view)", "Cost per 2-second continuous video play", "Cost per video play"],
    "CTR_PERCENT": ["CTR", "CTR %", "Click-Through Rate", "CTR%", "Click Through Rate", "Click-through rate", "CTR (%)", 
                   "Link click-through rate", "Clickthrough rate", "LCTR", "Link CTR"],
    "VTR_PERCENT": ["VTR", "VTR %", "View-Through Rate", "VTR%", "View Through Rate", "Video completion rate", "View completion rate", "VCR", 
                   "Video through rate", "Video play rate", "Video view rate", "VTR (%)"],
    "PLATFORM_FEE_LOCAL": ["Platform Fee", "Platform Fee [LOCAL]", "Tech Fee", "Platform Fee Local", "Fee (Local)", "PLATFORM FEE\n[LOCAL]", 
                          "Technology Fee", "Tech Cost", "Platform commission", "Platform Commission", "Platform tech fee"],
    "PLATFORM_BUDGET_LOCAL": ["Platform Budget [LOCAL]", "Platform Budget", "Platform Budget Local", "Budget w/ Fee", "Total Budget", 
                             "Budget with Tech Fee", "Total Cost", "PLATFORM BUDGET\n[LOCAL]", "Budget with fee", "Total budget incl. fee", 
                             "Total cost with fee"],
    "TA_SIZE": ["TA Size", "Audience Size", "Target Size", "Universe Size", "Target audience size", "Potential Reach", "Potential reach", 
               "Audience reach", "Maximum potential reach"],
    "MEDIA_KPIS": ["Media KPIs", "Media KPI", "KPIs", "Campaign KPIs", "Media Key Performance Indicators", "Performance KPIs", 
                  "Campaign performance KPIs", "Media objectives"],
    "COMMENTS": ["Comments", "Notes", "Remarks", "Comment", "Additional notes", "Campaign notes", "Additional information", "Special instructions"],
    "CREATIVE_NAME": ["Creative Name", "Creative", "Ad Name", "Ad Creative", "Creative Description", "Creative Asset", "Creative Title", 
                     "Ad creative", "Creative details", "Creative content"]
}

# PCA/Delivered format content indicators - now global
PCA_INDICATORS = [
    "METRICS/ MARKET", "METRICS / MARKET", "NTM (SPEND)",
    "CAMPAIGN REACH", "AWARENESS REACH", "AD RECALL LIFT",
    "CAMPAIGN FREQ", "PERFORMANCE SUMMARY", "ACTUAL DELIVERY",
    "DELIVERED METRICS", "CAMPAIGN RESULTS", "REACH & FREQUENCY", "REACH AND FREQUENCY",
    "ABSOLUTE LIFT", "RELATIVE LIFT",
    # Add all R&F metrics we're now handling
    "CAMPAIGN REACH (ABSL)", "CAMPAIGN REACH ABSL", 
    "AWARENESS REACH ABSL", 
    "CONSIDERATION REACH ABSL - FOR GNE", "CONSIDERATION REACH ABSL - FOR GNE (COMBINE PROG BUYS)",
    "PURCHASE REACH ABSL", 
    "CAMPAIGN FREQ.", "CAMPAIGN FREQUENCY",
    "AWARENESS FREQ.", "CONSIDERATION FREQ.", "PURCHASE FREQ.",
    "UNIQUE REACH", "UNIQUE USERS", "ESTIMATED AD RECALL LIFT (PEOPLE)"
]

# --- Helper Functions ---
def setup_logging(log_level=None):
    """
    Set up logging configuration with both file and console handlers.
    Logs are saved in the logs directory.
    
    Args:
        log_level: Optional log level string. If not provided, uses:
                  1. EXCEL_EXTRACTOR_LOG_LEVEL environment variable
                  2. Default to 'INFO'
    """
    # Create logs directory if it doesn't exist
    log_dir = Path('logs')
    log_dir.mkdir(parents=True, exist_ok=True)
    
    # Set up log file path (using global constant)
    log_file = LOG_FILE 
    
    # Determine log level from: 1) parameter, 2) env var, 3) default
    if log_level:
        current_log_level_name = log_level.upper()
    else:
        current_log_level_name = os.environ.get('EXCEL_EXTRACTOR_LOG_LEVEL', 'INFO').upper()
    
    if current_log_level_name not in VALID_LOG_LEVELS:
        current_log_level_name = 'INFO'
    current_log_level = getattr(logging, current_log_level_name, logging.INFO)
    
    # Create formatters
    # file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s') 
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s') # Revert to original format
    console_formatter = logging.Formatter('%(message)s') # Revert to original format
    
    # Create handlers
    file_handler = logging.FileHandler(log_file, mode='w') # Overwrite log each run
    file_handler.setFormatter(file_formatter)
    
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(console_formatter)
    # Set console handler level explicitly if needed (e.g., INFO for cleaner console)
    # console_handler.setLevel(logging.INFO)
    
    # Configure root logger
    root_logger = logging.getLogger()
    # Apply level (now defaults to DEBUG)
    root_logger.setLevel(getattr(logging, LOG_LEVEL, logging.DEBUG))
    
    # Remove any existing handlers to prevent duplicate logging
    if root_logger.hasHandlers():
        root_logger.handlers.clear()
    
    # Add handlers
    root_logger.addHandler(file_handler)
    root_logger.addHandler(console_handler)
    
    logging.info(f"Logging initialized with level: {logging.getLevelName(current_log_level)} to {log_file}") # Use INFO for this message

def check_required_libraries() -> bool:
    """Check if required libraries are installed."""
    try:
        import openpyxl
        import pandas
        return True
    except ImportError as e:
        logging.error(f"Missing required library: {e.name}. Please install it using 'pip install {e.name}'.")
        return False

def clean_column_name(header: str) -> str:
    """Clean and standardize column names."""
    if header is None:
        return "Unnamed"
    header_str = str(header)
    # Remove special characters, convert to uppercase, replace spaces with underscores
    cleaned = re.sub(r'[^\w\s-]', '', header_str).strip()
    cleaned = re.sub(r'[\s-]+|-\s*', '_', cleaned)
    return cleaned.upper() if cleaned else "Unnamed"

def _levenshtein_distance(s1: str, s2: str) -> int:
    """
    Calculate the Levenshtein distance between two strings.
    This is used for fuzzy string matching in header detection.
    """
    if len(s1) < len(s2):
        return _levenshtein_distance(s2, s1)
    
    if len(s2) == 0:
        return len(s1)
    
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    
    return previous_row[-1]

def _calculate_string_similarity(str1: str, str2: str) -> float:
    """
    Calculate string similarity using a combination of methods.
    Returns a score between 0 and 1.
    """
    if not str1 or not str2:
        return 0.0
        
    # Convert to lowercase for comparison
    s1 = str1.lower()
    s2 = str2.lower()
    
    # Calculate Levenshtein distance
    distance = _levenshtein_distance(s1, s2)
    max_len = max(len(s1), len(s2))
    if max_len == 0:
        return 0.0
        
    # Convert distance to similarity score (0-1)
    base_score = 1 - (distance / max_len)
    
    # Boost score for common word matches
    words1 = set(s1.split())
    words2 = set(s2.split())
    common_words = words1.intersection(words2)
    
    if common_words:
        word_match_score = len(common_words) / max(len(words1), len(words2))
        # Weight base score more heavily but consider word matches
        return (base_score * 0.7) + (word_match_score * 0.3)
    
    return base_score

def map_header_to_standard(header: str, sheet_name: str) -> Tuple[str, float]:
    """
    Maps a header string to a standardized column name with improved matching.
    Uses the global COLUMN_ALTERNATIVES for mapping rules.
    
    Args:
        header: Raw header string from Excel
        sheet_name: Name of the sheet being processed (used to get config)
    
    Returns:
        Tuple of (mapped column name, confidence score 0.0-1.0)
    """
    if not header or not str(header).strip():
        # Return a default name for empty headers and low confidence
        return "UNNAMED_COLUMN", 0.0
        
    # Ensure header is a string for all operations
    header_str = str(header)
    cleaned_header = clean_column_name(header_str)
    original_header_lower = header_str.strip().lower()
    
    best_match_name = None
    highest_score = 0.0
    
    # Iterate through standard names and their alternatives defined in COLUMN_ALTERNATIVES
    for standard_name, alternatives_list in COLUMN_ALTERNATIVES.items():
        # 1. Direct match with standard name (cleaned or original case-insensitive)
        if cleaned_header == standard_name or original_header_lower == standard_name.lower():
            logging.debug(f"[{sheet_name}] Header '{header_str}' matched standard '{standard_name}' directly.")
            return standard_name, 1.0
            
        # 2. Direct match with alternatives (cleaned or original case-insensitive)
        for alt in alternatives_list:
            # Ensure alt is also treated as a string for robust comparison
            alt_str = str(alt)
            if cleaned_header == clean_column_name(alt_str) or original_header_lower == alt_str.strip().lower():
                logging.debug(f"[{sheet_name}] Header '{header_str}' matched alternative '{alt_str}' for '{standard_name}'.")
                return standard_name, 1.0
                
        # 3. Similarity-based matching against the standard name and its alternatives
        all_forms_to_check = [standard_name] + alternatives_list
        for form_to_check in all_forms_to_check:
            # Ensure form_to_check is string for similarity calculation
            score = _calculate_string_similarity(original_header_lower, str(form_to_check).lower())
            if score > highest_score:
                highest_score = score
                best_match_name = standard_name
                
    # Apply a threshold for similarity-based matches.
    if best_match_name and highest_score >= 0.75: # Adjusted threshold
        logging.info(f"[{sheet_name}] Header '{header_str}' mapped to '{best_match_name}' with similarity {highest_score:.2f}")
        return best_match_name, highest_score
    
    # If no good match, return the cleaned original header with _UNMAPPED and a low score
    unmapped_name = f"{cleaned_header}_UNMAPPED"
    logging.warning(f"[{sheet_name}] Header '{header_str}' (cleaned: {cleaned_header}) could not be mapped. Max similarity score: {highest_score:.2f}. Returning as '{unmapped_name}'.")
    return unmapped_name, 0.1 # Low confidence for unmapped

def find_markers(sheet: Worksheet, sheet_name: str) -> Dict[str, Tuple[int, int]]:
    """
    Find START and END markers in the sheet.
    Returns positions as (row, col) tuples using 0-based indexing.
    
    Args:
        sheet: openpyxl worksheet object
        sheet_name: Name of the sheet being processed (used to get config)
    
    Returns:
        Dictionary with marker positions {marker_type_N: (row, col)}
    """
    markers = {}
    marker_counts = {'START': 0, 'END': 0}
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Get sheet-specific rules from loaded CONFIG
    # Use normalized_name for looking up sheet_specific_rules if sheet_name is raw
    # For simplicity here, assuming sheet_name passed to find_markers is normalized or matches config keys
    current_sheet_rules = CONFIG.get('sheet_specific_rules', {}).get(sheet_name, CONFIG.get('default_sheet_rules', {"ignore_markers_in_columns": [], "custom_markers": {}}))
    ignore_cols_for_sheet = current_sheet_rules.get('ignore_markers_in_columns', [])
    
    # Get global markers (and potentially merge with custom ones later if implemented)
    # Ensure 'global_markers' key exists, providing a default empty dict if not (though load_config handles this)
    markers_to_check = CONFIG.get('global_markers', {})
    # TODO: Add logic for custom_markers from current_sheet_rules.get('custom_markers', {}) if/when implemented

    logging.info(f"Searching for markers in sheet '{sheet_name}' (max_row={max_row}, max_col={max_col}) using config. Ignoring cols: {ignore_cols_for_sheet}")
    
    for r_idx in range(1, max_row + 1):
        for c_idx in range(1, max_col + 1):
            cell_value = sheet.cell(row=r_idx, column=c_idx).value
            current_col_0_indexed = c_idx - 1 # current column being checked

            if cell_value:
                # Check if this column should be ignored for marker detection for this sheet
                if current_col_0_indexed in ignore_cols_for_sheet:
                    # Optional: Add a debug log if a marker is found in an ignored column but skipped
                    # temp_cell_str_for_debug = str(cell_value).strip().upper()
                    # for marker_type_debug, variations_debug in markers_to_check.items():
                    #     if temp_cell_str_for_debug in variations_debug:
                    #         logging.debug(f"Marker '{temp_cell_str_for_debug}' of type '{marker_type_debug}' found in ignored column {current_col_0_indexed} at ({r_idx -1}, {current_col_0_indexed}) and was skipped.")
                    #         break
                    continue # Skip marker check for this cell if column is in ignore list

                cell_str = str(cell_value).strip().upper()
                for marker_type, variations in markers_to_check.items(): # Use markers_to_check from config
                    if cell_str in variations:
                        marker_counts[marker_type] += 1
                        marker_key = f"{marker_type}_{marker_counts[marker_type]}"
                        # Store 0-based index
                        markers[marker_key] = (r_idx - 1, c_idx - 1)
                        logging.debug(f"Found {marker_type} marker '{cell_str}' at position ({r_idx - 1}, {c_idx - 1}) (1-based: {r_idx}, {c_idx})")
                        # Now log marker identification at DEBUG level instead of INFO
                        logging.debug(
                            f"Marker Identification: Sheet '{sheet_name}', Row {r_idx}, Col {c_idx}. "
                            f"Cell value: '{cell_value}' matched marker variation: '{cell_str}' (type: {marker_type}). "
                            f"This marker influences data region boundaries.")
                        break # Found a marker in this cell, move to next cell
    
    # Add summary logging after scanning the sheet
    logging.info(f"Sheet '{sheet_name}': Found {marker_counts['START']} START and {marker_counts['END']} END markers.")
                        
    if not markers:
        logging.info(f"No markers found in sheet '{sheet_name}'. Checking all cells for debugging:")
        try:
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                     logging.debug(f"Cell ({r}, {c}): {sheet.cell(row=r, column=c).value!r}")
        except Exception as e:
             logging.warning(f"Could not read debug cells from '{sheet_name}': {e}")
    return markers

def validate_marker_alignment(sheet: Worksheet, markers: Dict[str, Tuple[int, int]], sheet_name: str, is_delivered_format: bool = False) -> List[Dict[str, int]]:
    """
    Validates alignment of START/END markers and determines potentially MULTIPLE table boundaries.
    Attempts to group markers into logical table regions.
    Returns a list of region boundary dictionaries (0-based index).
    
    Args:
        sheet: The worksheet
        markers: Dictionary of markers with their positions
        sheet_name: Name of the sheet for logging
        is_delivered_format: Whether this is a delivered/PCA format (more lenient validation)
    """
    start_positions = sorted([pos for key, pos in markers.items() if key.startswith('START')], key=lambda x: (x[0], x[1]))
    end_positions = sorted([pos for key, pos in markers.items() if key.startswith('END')], key=lambda x: (x[0], x[1]))
    identified_regions = []
    used_markers = set() # Keep track of markers assigned to a region
    
    # Track anchor attempts and successes for summary logging
    anchor_attempts = 0
    anchor_successes = 0
    failed_reasons = {}  # Track reasons for failures with counts

    if not start_positions or not end_positions:
        logging.info(f"[{sheet_name}] Missing START or END markers for marker validation.")
        return []

    logging.debug(f"[{sheet_name}] Validating {len(start_positions)} START and {len(end_positions)} END markers.")

    # Heuristic: Try to find top-left START markers as anchors for potential tables
    # Sort starts primarily by row, then column
    potential_anchors = sorted([pos for pos in start_positions if pos not in used_markers], key=lambda x: (x[0], x[1]))

    for start_anchor_row, start_anchor_col in potential_anchors:
        anchor_attempts += 1
        
        if (start_anchor_row, start_anchor_col) in used_markers:
            failed_reasons["already_used"] = failed_reasons.get("already_used", 0) + 1
            continue # Skip if already part of another found region
            
        logging.debug(f"[{sheet_name}] Trying anchor START at ({start_anchor_row}, {start_anchor_col})")

        # Find associated markers for this potential table (simplistic approach: nearby markers)
        # 1. Find the horizontal line of STARTs at the anchor row
        top_start_row = start_anchor_row
        top_starts = sorted([ (r,c) for r,c in start_positions if r == top_start_row and (r,c) not in used_markers], key=lambda x: x[1])
        if not top_starts:
             failed_reasons["no_top_starts"] = failed_reasons.get("no_top_starts", 0) + 1
             continue # Should not happen if anchor is valid
        min_top_start_col = top_starts[0][1]
        max_top_start_col = top_starts[-1][1]
        logging.debug(f"  - Top START row {top_start_row}, Cols {min_top_start_col} to {max_top_start_col}")

        # 2. Find the vertical line of STARTs at the min col
        left_start_col = min_top_start_col
        left_starts = sorted([ (r,c) for r,c in start_positions if c == left_start_col and r >= top_start_row and (r,c) not in used_markers ], key=lambda x: x[0])
        if not left_starts:
            failed_reasons["no_left_starts"] = failed_reasons.get("no_left_starts", 0) + 1
            continue
        max_left_start_row = left_starts[-1][0]
        logging.debug(f"  - Left START col {left_start_col}, Rows {top_start_row} to {max_left_start_row}")

        # 3. Look for END markers that form the bottom and right boundaries
        # Find bottom ENDs: Below max_left_start_row, within the column range
        possible_bottom_ends = sorted([ (r,c) for r,c in end_positions 
                                       if r > max_left_start_row 
                                       and c >= min_top_start_col and c <= max_top_start_col
                                       and (r,c) not in used_markers ], key=lambda x: (x[0], x[1]))
        if not possible_bottom_ends:
            logging.debug(f"  - No suitable bottom END markers found below row {max_left_start_row}.")
            failed_reasons["no_bottom_end"] = failed_reasons.get("no_bottom_end", 0) + 1
            continue
        bottom_end_row = possible_bottom_ends[0][0] # Assume first row of ends is the bottom
        bottom_ends = [ (r,c) for r,c in possible_bottom_ends if r == bottom_end_row ]
        min_bottom_end_col = min(c for r,c in bottom_ends)
        max_bottom_end_col = max(c for r,c in bottom_ends)
        logging.debug(f"  - Bottom END row {bottom_end_row}, Cols {min_bottom_end_col} to {max_bottom_end_col}")

        # Find right ENDs: At max_top_start_col, between top_start_row and bottom_end_row
        possible_right_ends = sorted([ (r,c) for r,c in end_positions 
                                      if c == max_top_start_col # Tentative right edge
                                      and r > top_start_row and r <= bottom_end_row
                                      and (r,c) not in used_markers ], key=lambda x: (x[0],x[1]))
        # Adjust right col based on alignment of start/end cols? For now, use max start col.
        # A more robust check would ensure alignment with max_bottom_end_col if possible
        right_end_col = max_top_start_col 
        # A more robust check would ensure alignment with max_bottom_end_col if possible
        logging.debug(f"  - Right END col {right_end_col} (estimated), Rows {top_start_row + 1} to {bottom_end_row}")


        # Basic validation of the formed box
        # Check if the identified boundaries seem coherent
        # For delivered format, be more permissive with column alignment due to typical tables structure
        columns_aligned = min_top_start_col == min_bottom_end_col and max_top_start_col == max_bottom_end_col
        
        # Special handling for DELIVERED (PCA) format files
        if not columns_aligned and is_delivered_format:
            # Check if there's significant overlap in the column range
            top_cols = set(range(min_top_start_col, max_top_start_col + 1))
            bottom_cols = set(range(min_bottom_end_col, max_bottom_end_col + 1))
            overlap = top_cols.intersection(bottom_cols)
            
            # For delivered format, accept if there's any overlap at all
            columns_aligned = len(overlap) > 0
            
            # Also accept if columns appear to form an R&F table (this is a common structure in delivered files)
            # In R&F tables, the first column often contains metric names and other columns are markets
            first_cell_value = str(sheet.cell(top_start_row + 1, min_top_start_col + 1).value or '').strip().upper()
            rf_table_indicator = any(kw in first_cell_value for kw in ["METRICS", "MARKET"])
            
            if rf_table_indicator:
                logging.debug(f"[{sheet_name}] R&F table structure detected at ({top_start_row+1}, {min_top_start_col+1}): '{first_cell_value}'")
                columns_aligned = True
            
            if columns_aligned:
                logging.debug(f"[{sheet_name}] Column alignment relaxed for delivered format: top ({min_top_start_col}-{max_top_start_col}) bottom ({min_bottom_end_col}-{max_bottom_end_col})")
        
        if columns_aligned: 
            logging.debug(f"[{sheet_name}] Marker group column alignment check passed for anchor ({start_anchor_row},{start_anchor_col}).")
            
            # For Planned template, header is BELOW the top START marker row.
            # Top START markers are at top_start_row (0-based).
            # Header is at top_start_row + 1.
            # Data starts at top_start_row + 2.
            header_row_0based = top_start_row + 1 
            data_start_row_0based = header_row_0based + 1 # which is top_start_row + 2
            
            # Data ends one row ABOVE the bottom END marker row.
            # SPECIAL HANDLING for DELIVERED format: scan for actual data end instead of using first END marker
            if is_delivered_format and data_start_row_0based >= 12:  # Media section starts around row 13 (0-based 12)
                # For DELIVERED media sections, scan for actual end of data
                actual_data_end = data_start_row_0based
                for scan_row in range(data_start_row_0based, min(sheet.max_row, data_start_row_0based + 30)):
                    # Check if this row has market data in column 2 (0-based column 1)
                    market_cell = sheet.cell(scan_row + 1, 2).value  # Convert to 1-based for openpyxl
                    if market_cell and str(market_cell).strip() and str(market_cell).strip() != 'END':
                        actual_data_end = scan_row
                logging.debug(f"[{sheet_name}] DELIVERED media section: Extended data end from row {bottom_end_row} to row {actual_data_end + 1}")
                data_end_row_0based = actual_data_end
            else:
                data_end_row_0based = bottom_end_row - 1 
            data_start_col_0based = left_start_col # Uses the identified left-most start column
            data_end_col_0based = right_end_col   # Uses the identified right-most start column (aligned with top)

            # Additional check: ensure header_row_0based is not beyond sheet limits
            if header_row_0based >= sheet.max_row:
                logging.warning(f"[{sheet_name}] Calculated header row {header_row_0based + 1} is out of sheet bounds (max_row: {sheet.max_row}). Skipping this marker group.")
                failed_reasons["header_out_of_bounds"] = failed_reasons.get("header_out_of_bounds", 0) + 1
                continue

            # SPECIAL HANDLING for PLANNED format: adjust data range based on actual structure
            if not is_delivered_format:  # PLANNED format
                # For PLANNED files, find the data range based on START markers in column A (0-based col 0)
                # and END markers in the rightmost column where data ends
                actual_start_row = None
                actual_end_row = None
                
                # First, find the range of rows that have START in column A with actual market data
                for scan_row in range(max(0, header_row_0based - 5), min(sheet.max_row, header_row_0based + 60)):
                    cell_a = sheet.cell(scan_row + 1, 1).value  # Column A (0-based column 0)
                    cell_b = sheet.cell(scan_row + 1, 2).value  # Column B - should contain market name
                    
                    # Look for rows with START in column A and a valid market name in column B
                    if (cell_a == 'START' and cell_b and 
                        str(cell_b).strip() and 
                        str(cell_b).strip().upper() not in ['MARKET', 'START', 'END']):
                        
                        if actual_start_row is None:
                            actual_start_row = scan_row
                        actual_end_row = scan_row
                        
                        # Log each market found for debugging
                        logging.debug(f"[{sheet_name}] PLANNED: Found market data row {scan_row + 1}: '{cell_b}'")
                
                if actual_start_row is not None and actual_end_row is not None:
                    # Find header row (should be 1 row before first data row)
                    potential_header_row = actual_start_row - 1
                    if potential_header_row >= 0:
                        header_cell = sheet.cell(potential_header_row + 1, 2).value
                        if header_cell and str(header_cell).strip().upper() == 'MARKET':
                            header_row_0based = potential_header_row
                            logging.debug(f"[{sheet_name}] PLANNED: Found header row {header_row_0based + 1}")
                    
                    data_start_row_0based = actual_start_row
                    data_end_row_0based = actual_end_row
                    
                    logging.info(f"[{sheet_name}] PLANNED format: Found data range rows {actual_start_row + 1}-{actual_end_row + 1} (header: {header_row_0based + 1})")
                    logging.info(f"[{sheet_name}] PLANNED format: Total data rows: {actual_end_row - actual_start_row + 1}")
                else:
                    logging.warning(f"[{sheet_name}] PLANNED format: No valid data range found with START markers")
            
            # For delivered format, allow empty data regions (they will be validated later)
            min_data_rows = -1 if is_delivered_format else 0
            if data_end_row_0based - data_start_row_0based >= min_data_rows and data_end_col_0based >= data_start_col_0based:
                region = {
                    'header_row': header_row_0based + 1,    # Convert to 1-based
                    'start_row': data_start_row_0based + 1, # Rename and convert to 1-based
                    'end_row': data_end_row_0based + 1,     # Rename and convert to 1-based
                    'start_col': data_start_col_0based + 1, # Rename and convert to 1-based
                    'end_col': data_end_col_0based + 1,     # Rename and convert to 1-based
                    'detection_method': 'markers',          # Add detection method
                    'is_delivered_format': is_delivered_format  # Flag for format awareness
                }
                
                # Extra check for R&F tables in delivered format
                if is_delivered_format:
                    header_cell_value = str(sheet.cell(header_row_0based + 1, data_start_col_0based + 1).value or '').strip().upper()
                    if any(kw in header_cell_value for kw in ["METRICS", "MARKET"]):
                        region['rf_table_candidate'] = True
                        logging.debug(f"[{sheet_name}] Potential R&F table detected at header ({header_row_0based + 1}, {data_start_col_0based + 1}): '{header_cell_value}'")
                
                identified_regions.append(region)
                anchor_successes += 1
                logging.info(f"[{sheet_name}] Valid marker region identified (1-based): Header R{header_row_0based + 1}, Data R{data_start_row_0based + 1}:{data_end_row_0based + 1}, C{data_start_col_0based + 1}:{data_end_col_0based + 1}")
                logging.debug(f"[{sheet_name}] Exact 0-based coordinates for this region: header_row={header_row_0based}, data_start_row={data_start_row_0based}, data_end_row={data_end_row_0based}, data_start_col={data_start_col_0based}, data_end_col={data_end_col_0based}")

                # Mark used markers for this region
                markers_in_region = top_starts + left_starts + bottom_ends + possible_right_ends
                for r, c in markers_in_region:
                    used_markers.add((r, c))
            else:
                 logging.debug(f"[{sheet_name}] Skipping potential marker region anchored at ({start_anchor_row}, {start_anchor_col}) because calculated data region is invalid (Data R{data_start_row_0based+1}:{data_end_row_0based+1}, C{data_start_col_0based+1}:{data_end_col_0based+1}).")
                 failed_reasons["invalid_data_region"] = failed_reasons.get("invalid_data_region", 0) + 1
        else:
            logging.debug(f"[{sheet_name}] Skipping potential marker region anchored at ({start_anchor_row}, {start_anchor_col}) due to column mismatch between top ({min_top_start_col}-{max_top_start_col}) and bottom ({min_bottom_end_col}-{max_bottom_end_col}).")
            failed_reasons["column_mismatch"] = failed_reasons.get("column_mismatch", 0) + 1

    # Add summary logging for anchor attempts
    logging.info(f"[{sheet_name}] Tried {anchor_attempts} anchors, {anchor_successes} succeeded.")
    
    # Only log detailed failure reasons if there were failures
    if anchor_attempts > anchor_successes and logging.getLogger().isEnabledFor(logging.DEBUG):
        failure_summary = ", ".join([f"{reason}: {count}" for reason, count in failed_reasons.items()])
        logging.debug(f"[{sheet_name}] Anchor failure reasons: {failure_summary}")

    if not identified_regions:
         logging.info(f"[{sheet_name}] Could not identify any valid marker regions after grouping.")
         
    return identified_regions # Return list of 0-based region dicts

def _find_regions_with_identifiers(sheet: Worksheet, sheet_name: str) -> List[Dict[str, int]]:
    """
    Find table regions primarily using TABLE_IDENTIFIER_KEYWORDS in potential header rows.
    Returns a list of region dictionaries (1-based index).
    """
    regions = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    used_header_rows = set()  # Track rows identified as headers to avoid overlap

    logging.info(f"[{sheet_name}] Attempting identifier-based region detection.")

    # Define a helper to check if a row looks like a header
    def is_potential_header(row_idx):
        if row_idx > max_row:
            return False
        try:
            # Check ALL columns to ensure we don't miss any data
            cell_values = [str(sheet.cell(row=row_idx, column=c).value or '').strip().upper() for c in range(1, max_col + 1)]
            
            # Count identifiers and analyze cell patterns
            identifier_count = 0
            numeric_count = 0
            text_count = 0
            non_empty_count = 0
            
            for cell_val in cell_values:
                if not cell_val:
                    continue
                    
                non_empty_count += 1
                
                # Check for identifiers
                if any(identifier in cell_val for identifier in TABLE_IDENTIFIER_KEYWORDS):
                    identifier_count += 1
                
                # Check if numeric (allowing for common numeric formats)
                try:
                    # Remove common numeric formatting characters
                    clean_val = cell_val.replace(',', '').replace('$', '').replace('%', '').replace('-', '')
                    float(clean_val)
                    numeric_count += 1
                except ValueError:
                    text_count += 1
            
            if non_empty_count == 0:
                return False
                
            # Calculate ratios
            text_ratio = text_count / non_empty_count if non_empty_count > 0 else 0
            numeric_ratio = numeric_count / non_empty_count if non_empty_count > 0 else 0
            
            # A row is a potential header if it meets ANY of these criteria:
            # 1. Has multiple identifier keywords
            has_multiple_identifiers = identifier_count >= 2
            
            # 2. Has a good mix of text cells (headers are usually text)
            has_good_text_ratio = text_ratio >= 0.6 and non_empty_count >= 3
            
            # 3. Has at least one identifier and mostly text cells
            has_identifier_and_text = identifier_count >= 1 and text_ratio >= 0.7 and non_empty_count >= 3
            
            # 4. Has very high text ratio (almost all cells are text)
            is_mostly_text = text_ratio >= 0.8 and non_empty_count >= 3
            
            # 5. Check if this row has significantly more text cells than numeric cells
            # (data rows often have more numeric cells)
            has_text_dominance = text_count > numeric_count * 2 and non_empty_count >= 3
            
            # Reject if mostly numeric - likely a data row
            if numeric_ratio > 0.7:
                return False
            
            return (has_multiple_identifiers or 
                   has_good_text_ratio or 
                   has_identifier_and_text or 
                   is_mostly_text or 
                   has_text_dominance)
        except Exception as e:
            logging.warning(f"[{sheet_name}] Error checking potential header at row {row_idx}: {e}")
            return False

    # Define a helper to find the end of a data block starting below a header
    def find_data_end_row(start_data_row, start_col, end_col):
        current_row = start_data_row
        consecutive_blank_rows = 0
        max_blank_rows = 3  # Number of consecutive blank rows to consider end of data
        last_data_row = start_data_row - 1 # Initialize to before data actually starts
        
        # Keywords that indicate end of data or a summary section
        SUMMARY_END_KEYWORDS = ["TOTAL", "SUBTOTAL", "SUB TOTAL", "GRAND TOTAL", "SUM", "AVERAGE", "AVG"]
        ALL_END_MARKERS = [m.upper() for m in MARKER_VARIATIONS['END']]
        
        while current_row <= max_row:
            try:
                # Check ALL columns to ensure we don't miss any data
                cell_values = [str(sheet.cell(row=current_row, column=c).value or '').strip().upper() for c in range(1, max_col + 1)]
                
                # Check for blank rows
                is_blank_row = all(cell_val == '' for cell_val in cell_values)
                
                # Check for summary keywords or end markers
                first_cell_in_row_str = cell_values[0] if cell_values else ''
                row_has_end_marker = any(marker in first_cell_in_row_str for marker in ALL_END_MARKERS)
                looks_like_next_header = is_potential_header(current_row)
                is_summary_keyword_row = any(keyword in first_cell_in_row_str for keyword in SUMMARY_END_KEYWORDS)
                
                if not is_blank_row and not looks_like_next_header and not is_summary_keyword_row and not row_has_end_marker:
                    last_data_row = current_row
                    consecutive_blank_rows = 0
                elif is_blank_row:
                    consecutive_blank_rows += 1
                else:
                    consecutive_blank_rows = 0 
                    if looks_like_next_header:
                        logging.debug(f"[{sheet_name}] End data (R{start_data_row}-?) at R{current_row-1}. Reason: Potential next header at R{current_row}.")
                        return last_data_row # Data ended on the previous valid row
                    elif is_summary_keyword_row:
                        logging.debug(f"[{sheet_name}] End data (R{start_data_row}-?) at R{current_row-1}. Reason: Summary keyword '{first_cell_in_row_str}' in R{current_row}.")
                        return last_data_row 
                    elif row_has_end_marker:
                        logging.debug(f"[{sheet_name}] End data (R{start_data_row}-?) at R{current_row-1}. Reason: END marker in R{current_row}.")
                        return last_data_row 
                
                if consecutive_blank_rows >= max_blank_rows:
                    logging.debug(f"[{sheet_name}] End data (R{start_data_row}-?) at R{current_row - consecutive_blank_rows}. Reason: {consecutive_blank_rows} blank rows.")
                    return last_data_row

            except Exception as e_scan:
                logging.warning(f"[{sheet_name}] Error during end row detection for data starting R{start_data_row} at actual row {current_row}: {e_scan}")
                return last_data_row 
            current_row += 1

        logging.debug(f"[{sheet_name}] End data (R{start_data_row}-?) at R{max_row} (end of sheet). Effective last data row: {last_data_row}")
        return last_data_row
    
    # Iterate through rows to find potential headers
    # Process all rows to ensure no data is missed
    for r_idx in range(1, max_row + 1):
        if r_idx in used_header_rows:
            continue

        if is_potential_header(r_idx):
            logging.debug(f"[{sheet_name}] Potential header identified at row {r_idx}.")
            header_row_actual = r_idx
            
            # Find the column range for this header
            header_values = []
            for c_idx in range(1, max_col + 1):
                cell_value = sheet.cell(row=header_row_actual, column=c_idx).value
                header_values.append(cell_value)
            
            # Determine the first and last non-empty columns
            first_col_index = -1
            last_col_index = -1
            
            for c_idx, val in enumerate(header_values):
                if val is not None and str(val).strip():
                    if first_col_index == -1:
                        first_col_index = c_idx
                    last_col_index = c_idx
            
            if first_col_index == -1:
                logging.debug(f"[{sheet_name}] Skipping potential header at row {header_row_actual}: No non-empty cells found.")
                continue
            
            # Apply more aggressive column range determination
            # Look for gaps in header columns - if a gap is more than 3 empty cells, consider it a boundary
            col_groups = []
            current_group = []
            gap_count = 0
            max_gap = 3  # Maximum number of empty columns to consider within the same region
            
            for c_idx, val in enumerate(header_values):
                if val is not None and str(val).strip():
                    # Non-empty cell
                    current_group.append(c_idx)
                    gap_count = 0
                else:
                    # Empty cell
                    gap_count += 1
                    if gap_count > max_gap and current_group:
                        # Gap exceeded threshold, finalize current group
                        if len(current_group) >= 3:  # Require at least 3 header cells to form a group
                            col_groups.append((current_group[0], current_group[-1]))
                        current_group = []
                        gap_count = 0
            
            # Add the last group if it exists
            if current_group and len(current_group) >= 3:
                col_groups.append((current_group[0], current_group[-1]))
            
            # If we found column groups, use them to define regions
            # Otherwise fall back to the simple first/last non-empty determination
            if col_groups:
                for group_start, group_end in col_groups:
                    start_col = group_start + 1  # Convert to 1-based
                    end_col = group_end + 1      # Convert to 1-based
                    
                    # Mark only the header row as used
                    used_header_rows.add(header_row_actual)
                    
                    # Find data rows
                    start_row = header_row_actual + 1
                    end_row = find_data_end_row(start_row, start_col, end_col)
                    
                    if start_row > end_row:
                        logging.warning(f"[{sheet_name}] Skipping region for header group at row {header_row_actual}: No data rows found below it (calculated end_row: {end_row}).")
                        continue
                    
                    region = {
                        "start_row": start_row,
                        "end_row": end_row,
                        "start_col": start_col,
                        "end_col": end_col,
                        "header_row": header_row_actual, 
                        "detection_method": "identifier_group"
                    }
                    regions.append(region)
                    logging.info(f"[{sheet_name}] Identifier group-based region added: Header R{header_row_actual}, Data R{start_row}:R{end_row}, C{start_col}:C{end_col}")
            else:
                # Fall back to simple first/last non-empty approach
                start_col = first_col_index + 1  # 1-based
                end_col = last_col_index + 1     # 1-based
                
                # Mark only the header row as used
                used_header_rows.add(header_row_actual)
                
                start_row = header_row_actual + 1
                end_row = find_data_end_row(start_row, start_col, end_col)

                if start_row > end_row:
                    logging.warning(f"[{sheet_name}] Skipping region for header at row {header_row_actual}: No data rows found below it (calculated end_row: {end_row}).")
                    continue 

                region = {
                    "start_row": start_row,
                    "end_row": end_row,
                    "start_col": start_col,
                    "end_col": end_col,
                    "header_row": header_row_actual, 
                    "detection_method": "identifier"
                }
                regions.append(region)
                logging.info(f"[{sheet_name}] Identifier-based region added: Header R{header_row_actual}, Data R{start_row}:R{end_row}, C{start_col}:C{end_col}")

    if not regions:
        logging.warning(f"[{sheet_name}] No regions found using identifier-based detection.")

    return regions

def regions_overlap(region1: Dict[str, int], region2: Dict[str, int], overlap_threshold: float = 0.5) -> bool:
    """
    Check if two table regions overlap significantly.

    Args:
        region1: First region dictionary with start_row, end_row, start_col, end_col
        region2: Second region dictionary with same keys
        overlap_threshold: Minimum overlap ratio to consider regions as overlapping

    Returns:
        bool: True if regions overlap significantly
    """
    # Calculate row and column ranges
    rows1 = set(range(region1['start_row'], region1['end_row'] + 1))
    cols1 = set(range(region1['start_col'], region1['end_col'] + 1))
    rows2 = set(range(region2['start_row'], region2['end_row'] + 1))
    cols2 = set(range(region2['start_col'], region2['end_col'] + 1))
    
    # Calculate overlaps
    row_overlap = len(rows1.intersection(rows2))
    col_overlap = len(cols1.intersection(cols2))
    
    # Calculate overlap ratios
    row_ratio = row_overlap / min(len(rows1), len(rows2))
    col_ratio = col_overlap / min(len(cols1), len(cols2))
    
    # Consider regions overlapping if both row and column overlap ratios exceed threshold
    return row_ratio > overlap_threshold and col_ratio > overlap_threshold

def find_table_regions(sheet: Worksheet, sheet_name: str, source_workbook_path: Optional[str] = None, effective_max_row: Optional[int] = None, effective_max_col: Optional[int] = None) -> List[Dict[str, int]]:
    """
    Find table regions in a worksheet using multiple detection methods.
    Enhanced to better handle delivered template layouts.
    Can now accept effective_max_row and effective_max_col to limit scanning range.
    """
    log_prefix = f"[{sheet_name}]"
    regions = []
    
    # Use effective bounds if provided, otherwise use sheet's actual max_row/max_column
    max_row = effective_max_row if effective_max_row is not None else sheet.max_row
    max_col = effective_max_col if effective_max_col is not None else sheet.max_column
    
    # Check if this is a delivered/PCA file format based on workbook path or content
    is_delivered_format = False
    if source_workbook_path:
        is_delivered_format = ('pca' in Path(source_workbook_path).name.lower() or 
                              'delivered' in Path(source_workbook_path).name.lower() or 
                              'post' in Path(source_workbook_path).name.lower())
    
    # If path-based check isn't definitive, check for PCA content indicators
    if not is_delivered_format and sheet_name in ['DV360', 'META', 'TIKTOK']:
        # Check first 20 rows for PCA indicators
        for row in range(1, min(21, max_row + 1)):
            for col in range(1, min(10, max_col + 1)):
                cell_value = str(sheet.cell(row, col).value or '').strip().upper()
                if any(indicator in cell_value for indicator in PCA_INDICATORS):
                    is_delivered_format = True
                    logging.info(f"{log_prefix} Detected delivered/PCA format based on content indicator '{cell_value}'")
                    break
            if is_delivered_format:
                break
    
    logging.info(f"{log_prefix} File format detection: {'DELIVERED/PCA' if is_delivered_format else 'PLANNED/STANDARD'}")
    
    # First try marker-based detection
    marker_regions = validate_marker_alignment(sheet, find_markers(sheet, sheet_name), sheet_name, is_delivered_format)
    
    if marker_regions:
        # If markers are found and define valid regions, use ONLY these.
        regions.extend(marker_regions)
        logging.info(f"{log_prefix} Found {len(marker_regions)} regions using START/END markers. These will be used exclusively for this sheet.")
    else:
        # If no valid marker regions found, then try identifier-based detection
        logging.info(f"{log_prefix} No valid START/END marker regions found. Falling back to identifier-based detection.")
        identifier_regions = _find_regions_with_identifiers(sheet, sheet_name)
        if identifier_regions:
            # Assign detection_method, as _find_regions_with_identifiers might produce different types
            # This is already handled within _find_regions_with_identifiers which sets 'identifier_group' or 'identifier'
            regions.extend(identifier_regions)
            logging.info(f"{log_prefix} Found {len(identifier_regions)} regions using identifiers.")
        
        # If still no regions, try scanning for metric columns as a last resort
        # This part might be too aggressive or lead to incorrect regions if not careful
        # For now, let's keep it but be mindful of its potential to find sparse/incorrect tables
        if not regions:
            logging.info(f"{log_prefix} No regions found by markers or identifiers. Falling back to metric column scan.")
            # Scan each row for potential metric headers
            for row in range(1, max_row + 1):
                metric_cols = []
                for col in range(1, max_col + 1):
                    cell_value = str(sheet.cell(row, col).value or '').strip().upper()
                    if any(keyword in cell_value for keyword in TABLE_IDENTIFIER_KEYWORDS):
                        metric_cols.append(col)
                
                if len(metric_cols) >= 2:  # Require at least 2 metric columns
                    # Found potential header row
                    start_col = max(1, min(metric_cols) - 2)  # Include up to 2 columns before first metric
                    end_col = min(max_col, max(metric_cols) + 2)  # Include up to 2 columns after last metric
                    
                    # Find data rows
                    start_row = row + 1
                    end_row = start_row
                    empty_row_count = 0
                    for r in range(start_row, max_row + 1):
                        row_cells = [sheet.cell(r, c).value for c in range(start_col, end_col + 1)]
                        if all(cell is None or str(cell).strip() == '' for cell in row_cells):
                            empty_row_count += 1
                            if empty_row_count >= 3:
                                break
        else:
                        empty_row_count = 0
                        end_row = r
        
        if end_row > start_row:  # Ensure we found some data rows
                    regions.append({
                        'start_row': start_row,
                        'end_row': end_row,
                        'start_col': start_col,
                        'end_col': end_col,
                        'header_row': row,
                        'detection_method': 'metrics'
                    })

    # Merge overlapping regions
    final_regions = []
    for region_to_add in regions: # Renamed from region to avoid confusion with existing_region's properties
        overlap_found_and_merged = False
        for i, existing_region in enumerate(final_regions):
            if regions_overlap(region_to_add, existing_region):
                # Check if one is an RF candidate and the other is not, or if specific conditions apply
                is_new_rf = region_to_add.get('rf_table_candidate', False)
                is_existing_rf = existing_region.get('rf_table_candidate', False)

                # Priority: If one is a clear R&F table (e.g., header starts with METRICS/MARKET) and the other is not,
                # it might be better not to merge them, or merge very carefully.
                # For now, let's refine the header selection upon merging.

                # Do NOT merge if one region is clearly R&F and the other is not – keep them distinct
                if (is_new_rf and not is_existing_rf) or (is_existing_rf and not is_new_rf):
                    continue  # treat as separate table regions despite overlap

                # Otherwise perform standard merge
                merged_start_row = min(existing_region['start_row'], region_to_add['start_row'])
                merged_end_row = max(existing_region['end_row'], region_to_add['end_row'])
                merged_start_col = min(existing_region['start_col'], region_to_add['start_col'])
                merged_end_col = max(existing_region['end_col'], region_to_add['end_col'])
                
                # Header selection logic: prefer R&F candidate header if distinct
                chosen_header_row = existing_region['header_row'] # Default to existing
                
                if is_new_rf and not is_existing_rf:
                    chosen_header_row = region_to_add['header_row']
                elif is_new_rf and is_existing_rf:
                    # Both are R&F, choose based on some heuristic (e.g., earlier row, or more specific pattern)
                    # For now, let's prefer the one that starts earlier (lower header_row index)
                    if region_to_add['header_row'] < existing_region['header_row']:
                        chosen_header_row = region_to_add['header_row']
                elif not is_new_rf and is_existing_rf:
                    chosen_header_row = existing_region['header_row'] # Keep existing R&F header
                else: # Neither are R&F candidates, or both are general, use metric count
                    header_cells_new = [str(sheet.cell(region_to_add['header_row'], c).value or '').strip().upper() 
                                      for c in range(region_to_add['start_col'], region_to_add['end_col'] + 1)]
                    header_cells_existing = [str(sheet.cell(existing_region['header_row'], c).value or '').strip().upper() 
                                       for c in range(existing_region['start_col'], existing_region['end_col'] + 1)]
                    metrics_new = sum(1 for cell in header_cells_new if any(keyword in cell for keyword in TABLE_IDENTIFIER_KEYWORDS))
                    metrics_existing = sum(1 for cell in header_cells_existing if any(keyword in cell for keyword in TABLE_IDENTIFIER_KEYWORDS))
                    if metrics_new > metrics_existing:
                        chosen_header_row = region_to_add['header_row']
                
                # Update the existing_region with merged boundaries and chosen header
                final_regions[i]['start_row'] = merged_start_row
                final_regions[i]['end_row'] = merged_end_row
                final_regions[i]['start_col'] = merged_start_col
                final_regions[i]['end_col'] = merged_end_col
                final_regions[i]['header_row'] = chosen_header_row
                # Ensure rf_table_candidate flag is OR-ed if either was an R&F candidate
                final_regions[i]['rf_table_candidate'] = is_new_rf or is_existing_rf
                # Update detection method if one was more specific (e.g., prefer 'markers' over 'identifier')
                if region_to_add.get('detection_method') == 'markers' and existing_region.get('detection_method') != 'markers':
                    final_regions[i]['detection_method'] = 'markers'
                
                overlap_found_and_merged = True
                logging.debug(f"{log_prefix} Merged overlapping region. New bounds for existing region {i}: R{merged_start_row}-{merged_end_row}, C{merged_start_col}-{merged_end_col}, H{chosen_header_row}")
                break
        if not overlap_found_and_merged:
            final_regions.append(region_to_add)
                
    # Add region indices
    for idx, region in enumerate(final_regions):
        region['region_index'] = idx + 1
    
    # Validate and adjust regions
    validated_regions = []
    for region in final_regions:
        # Ensure minimum data rows - be more lenient for delivered format
        min_data_rows = 0 if is_delivered_format else 1
        if region['end_row'] - region['start_row'] < min_data_rows:
            logging.warning(f"{log_prefix} Skipping region {region['region_index']}: insufficient data rows")
            continue
            
        # Validate header row
        header_row = region['header_row']
        header_cells = [str(sheet.cell(header_row, c).value or '').strip().upper() 
                       for c in range(region['start_col'], region['end_col'] + 1)]
        
        # Count metric columns in header
        metric_count = sum(1 for cell in header_cells 
                         if any(keyword in cell for keyword in TABLE_IDENTIFIER_KEYWORDS))
        
        # For delivered format, be more lenient with metric column requirements
        min_metric_columns = 1 if is_delivered_format else 2
        if metric_count < min_metric_columns:
            logging.warning(f"{log_prefix} Skipping region {region['region_index']}: insufficient metric columns (found {metric_count}, need {min_metric_columns})")
            continue
            
        # Look for better header row within 3 rows
        for test_row in range(max(1, header_row - 3), min(header_row + 4, region['start_row'])):
            test_cells = [str(sheet.cell(test_row, c).value or '').strip().upper() 
                         for c in range(region['start_col'], region['end_col'] + 1)]
            test_metric_count = sum(1 for cell in test_cells 
                                  if any(keyword in cell for keyword in TABLE_IDENTIFIER_KEYWORDS))
            if test_metric_count > metric_count:
                region['header_row'] = test_row
                logging.info(f"{log_prefix} Found better header row at {test_row}")
                break
        
        # Adjust data start row if needed
        if region['start_row'] <= region['header_row']:
            region['start_row'] = region['header_row'] + 1
            
        # Find actual data end by scanning for empty rows
        empty_row_threshold = 3
        empty_row_count = 0
        actual_end_row = region['start_row']
        
        for r in range(region['start_row'], region['end_row'] + 1):
            row_cells = [sheet.cell(r, c).value for c in range(region['start_col'], region['end_col'] + 1)]
            if all(cell is None or str(cell).strip() == '' for cell in row_cells):
                empty_row_count += 1
                if empty_row_count >= empty_row_threshold:
                    break
            else:
                empty_row_count = 0
                actual_end_row = r
        
        region['end_row'] = actual_end_row
        region['is_delivered_format'] = is_delivered_format  # Add flag to region for later processing
        validated_regions.append(region)
    
    if not validated_regions:
        logging.warning(f"{log_prefix} No valid table regions found in sheet")
    else:
        logging.info(f"{log_prefix} Found {len(validated_regions)} valid table regions")
        
    return validated_regions

def map_to_standard_column(header_str: str, column_alternatives: dict) -> str:
    """
    Maps a header string to the standardized column name from
    the provided alternatives dictionary.
    
    Args:
        header_str: The header string to map
        column_alternatives: Dictionary of standardized names with alternative versions
        
    Returns:
        The standardized column name, or None if no match is found
    """
    if not header_str:
        return None
        
    header_clean = clean_column_name(header_str)
    
    # 1. Try direct and exact matches first
    for std_name, alternatives in column_alternatives.items():
        if header_clean.upper() == std_name.upper():
            return std_name
        
        for alt in alternatives:
            if header_clean.upper() == clean_column_name(alt).upper():
                return std_name
    
    # 2. Only after trying exact matches, try substring matches with stricter priorities
    for std_name, alternatives in column_alternatives.items():
        # Check if header contains standard name
        if std_name.upper() in header_clean.upper():
            # Special case: If checking for PERCENT_UNIQUES, make sure it has a % character
            if std_name == "PERCENT_UNIQUES" and "%" in header_str:
                return std_name
            # Special case: If checking for UNIQUES_REACH, and it has % character, skip 
            # to avoid conflict with PERCENT_UNIQUES
            elif std_name == "UNIQUES_REACH" and "%" in header_str:
                continue
            else:
                return std_name
    
    # 3. Last resort: try fuzzy matches with alternatives
    for std_name, alternatives in column_alternatives.items():
        for alt in alternatives:
            alt_clean = clean_column_name(alt)
            
            # Avoid mapping '%' columns to non-percent standards
            if "%" in header_str and "%" not in alt:
                continue
                
            # Check if alternative is substring of header or vice versa
            if alt_clean.upper() in header_clean.upper() or header_clean.upper() in alt_clean.upper():
                # Special case: If mapping to UNIQUES_REACH but the header has '%', skip
                if std_name == "UNIQUES_REACH" and "%" in header_str:
                    continue
                return std_name
    
    # If no match found, return None
    return None

# --- Constants - R&F Table Detection & Normalization ---
def _count_campaign(df: pd.DataFrame, label: str) -> None:
    """
    Diagnostic function to count Campaign Reach (Absl) and Campaign Freq metrics.
    Logs the count to help track where these metrics are being lost in processing.
    
    Args:
        df: DataFrame to analyze
        label: Descriptive label for logging context
    """
    if df is None or df.empty:
        logging.info(f"[CAMPAIGN COUNT] {label}: DataFrame empty or None")
        return
    
    campaign_reach_count = 0
    campaign_freq_count = 0
    
    # Check in PLATFORM column for campaign metrics
    if 'PLATFORM' in df.columns:
        platform_col = df['PLATFORM'].fillna('').astype(str)
        campaign_reach_count = platform_col.str.contains('Campaign Reach', case=False).sum()
        campaign_freq_count = platform_col.str.contains('Campaign Freq', case=False).sum()
    
    logging.info(f"[CAMPAIGN COUNT] {label}: Campaign Reach rows={campaign_reach_count}, Campaign Freq rows={campaign_freq_count}, Total rows={len(df)}")

PCA_INDICATORS = ["NTM (SPEND)", "PCA REPORT", "MEDIA DELIVERY", "POST CAMPAIGN", "DELIVERED", "ACTUAL DELIVERY"]

RF_METRIC_MAP = {
    # Reach metrics
    "Campaign Reach (Absl)": "UNIQUES_REACH",
    "Campaign Reach": "UNIQUES_REACH",
    "Campaign Reach Absl": "UNIQUES_REACH",
    "Awareness Reach Absl": "UNIQUES_REACH",
    "Consideration Reach Absl - FOR GNE": "UNIQUES_REACH",
    "Consideration Reach Absl - FOR GNE (COMBINE PROG BUYS)": "UNIQUES_REACH",
    "Purchase Reach Absl": "UNIQUES_REACH",
    "Reach": "UNIQUES_REACH",
    "Unique Reach": "UNIQUES_REACH",
    "Unique Users": "UNIQUES_REACH",
    "Estimated Ad Recall Lift (people)": "UNIQUES_REACH",
    "Ad Recall Lift": "UNIQUES_REACH",
    "Total Reach": "UNIQUES_REACH",
    
    # Frequency metrics
    "Campaign Freq.": "FREQUENCY",
    "Campaign Frequency": "FREQUENCY",
    "Awareness Freq.": "FREQUENCY",
    "Consideration Freq.": "FREQUENCY",
    "Purchase Freq.": "FREQUENCY",
    "Frequency": "FREQUENCY",
    "Avg. Frequency": "FREQUENCY",
    "Average Frequency": "FREQUENCY",
    
    # Standard delivery metrics
    "Impressions": "IMPRESSIONS",
    "Total Impressions": "IMPRESSIONS",
    "Delivered Impressions": "IMPRESSIONS",
    "Clicks": "CLICKS_ACTIONS",
    "Total Clicks": "CLICKS_ACTIONS",
    "Link Clicks": "CLICKS_ACTIONS",
    "Views": "VIDEO_VIEWS",
    "Video Views": "VIDEO_VIEWS",
    "Total Video Views": "VIDEO_VIEWS",
    "Video Completes": "VIDEO_VIEWS",
    "NTM (Spend)": "BUDGET_LOCAL",
    "NTM": "BUDGET_LOCAL",
    "Spend": "BUDGET_LOCAL",
    "Total Spend": "BUDGET_LOCAL",
    "Media Cost": "BUDGET_LOCAL",
    "Media Spend": "BUDGET_LOCAL",
    "Investment": "BUDGET_LOCAL",
    
    # Performance metrics
    "CTR": "CTR_PERCENT",
    "VTR": "VTR_PERCENT",
    "CPM": "CPM_LOCAL",
    "CPC": "CPC_LOCAL",
    "CPV": "CPV_LOCAL"
    
    # NOTE: Market names like "UAE", "Bahrain" etc. are handled by the pivot's index, not this map.
}

# Mapping for CEJ_OBJECTIVES based on R&F metrics
RF_CEJ_MAP = {
    "Campaign Reach (Absl)": "N/A", 
    "Campaign Reach": "N/A",
    "Campaign Reach Absl": "N/A",
    "Campaign Freq.": "N/A",
    "Campaign Frequency": "N/A",
    "Awareness Reach Absl": "Awareness",
    "Awareness Freq.": "Awareness",
    "Consideration Reach Absl - FOR GNE": "Consideration",
    "Consideration Reach Absl - FOR GNE (COMBINE PROG BUYS)": "Consideration",
    "Consideration Freq.": "Consideration",
    "Purchase Reach Absl": "Purchase",
    "Purchase Freq.": "Purchase"
}

def process_rf_table(raw_sheet_df: pd.DataFrame, metric_col_actual_name: str, sheet_name: str, source_file: str) -> pd.DataFrame:
    """
    Special processor for R&F tables from DELIVERED files.
    Transforms wide R&F data (metrics in rows, markets in columns) into a long,
    row-based format as per user specification.
    
    Args:
        raw_sheet_df: DataFrame containing the RAW R&F table data from the Excel sheet,
                      BEFORE generic header mapping by extract_data_to_dataframe.
                      The first column of this raw data is expected to contain R&F metric names.
        metric_col_actual_name: The actual name of the first column in raw_sheet_df that contains the metrics.
        sheet_name: Name of the current sheet.
        source_file: Source file name.
        
    Returns:
        DataFrame containing the processed R&F data in the specified long format,
        or an empty DataFrame if processing fails or no R&F data is found.
    """
    log_prefix = f"[{sheet_name} R&F Process]"
    logging.info(f"{log_prefix} Starting specialized R&F processing for DELIVERED format. Metric Col Hint: '{metric_col_actual_name}'")
    
    if raw_sheet_df.empty or len(raw_sheet_df.columns) < 2:
        logging.warning(f"{log_prefix} Input raw DataFrame is empty or has less than 2 columns. Cannot process.")
        return pd.DataFrame()
    
    # Identify the metric column and market columns from the raw DataFrame
    # The metric_col_actual_name is the name of the column holding R&F metric strings.
    # We need to find this column in raw_sheet_df.columns. It's typically the first one.

    if metric_col_actual_name not in raw_sheet_df.columns:
        # Fallback if the provided name isn't directly in columns (e.g. if it was cleaned slightly)
        # or if the first column is the reliable source.
        # For safety, let's assume the first column of the raw_sheet_df IS the metric column.
        # This was the previous implicit assumption based on structure.
        identified_metric_column_name = raw_sheet_df.columns[0]
        logging.warning(f"{log_prefix} Provided metric_col_actual_name '{metric_col_actual_name}' not found in raw_sheet_df columns. Assuming first column '{identified_metric_column_name}' holds metrics.")
    else:
        identified_metric_column_name = metric_col_actual_name
    
    market_column_names = [col for col in raw_sheet_df.columns if col != identified_metric_column_name]

    if not market_column_names:
        logging.warning(f"{log_prefix} No market columns found (expected at least one after the metric column '{identified_metric_column_name}').")
        return pd.DataFrame()

    logging.debug(f"{log_prefix} Identified Metric column: '{identified_metric_column_name}'. Identified Market columns: {market_column_names}")

    # Ensure RF_CEJ_MAP is available (it should be global)
    global RF_CEJ_MAP
    if 'RF_CEJ_MAP' not in globals():
        logging.error(f"{log_prefix} RF_CEJ_MAP is not defined globally. Cannot proceed.")
        # Define a default empty map to avoid crashing, but this is an error condition
        RF_CEJ_MAP = {}

    # Define which R&F metrics map to which final output columns
    # Based on user feedback: UNIQUES_REACH for reach, FREQUENCY for freq.
    rf_metric_to_output_col_map = {
        "Campaign Reach (Absl)": "UNIQUES_REACH",
        "Awareness Reach Absl": "UNIQUES_REACH",
        "Consideration Reach Absl - FOR GNE": "UNIQUES_REACH",
        "Consideration Reach Absl - FOR GNE (COMBINE PROG BUYS)": "UNIQUES_REACH",
        "Purchase Reach Absl": "UNIQUES_REACH",
        "Campaign Freq.": "FREQUENCY",
        "Awareness Freq.": "FREQUENCY",
        "Consideration Freq.": "FREQUENCY",
        "Purchase Freq.": "FREQUENCY"
        # Add other specific R&F metrics if they appear and need mapping
    }
    # All R&F metric names that this function is designed to pick up from the input table.
    # This also helps to filter rows from sheet_df that are actual R&F metrics.
    known_rf_input_metrics = list(rf_metric_to_output_col_map.keys())

    processed_rows = []

    for index, row_data in raw_sheet_df.iterrows():
        # Get the R&F metric string from the identified_metric_column_name
        current_rf_metric_input = str(row_data[identified_metric_column_name]).strip()

        # Enhanced logging for debugging Region 1 issues
        if index < 5:  # Log first 5 rows for debugging
            logging.info(f"{log_prefix} Row {index} metric: '{current_rf_metric_input}'")

        # Only process if the metric in the first column is one of the known R&F types
        if not current_rf_metric_input or current_rf_metric_input not in known_rf_input_metrics:
            logging.debug(f"{log_prefix} Skipping row {index} with metric '{current_rf_metric_input}' as it's not a targeted R&F metric.")
            continue
            
        logging.debug(f"{log_prefix} Processing input R&F metric: '{current_rf_metric_input}'")
        
        # Special debug logging for Campaign metrics
        if 'Campaign' in current_rf_metric_input:
            logging.info(f"{log_prefix} CAMPAIGN METRIC DETECTED: '{current_rf_metric_input}' at row {index}")
        
        for market_name_original in market_column_names:
            market_name = str(market_name_original).strip()
            value = row_data[market_name_original]
            
            # Skip rows with empty market values or "END" in market name
            if not market_name or market_name.upper() == "END" or market_name.upper() == "N/A" or market_name == "":
                continue
                
            # Handle "n/a" values as valid business data (convert to 0 for calculations)
            if str(value).lower() == "n/a":
                value = 0  # Convert n/a to 0 for numeric calculations
            # Skip rows where the value is completely empty or END marker
            elif pd.isna(value) or value == "" or value == "END":
                continue

            # Create a new row dictionary based on the detailed user specification
            output_row = {
                "Source_File": os.path.basename(source_file),
                "Source_Sheet": sheet_name,
                "MARKET": market_name,
                "BRAND": "N/A",
                "CAMPAIGN": "N/A",
                "PLATFORM": current_rf_metric_input, # User spec: PLATFORM is the R&F metric name
                "CEJ_OBJECTIVES": RF_CEJ_MAP.get(current_rf_metric_input, "N/A"),
                "FORMAT_TYPE": "N/A",
                "PLACEMENT": "N/A",
                "AD_UNIT_TYPE": "N/A",
                "DEVICE": "N/A",
                "TARGET_AUDIENCE": "N/A",
                "BUYING_MODEL": "N/A",
                "START_DATE": pd.NA, # User spec: N/A for R&F specific rows
                "END_DATE": pd.NA,   # User spec: N/A
                "WEEKS": pd.NA,      # User spec: N/A
                "LOCAL_CURRENCY": "N/A",
                "BUDGET_LOCAL": pd.NA, # User spec: N/A
                "IMPRESSIONS": pd.NA,  # User spec: N/A
                "CLICKS_ACTIONS": pd.NA,# User spec: N/A
                "VIDEO_VIEWS": pd.NA,  # User spec: N/A
                "FREQUENCY": pd.NA,    # Initialize to N/A
                "UNIQUES_REACH": pd.NA,# Initialize to N/A
                "PERCENT_UNIQUES": pd.NA, # User spec: N/A
                "CPM_LOCAL": pd.NA,      # User spec: N/A
                "CPC_LOCAL": pd.NA,      # User spec: N/A
                "CPV_LOCAL": pd.NA,      # User spec: N/A
                "CTR_PERCENT": pd.NA,    # User spec: N/A
                "VTR_PERCENT": pd.NA,    # User spec: N/A
                "PLATFORM_FEE_LOCAL": pd.NA, # User spec: N/A
                "PLATFORM_BUDGET_LOCAL": pd.NA, # User spec: N/A
                "TA_SIZE": pd.NA,          # User spec: N/A
                "MEDIA_KPIS": pd.NA,       # User spec: N/A
                "COMMENTS": pd.NA,         # User spec: N/A
                "CREATIVE_NAME": pd.NA,    # User spec: N/A,
                "Source_Type": "DELIVERED R&F"  # Updated to differentiate R&F data
            }

            # Populate FREQUENCY or UNIQUES_REACH based on the current R&F metric
            target_output_column = rf_metric_to_output_col_map.get(current_rf_metric_input)
            if target_output_column:
                output_row[target_output_column] = value
            else:
                # Try mapping via RF_METRIC_MAP as a fallback if direct key not in rf_metric_to_output_col_map
                mapped_to_std_metric = map_rf_metric_to_standard(current_rf_metric_input)
                if mapped_to_std_metric in ["UNIQUES_REACH", "FREQUENCY"]: # Check if it maps to one of the target cols
                    output_row[mapped_to_std_metric] = value
                    logging.debug(f"{log_prefix} Metric '{current_rf_metric_input}' mapped to '{mapped_to_std_metric}' and value '{value}' placed.")
                else:
                    logging.warning(f"{log_prefix} R&F metric '{current_rf_metric_input}' (maps to: {mapped_to_std_metric}) does not have a defined target output column (UNIQUES_REACH/FREQUENCY). Value '{value}' for market '{market_name}' will not be placed.")
            
            processed_rows.append(output_row)

    if not processed_rows:
        logging.warning(f"{log_prefix} No R&F data rows were processed. Check input table structure and metric names.")
        return pd.DataFrame()

    final_rf_df = pd.DataFrame(processed_rows)
    
    # Ensure all FINAL_OUTPUT_COLUMNS are present, adding missing ones with pd.NA
    for col in FINAL_OUTPUT_COLUMNS:
        if col not in final_rf_df.columns:
            final_rf_df[col] = pd.NA
    
    # Reorder columns to match FINAL_OUTPUT_COLUMNS
    final_rf_df = final_rf_df[FINAL_OUTPUT_COLUMNS]
    
    logging.info(f"{log_prefix} Successfully processed {len(final_rf_df)} R&F rows. Shape: {final_rf_df.shape}")
    return final_rf_df

def normalize_rf_table(df_mapped: pd.DataFrame, df_raw: pd.DataFrame, raw_first_column_name: str, sheet_name: str, source_file:str, source_region_idx:int, detection_method:str, file_format_for_rf_logic: str) -> Optional[pd.DataFrame]:
    """
    Detects if a DataFrame is a Reach & Frequency (R&F) table and normalizes it.
    For 'delivered' format, it calls process_rf_table with the raw data.
    Otherwise, or if R&F processing fails, it returns the original mapped DataFrame.

    Args:
        df_mapped: The DataFrame with headers already mapped to standard names.
        df_raw: The raw DataFrame from the Excel sheet for this region, before header mapping.
        raw_first_column_name: The original string name of the first column from the raw headers.
        sheet_name: Name of the sheet.
        source_file: Source file name.
        source_region_idx: Index of the region within the sheet.
        detection_method: How this region was initially detected (e.g., 'markers').
        file_format_for_rf_logic: The format ('planned', 'delivered') to guide R&F logic.

    Returns:
        A processed DataFrame if R&F normalization was successful, 
        otherwise the original df_mapped.
    """
    log_prefix = f"[{sheet_name} Reg:{source_region_idx} M:{detection_method} FF:{file_format_for_rf_logic}] normalize_rf_table:"
    
    # Use df_mapped for initial checks and R&F detection heuristics if needed, as it has standardized names.
    # However, df_raw is crucial if we need to pass original structure to process_rf_table.
    if df_mapped.empty or len(df_mapped.columns) < 2:
        logging.debug(f"{log_prefix} Input mapped DataFrame is empty or has less than 2 columns. Skipping R&F normalization.")
        return df_mapped # Return original mapped df

    # Ensure df_raw is also available if we might need it for 'delivered' format
    if file_format_for_rf_logic == "delivered" and (df_raw is None or df_raw.empty):
        logging.warning(f"{log_prefix} 'delivered' format specified for R&F, but raw DataFrame is missing or empty. Cannot perform specialized R&F processing.")
        return df_mapped # Return original mapped df

    logging.debug(f"{log_prefix} Input mapped_df.head():\n{df_mapped.head().to_string()}")
    logging.debug(f"{log_prefix} Input mapped_df.columns: {df_mapped.columns.tolist()}")
    if df_raw is not None:
        logging.debug(f"{log_prefix} Input raw_df.head():\n{df_raw.head().to_string()}")
        logging.debug(f"{log_prefix} Input raw_df.columns: {df_raw.columns.tolist()}")
    logging.debug(f"{log_prefix} Received raw_first_column_name: '{raw_first_column_name}'")

    platform_value = 'DEFAULT'
    if 'PLATFORM' in df_mapped.columns and not df_mapped['PLATFORM'].empty and pd.notna(df_mapped['PLATFORM'].iloc[0]):
        platform_value = df_mapped['PLATFORM'].iloc[0]
    
    config = get_default_platform_config(platform_value)
    market_threshold = config.get('market_threshold', 0.5)
    logging.debug(f"{log_prefix} Using platform '{platform_value}' with market threshold {market_threshold}")
    
    # Add targeted diagnostic logging for Region 0 specifically
    if source_region_idx == 0:
        logging.debug(f"[{sheet_name} Reg0] first_col='{raw_first_column_name}'")
        logging.debug(f"[{sheet_name} Reg0] df_mapped.columns={list(df_mapped.columns)}")
        logging.debug(f"[{sheet_name} Reg0] market_threshold={market_threshold}")
    
    is_rf_table_detected = False
    
    # Enhanced R&F detection for DELIVERED format
    if file_format_for_rf_logic == "delivered" and raw_first_column_name and any(kw.lower() in raw_first_column_name.lower() for kw in ['metrics', '/market', 'metric', 'market']):
        # For DELIVERED format, check the actual data content in the first column to distinguish between:
        # 1. R&F table: First column contains metric names like "Campaign Reach", "Campaign Freq"
        # 2. Media table: First column contains market names like "UAE", "QAT" 
        rf_metric_indicators_found = 0
        market_name_indicators_found = 0
        
        # Check the mapped MARKET column to see what kind of data it contains
        if 'MARKET' in df_mapped.columns and not df_mapped['MARKET'].empty:
            market_col_values = df_mapped['MARKET'].astype(str).str.strip().str.upper()
            
            # Filter out marker/boundary values before analysis
            marker_values = ['END', 'START', '-', 'NAN', 'NONE', '', 'METRICS/ MARKET', 'METRICS / MARKET', 'METRICS / MARKETS']
            clean_values = [val for val in market_col_values if val not in marker_values and len(val.strip()) > 0]
            
            # Check for R&F metric indicators in the MARKET column
            # R&F table: MARKET column contains metric names like "Campaign Reach (Absl)", "Campaign Freq."
            rf_indicators = ['REACH', 'FREQ', 'CAMPAIGN', 'AWARENESS', 'CONSIDERATION', 'PURCHASE']
            for value in clean_values:
                if any(indicator in value for indicator in rf_indicators):
                    rf_metric_indicators_found += 1
            
            # Check for market name indicators in the MARKET column  
            # Media table: MARKET column contains market names like "UAE", "QAT"
            market_indicators = ['UAE', 'QAT', 'BAH', 'LEB', 'OMAN', 'QATAR', 'BAHRAIN', 'LEBANON']
            for value in clean_values:
                if any(indicator in value for indicator in market_indicators):
                    market_name_indicators_found += 1
            
            # Additional check: look for header patterns that indicate R&F structure
            header_like_values = market_col_values.head(3)  # Check first few values including headers
            contains_metrics_header = any('METRICS' in str(val) for val in header_like_values)
            
            logging.debug(f"{log_prefix} MARKET column analysis: {rf_metric_indicators_found} metrics, {market_name_indicators_found} markets, header_pattern: {contains_metrics_header}, clean_values: {len(clean_values)}")
            
            # Strong indicator of R&F table if header contains "METRICS" pattern
            if contains_metrics_header:
                rf_metric_indicators_found += 10  # Boost R&F score significantly
        
        # Classify based on data content - R&F table should have metric names in first column
        # For R&F table: Look for metric names in the first column (values like "Campaign Reach", "Campaign Freq")
        # For Media table: Look for market names in the first column (values like "UAE", "QAT")
        first_col_contains_metrics = rf_metric_indicators_found > 0
        first_col_contains_markets = market_name_indicators_found > 0
        
        # Check for clear R&F indicators - if we have the classic R&F metrics, it's definitely R&F
        classic_rf_metrics = ['CAMPAIGN REACH', 'CAMPAIGN FREQ', 'AWARENESS REACH', 'AWARENESS FREQ', 'CONSIDERATION REACH', 'PURCHASE REACH']
        has_classic_rf_pattern = any(classic in str(val) for val in clean_values for classic in classic_rf_metrics)
        
        if has_classic_rf_pattern or contains_metrics_header:
            # Clear R&F indicators found
            is_rf_table_detected = True
            logging.debug(f"{log_prefix} R&F table detected (classic pattern): classic_rf={has_classic_rf_pattern}, metrics_header={contains_metrics_header}")
        elif first_col_contains_metrics and not first_col_contains_markets:
            # First column has metrics, not markets = R&F table
            is_rf_table_detected = True
            logging.debug(f"{log_prefix} R&F table detected: first column contains {rf_metric_indicators_found} metric indicators, {market_name_indicators_found} market indicators")
        elif first_col_contains_markets and not first_col_contains_metrics:
            # First column has markets, not metrics = Media table  
            is_rf_table_detected = False
            logging.debug(f"{log_prefix} Media table detected: first column contains {market_name_indicators_found} market indicators, {rf_metric_indicators_found} metric indicators")
        elif rf_metric_indicators_found >= 6 and market_name_indicators_found > 0:
            # If we have at least 6 R&F metrics (standard count), prefer R&F even if markets are present
            is_rf_table_detected = True
            logging.debug(f"{log_prefix} R&F table detected (metric count preference): {rf_metric_indicators_found} metric indicators vs {market_name_indicators_found} market indicators")
        else:
            # Default to media table when unclear
            is_rf_table_detected = False
            logging.debug(f"{log_prefix} Media table detected (default): {rf_metric_indicators_found} metric indicators vs {market_name_indicators_found} market indicators")
    elif raw_first_column_name and any(kw.lower() in raw_first_column_name.lower() for kw in ['metrics', '/market', 'metric', 'market']):
        # Original logic for non-delivered formats
        is_rf_table_detected = True
        logging.debug(f"{log_prefix} R&F hint from first column name: '{raw_first_column_name}'")

    if not is_rf_table_detected and 'MARKET' in df_mapped.columns:
        if any('REACH' in str(col).upper() or 'FREQ' in str(col).upper() for col in df_mapped.columns if col != 'MARKET'):
            is_rf_table_detected = True
            logging.debug(f"{log_prefix} R&F hint from existing 'MARKET' column and REACH/FREQ metrics in other columns.")

    if not is_rf_table_detected:
        potential_market_data_cols = df_mapped.columns[1:] 
        total_data_cols_for_threshold = len(potential_market_data_cols)
        if total_data_cols_for_threshold > 0:
            global KNOWN_MARKETS
            if 'KNOWN_MARKETS' not in globals():
                logging.warning(f"{log_prefix} KNOWN_MARKETS is not defined. Market name matching will be limited.")
                KNOWN_MARKETS = set()
            market_like_cols_found = [col for col in potential_market_data_cols 
                                      if str(col).upper() in KNOWN_MARKETS or 
                                         any(m_kw.upper() in str(col).upper() for m_kw in ['MARKET', 'COUNTRY', 'REGION'])]
            if 'MARKET' in potential_market_data_cols and 'MARKET' not in market_like_cols_found:
                 market_like_cols_found.append('MARKET')
            count_market_like_cols = len(set(market_like_cols_found))
            required_market_cols = int(total_data_cols_for_threshold * market_threshold)
            if total_data_cols_for_threshold > 0 and required_market_cols == 0: 
                 required_market_cols = 1
            logging.debug(f"{log_prefix} Market column check: Found {count_market_like_cols} market-like columns ... Required: {required_market_cols}.")
            if count_market_like_cols >= required_market_cols:
                is_rf_table_detected = True
            else:
                logging.info(f"{log_prefix} Failed market column check for R&F.")
        else:
            logging.debug(f"{log_prefix} Not enough data columns for market proportion check.")

    if not is_rf_table_detected:
        logging.info(f"{log_prefix} Table not classified as R&F.")
        return df_mapped # Return original mapped df

    logging.info(f"{log_prefix} Table classified as R&F. Proceeding.")

    if file_format_for_rf_logic == "delivered":
        logging.debug(f"{log_prefix} Applying specialized process_rf_table for 'delivered' format using df_raw.")
        if df_raw is None or df_raw.empty:
            logging.warning(f"{log_prefix} Cannot call process_rf_table: df_raw is missing or empty for 'delivered' format.")
            return df_mapped # Fallback to mapped if raw is not usable

        # Attempt to find the actual R&F header within df_raw for 'delivered' files
        actual_rf_header_name_in_raw = None
        rf_header_row_index_in_raw = -1

        # Scan first few rows of df_raw to find a better R&F header
        # Typically, the R&F metrics start with a column like "METRICS / MARKETS"
        # PCA_INDICATORS contains relevant keywords like "METRICS / MARKETS", "CAMPAIGN REACH", etc.
        # We are looking for a row where the first cell contains such an indicator.
        for i in range(min(5, len(df_raw))): # Scan up to first 5 rows of the raw data region
            try:
                first_cell_val_raw = str(df_raw.iloc[i, 0]).strip().upper()
                # Check if this first cell value in raw data looks like an R&F metric header identifier
                if any(indicator.upper() in first_cell_val_raw for indicator in PCA_INDICATORS if indicator.upper() in ["METRICS/ MARKET", "METRICS / MARKETS", "METRICS / MARKET"]):
                    # This looks like the actual R&F metrics column header
                    # The name of this column in df_raw.columns is what we need
                    actual_rf_header_name_in_raw = df_raw.columns[0]
                    rf_header_row_index_in_raw = i
                    logging.info(f"{log_prefix} Found potential R&F metrics header '{actual_rf_header_name_in_raw}' (value: '{df_raw.iloc[i, 0]}') at row {i} within the raw data region.")
                    break # Found it
            except Exception as e_scan:
                logging.debug(f"{log_prefix} Error scanning row {i} of df_raw for R&F header: {e_scan}")
                continue
        
        effective_metric_col_name_for_process = raw_first_column_name # Default to what was passed in
        df_to_process_for_rf = df_raw.copy() # Default to using the whole raw df

        if actual_rf_header_name_in_raw and rf_header_row_index_in_raw != -1:
            # We found a more specific R&F header within the df_raw region.
            # The actual R&F data starts from this row.
            effective_metric_col_name_for_process = actual_rf_header_name_in_raw
            # Slice df_raw to start from this identified R&F header row
            # The header of this new df_to_process_for_rf will be the values from this row,
            # and data will be subsequent rows.
            
            # Take the identified row as header
            new_header = df_raw.iloc[rf_header_row_index_in_raw]
            df_temp_rf = df_raw.iloc[rf_header_row_index_in_raw + 1:].copy() # Get data rows after this header
            if not df_temp_rf.empty:
                df_temp_rf.columns = new_header # Set the new header
                df_to_process_for_rf = df_temp_rf.reset_index(drop=True)
                # The first column name of this newly realigned df is what process_rf_table needs.
                effective_metric_col_name_for_process = df_to_process_for_rf.columns[0]
                logging.info(f"{log_prefix} Realigned df_raw for R&F processing. New metric col for process_rf_table: '{effective_metric_col_name_for_process}'. Data shape: {df_to_process_for_rf.shape}")
            else:
                logging.warning(f"{log_prefix} Found R&F header row but no data rows below it in df_raw. Proceeding with originally passed raw_first_column_name.")
        else:
            logging.info(f"{log_prefix} No specific R&F metrics header found within first few rows of df_raw. Using originally passed raw_first_column_name: '{raw_first_column_name}'.")
        
        # Call process_rf_table with the (potentially realigned) df_to_process_for_rf and its first column name
        rf_processed_df = process_rf_table(df_to_process_for_rf, str(effective_metric_col_name_for_process), sheet_name, source_file)
        
        if rf_processed_df is not None and not rf_processed_df.empty:
            logging.info(f"{log_prefix} Specialized R&F processing successful. Output shape: {rf_processed_df.shape}")
            return rf_processed_df # Return the successfully processed R&F data
        else:
            logging.warning(f"{log_prefix} Specialized R&F processing (process_rf_table) returned empty or None. Returning original mapped df.")
            return df_mapped # Fallback to mapped if R&F processing failed
    
    logging.info(f"{log_prefix} Format is not 'delivered' or R&F processing path not taken. Standard R&F melt/pivot is SKIPPED. Returning original mapped df.")
    return df_mapped # Return original mapped df for non-delivered or if other paths didn't result in processing

def map_rf_metric_to_standard(metric_name: str) -> Optional[str]:
    """
    Maps an R&F metric name to a standardized column name.
    """
    if not metric_name:
        return None
        
    # Try exact match first
    if metric_name in RF_METRIC_MAP:
        return RF_METRIC_MAP[metric_name]
        
    # Try case-insensitive match
    metric_lower = metric_name.lower()
    for pattern, std_name in RF_METRIC_MAP.items():
        if pattern.lower() == metric_lower:
            return std_name
            
    # Try partial matching
    for pattern, std_name in RF_METRIC_MAP.items():
        if pattern.lower() in metric_lower:
            return std_name
            
    # No match found
    return None

def validate_data_types(df: pd.DataFrame, quality_score: float, warnings: List[str]) -> Tuple[pd.DataFrame, float]:
    """
    Validates that essential data columns have appropriate values and types.
    Returns the validated DataFrame and an updated quality score.
    
    Args:
        df: The DataFrame to validate
        quality_score: Initial quality score
        warnings: List to append warnings to
        
    Returns:
        Tuple of (validated DataFrame, updated quality score)
    """
    if df is None or df.empty:
        warnings.append("Empty DataFrame provided for validation")
        return df, 0.0
    
    # Check essential columns
    for col in ESSENTIAL_DATA_COLUMNS_FOR_VALIDATION:
        if col not in df.columns:
            warnings.append(f"Essential column '{col}' missing")
            quality_score *= 0.8
        elif df[col].isna().all():
            warnings.append(f"Essential column '{col}' has all NA values")
            quality_score *= 0.9
    
    # For numeric columns, ensure they contain numeric data
    for col in [c for c in NUMERIC_COLUMNS if c in df.columns]:
        non_numeric_count = 0
        for val in df[col].dropna():
            if not isinstance(val, (int, float)) and not (isinstance(val, str) and val.replace('.', '', 1).replace('-', '', 1).isdigit()):
                non_numeric_count += 1
        
        if non_numeric_count > 0:
            warnings.append(f"Column '{col}' contains {non_numeric_count} non-numeric values")
            quality_score *= 0.95
    
    return df, quality_score

def validate_and_convert_numeric_data(df: pd.DataFrame, log_prefix: str) -> Tuple[pd.DataFrame, List[str], float]:
    """
    Thoroughly validates and converts numeric data in the DataFrame.
    Returns the processed DataFrame, warnings, and a confidence score.
    """
    # A "suspicious" value is one that technically can be converted to a numeric value,
    # but violates expected constraints for the column type, such as:
    # 1. Negative numbers in columns that should only contain positive values (e.g., IMPRESSIONS, CLICKS)
    # 2. Decimal values in columns that should only contain integers (e.g., count-based metrics)
    # 3. Values that are outside the expected range for the column type
    # Suspicious values are converted but marked for reporting to help identify potential data issues.
    
    warnings = []
    confidence = 1.0
    
    # Copy DataFrame to avoid modifying the original during validation
    df_validated = df.copy()
    
    # Track conversion success rates for each numeric column
    conversion_stats = {}
    
    # Define numeric column patterns
    numeric_patterns = {
        'BUDGET': {'min': 0, 'allow_decimals': True, 'currency_symbols': True, 'allow_negative': False},
        'IMPRESSIONS': {'min': 0, 'allow_decimals': False, 'currency_symbols': False, 'allow_negative': False},
        'VIEWS': {'min': 0, 'allow_decimals': False, 'currency_symbols': False, 'allow_negative': False},
        'CLICKS': {'min': 0, 'allow_decimals': False, 'currency_symbols': False, 'allow_negative': False},
        'ACTIONS': {'min': 0, 'allow_decimals': False, 'currency_symbols': False, 'allow_negative': False},
        'CPM': {'min': 0, 'allow_decimals': True, 'currency_symbols': True, 'allow_negative': False},
        'CPC': {'min': 0, 'allow_decimals': True, 'currency_symbols': True, 'allow_negative': False},
        'CPV': {'min': 0, 'allow_decimals': True, 'currency_symbols': True, 'allow_negative': False},
        'FREQUENCY': {'min': 0, 'allow_decimals': True, 'currency_symbols': False, 'allow_negative': False},
        'REACH': {'min': 0, 'allow_decimals': False, 'currency_symbols': False, 'allow_negative': False},
        'COST': {'min': 0, 'allow_decimals': True, 'currency_symbols': True, 'allow_negative': False},
        'PERCENT': {'min': None, 'allow_decimals': True, 'currency_symbols': False, 'allow_negative': False}
    }
    
    # Define known placeholder values that would cause conversion errors
    known_placeholders = ['n.a', 'n.a.', 'n/a', 'na', 'n.a-', '#DIV/0!', '#VALUE!', '#REF!', '#N/A']
    
    for col in df_validated.columns:
        # Determine if column should be numeric based on name or content
        is_numeric_column = False
        validation_rules = None
        
        # Check if column matches any numeric patterns
        for pattern, rules in numeric_patterns.items():
            if pattern in col.upper():
                is_numeric_column = True
                validation_rules = rules
            break
            
        # Also check NUMERIC_COLUMNS constant
        if col in NUMERIC_COLUMNS:
            is_numeric_column = True
            # Use default rules if not already set
            if not validation_rules:
                validation_rules = {'min': 0, 'allow_decimals': True, 'currency_symbols': False, 'allow_negative': False}
                
        if not is_numeric_column:
            # Check if column contains mostly numeric values
            sample = df_validated[col].dropna().head(100)  # Check first 100 non-null values
            if len(sample) > 0:
                numeric_count = sum(1 for val in sample if isinstance(val, (int, float)) or 
                                  (isinstance(val, str) and 
                                   bool(re.match(r'^[\s]*[+-]?[\d,.]+%?[\s]*$', val.strip()))))
                if numeric_count / len(sample) > 0.8:  # If >80% values appear numeric
                    is_numeric_column = True
                    validation_rules = {'min': None, 'allow_decimals': True, 'currency_symbols': False, 'allow_negative': True}
                    logging.info(f"{log_prefix} Auto-detected numeric column: {col}")
            
        if not is_numeric_column:
            continue
            
        total_non_null = len(df_validated[col].dropna())
        if total_non_null == 0:
            warnings.append(f"Numeric column '{col}' has no non-null values")
            confidence *= 0.95
            continue
            
        # Initialize stats
        conversion_stats[col] = {
            'total_values': total_non_null,
            'successful_conversions': 0,
            'failed_conversions': 0,
            'suspicious_values': 0,
            'known_placeholder_failures': 0,  # Track placeholder errors separately
            'validation_rules': validation_rules
        }
        
        # Create a temporary series for validation
        temp_series = df_validated[col].copy()
        
        # Pre-process values
        for idx, val in enumerate(temp_series):
            if pd.isna(val):
                continue
                
            try:
                # Convert to string and clean
                str_val = str(val).strip()
                original_val = str_val  # Keep original for logging
                
                # Handle percentage values
                if '%' in str_val:
                    str_val = str_val.replace('%', '')
                    is_percentage = True
                else:
                    is_percentage = False
                    
                # Remove currency symbols and commas
                if validation_rules['currency_symbols']:
                    str_val = re.sub(r'[$€£¥]', '', str_val)
                str_val = str_val.replace(',', '')
                
                # Handle special cases
                if str_val.lower() in ['na', 'n.a', 'n/a', '-', '', '--']:
                    temp_series.iloc[idx] = 0  # Convert to 0 for delivered templates
                    conversion_stats[col]['successful_conversions'] += 1
                    continue
                    
                # Handle parentheses for negative numbers
                if str_val.startswith('(') and str_val.endswith(')'):
                    str_val = '-' + str_val[1:-1]
                    
                # Convert to float
                float_val = float(str_val)
                
                # Handle percentages
                if is_percentage:
                    float_val = float_val / 100
                    
                # Validate the value
                if validation_rules['min'] is not None and float_val < validation_rules['min']:
                    if not validation_rules.get('allow_negative', False):
                        warnings.append(f"Negative value found in column '{col}': {original_val}")
                        confidence *= 0.98
                        conversion_stats[col]['suspicious_values'] += 1
                        # Set negative values to 0 for delivered templates
                        float_val = 0
                    
                if not validation_rules['allow_decimals'] and not float_val.is_integer():
                    warnings.append(f"Non-integer value found in count column '{col}': {original_val}")
                    confidence *= 0.98
                    conversion_stats[col]['suspicious_values'] += 1
                    # Round to nearest integer for count columns
                    float_val = round(float_val)
                    
                # Update the value in the temporary series
                temp_series.iloc[idx] = float_val
                conversion_stats[col]['successful_conversions'] += 1
                
            except (ValueError, TypeError) as e:
                # Check if this is a known placeholder value
                is_known_placeholder = False
                for placeholder in known_placeholders:
                    if str(val).strip() == placeholder:
                        conversion_stats[col]['known_placeholder_failures'] += 1
                        is_known_placeholder = True
                        break
                
                if not is_known_placeholder:
                    # Only log unexpected conversion errors individually
                    logging.warning(f"{log_prefix} Could not convert value '{val}' in column '{col}': {e}")
                
                conversion_stats[col]['failed_conversions'] += 1
                # Set failed conversions to 0 for delivered templates
                temp_series.iloc[idx] = 0
                
        # Calculate conversion success rate
        success_rate = conversion_stats[col]['successful_conversions'] / total_non_null
        
        if success_rate < 0.5:
            warnings.append(f"Low conversion success rate ({success_rate*100:.1f}%) for numeric column '{col}'")
            confidence *= 0.9
        elif success_rate < 0.8:
            warnings.append(f"Moderate conversion success rate ({success_rate*100:.1f}%) for numeric column '{col}'")
            confidence *= 0.95
            
        # Update the original DataFrame with validated values
        df_validated[col] = temp_series
        
        # Log conversion statistics
        if conversion_stats[col]['suspicious_values'] > 0:
            logging.debug(f"{log_prefix} Column '{col}' conversion stats: "
                         f"Success: {conversion_stats[col]['successful_conversions']}, "
                         f"Failed: {conversion_stats[col]['failed_conversions']}, "
                         f"Suspicious: {conversion_stats[col]['suspicious_values']} (may need review)")
        else:
            logging.debug(f"{log_prefix} Column '{col}' conversion stats: "
                         f"Success: {conversion_stats[col]['successful_conversions']}, "
                         f"Failed: {conversion_stats[col]['failed_conversions']}")
        
        # Log a summary message for known placeholder failures
        if conversion_stats[col]['known_placeholder_failures'] > 0:
            logging.warning(f"{log_prefix} Column '{col}': {conversion_stats[col]['known_placeholder_failures']} "
                          f"values could not be converted due to known placeholders ('n.a.', '#DIV/0!', etc.).")
                     
    return df_validated, warnings, confidence

# --- Main Processing Logic ---
def detect_file_format(workbook_path: str, wb: Workbook) -> str:
    """
    Detect the file format ("planned" or "delivered") using simple, reliable filename heuristics.

    This replacement avoids the previous complex, badly-indented logic while still satisfying
    current test expectations:
        • Any filename containing keywords like "PLANNED", "MEDIA PLAN", "PROPOSAL" ⇒ planned
        • Any filename containing "DELIVERED", "PCA", "POST CAMPAIGN" ⇒ delivered
        • Otherwise "unknown"
    """
    filename_upper = Path(workbook_path).name.upper()

    planned_keywords = [
        "PLANNED", "MEDIA PLAN", "MEDIA_PLAN", "PLAN TEMPLATE", "PROPOSAL", "BRIEF"
    ]
    delivered_keywords = [
        "DELIVERED", "PCA", "POST CAMPAIGN", "POST_CAMPAIGN", "WRAP", "ACTUAL", "EOC"
    ]

    if any(kw in filename_upper for kw in planned_keywords):
        return "planned"
    if any(kw in filename_upper for kw in delivered_keywords):
        return "delivered"
    return "unknown"

def map_column_name(header: str) -> Tuple[str, float]:
    """
    Wrapper function for map_header_to_standard that maps column names
    to standardized names with confidence scores.
    
    Args:
        header: Raw header string from Excel
        
    Returns:
        Tuple of (mapped column name, confidence score 0.0-1.0)
    """
    if not header:
        return None, 0.0
        
    # Use the implementation directly
    mapped_name = map_to_standard_column(header, COLUMN_ALTERNATIVES)
    
    # Simple confidence scoring - we could make this more sophisticated
    confidence = 1.0 if mapped_name else 0.0
    
    return mapped_name, confidence

def extract_data_to_dataframe(sheet, sheet_name, region):
    """
    Extracts data from a given region of a sheet into a pandas DataFrame.
    Identifies headers, maps them to standard names, and applies basic cleaning.
    
    Args:
        sheet: The openpyxl worksheet object.
        sheet_name: The name of the sheet.
        region (dict): A dictionary with keys 'header_row', 'data_start_row',
                         'data_end_row', 'data_start_col', 'data_end_col',
                         'format' (e.g., 'standard', 'delivered_rf')
        
    Returns:
        tuple: (df_mapped_region, df_raw_region, raw_headers, region_format_flag)
               df_mapped_region: DataFrame with standardized column names.
               df_raw_region: DataFrame with original column names from the sheet for that region.
               raw_headers: List of raw header strings.
               region_format_flag: The 'format' flag from the input region dict.
    """
    log_prefix = f"[{sheet_name} Extr]" 
    # region dictionary values (like header_row, start_row etc.) are 1-based from find_table_regions
    # Convert to 0-based for internal openpyxl usage
    header_row_idx = region['header_row'] -1 # 0-indexed for openpyxl
    data_start_row_idx = region['start_row'] -1 # 0-indexed
    data_end_row_idx = region['end_row'] -1 # 0-indexed
    data_start_col_idx = region['start_col'] -1 # 0-indexed
    data_end_col_idx = region['end_col'] -1 # 0-indexed
    region_format_flag = region.get('format', None) # e.g., 'standard', 'delivered_rf', 'planned_rf'

    logging.debug(f"{log_prefix} Extracting region (1-based input): Header Row {region['header_row']}, Data Rows {region['start_row']}-{region['end_row']}, Cols {region['start_col']}-{region['end_col']}. Format: {region_format_flag}")
    logging.debug(f"{log_prefix} Using 0-based indices: Header Row {header_row_idx}, Data Rows {data_start_row_idx}-{data_end_row_idx}, Cols {data_start_col_idx}-{data_end_col_idx}.")

    if header_row_idx < 0 or data_start_row_idx < 0: # Check for negative if original was 0 or less
        logging.error(f"{log_prefix} Invalid region definition: header_row or data_start_row is negative.")
        return pd.DataFrame(), pd.DataFrame(), [], region_format_flag

    # Extract raw header strings (initial attempt)
    raw_headers = []
    # Iterate over cells in the header row for the specified columns
    # openpyxl rows and columns are 1-indexed for iter_cols/iter_rows, but 0-indexed for direct cell access
    try:
        header_cells = list(sheet.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1,
                                            min_col=data_start_col_idx + 1, max_col=data_end_col_idx + 1))[0]
        for cell in header_cells:
            raw_headers.append(str(cell.value) if cell.value is not None else "")
    except IndexError:
        logging.warning(f"{log_prefix} Could not extract header cells for region. Header row index: {header_row_idx+1}, Col range: {data_start_col_idx+1}-{data_end_col_idx+1}")
        return pd.DataFrame(), pd.DataFrame(), [], region_format_flag
    
    logging.debug(f"{log_prefix} Raw headers extracted: {raw_headers}")

    # ------------------------------------------------------------------
    # Fix for Delivered-Media tables where the Market column ("Metrics / Market")
    # sits one column *left* of the detected region.
    # If we didn't capture a MARKET header but the cell immediately left of
    # start_col looks like a Market header, shift the whole region one column
    # left and re-extract headers & data.
    # ------------------------------------------------------------------

    if ('MARKET' not in [clean_column_name(h) for h in raw_headers]) and region['start_col'] > 1:
        left_cell_val = str(sheet.cell(header_row_idx + 1, data_start_col_idx).value or '').strip().upper()
        if any(key in left_cell_val for key in ['MARKET', 'METRICS/ MARKET', 'METRICS / MARKET']):
            logging.info(f"{log_prefix} Market column detected to the left—shifting region 1 column left.")
            # Update indices
            data_start_col_idx -= 1
            region['start_col'] -= 1
            # Re-extract headers with new column range
            raw_headers = []
            try:
                header_cells = list(sheet.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1,
                                                    min_col=data_start_col_idx + 1, max_col=data_end_col_idx + 1))[0]
                for cell in header_cells:
                    raw_headers.append(str(cell.value) if cell.value is not None else "")
            except Exception:
                pass
            logging.debug(f"{log_prefix} Raw headers re-extracted after shift: {raw_headers}")

    # Extract data rows
    data_rows = []
    if data_start_row_idx <= data_end_row_idx:
        for row_idx in range(data_start_row_idx, data_end_row_idx + 1):
            row_values = []
            try:
                current_row_cells = list(sheet.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1,
                                                       min_col=data_start_col_idx + 1, max_col=data_end_col_idx + 1))[0]
                for cell in current_row_cells:
                    row_values.append(cell.value)
                data_rows.append(row_values)
            except IndexError:
                logging.warning(f"{log_prefix} Could not extract full data row at sheet row index {row_idx+1}. This might indicate an irregular table end.")
                # If a partial row is problematic, one might choose to break or continue
                # For now, we log and continue, potentially resulting in shorter rows if any.
                if row_values: # if some cells were extracted before error
                    # Fill remaining cells with None to match header length if possible
                    missing_cols = len(raw_headers) - len(row_values)
                    if missing_cols > 0:
                        row_values.extend([None] * missing_cols)
                    data_rows.append(row_values)
                # else: skip this problematic row entirely if no cells were read

    if not data_rows and not raw_headers:
        logging.warning(f"{log_prefix} No data or headers found for the region.")
        return pd.DataFrame(), pd.DataFrame(), [], region_format_flag
    elif not data_rows:
        logging.warning(f"{log_prefix} Headers found but no data rows for the region.")
        # Create empty DataFrames with appropriate columns if headers exist
        df_raw_empty = pd.DataFrame(columns=raw_headers if raw_headers else None)
        df_mapped_empty_cols = [map_column_name(h)[0] if h else f"UNKNOWN_COL_{i}" for i, h in enumerate(raw_headers)]
        df_mapped_empty = pd.DataFrame(columns=df_mapped_empty_cols if df_mapped_empty_cols else None)
        return df_mapped_empty, df_raw_empty, raw_headers, region_format_flag

    # Create the raw DataFrame before header mapping
    # Ensure all data rows have the same length as raw_headers, padding with None if necessary
    expected_cols = len(raw_headers)
    processed_data_rows = []
    for r_idx, r_values in enumerate(data_rows):
        if len(r_values) < expected_cols:
            logging.debug(f"{log_prefix} Row {data_start_row_idx + r_idx +1} has fewer columns ({len(r_values)}) than headers ({expected_cols}). Padding with None.")
            r_values.extend([None] * (expected_cols - len(r_values)))
        elif len(r_values) > expected_cols:
            logging.debug(f"{log_prefix} Row {data_start_row_idx + r_idx +1} has more columns ({len(r_values)}) than headers ({expected_cols}). Truncating.")
            r_values = r_values[:expected_cols]
        processed_data_rows.append(r_values)

    try:
        df_raw_region = pd.DataFrame(processed_data_rows, columns=raw_headers)
    except ValueError as e:
        logging.error(f"{log_prefix} Error creating raw DataFrame: {e}. Header length: {len(raw_headers)}, Mismatched data row lengths might exist despite padding.")
        # Fallback: create with default integer columns if raw_headers are problematic
        df_raw_region = pd.DataFrame(processed_data_rows)
        # Attempt to still get raw_headers out for debugging even if df creation with them failed.
        # The caller (process_workbook) expects raw_headers list.

    # Create the mapped DataFrame
    # Map raw headers to standardized names
    mapped_headers = []
    unmapped_count = 0
    header_name_counts = {}
    if not raw_headers:
        # If no headers detected, generate generic column names based on the number of columns
        col_count = df_raw_region.shape[1] if not df_raw_region.empty else expected_cols
        for idx in range(col_count):
            base_name = f"COLUMN_{idx}"
            # ensure uniqueness (unlikely duplicates here but keep consistent)
            header_name_counts[base_name] = 1
            mapped_headers.append(base_name)
    else:
        for i, header_str in enumerate(raw_headers):
            clean_header = clean_column_name(str(header_str))
            mapped_name, confidence = map_column_name(clean_header)
            if mapped_name:
                base_name = mapped_name
            else:
                base_name = f"UNKNOWN_COL_{clean_header.replace(' ', '_')}_RAW_IDX{i}" if clean_header else f"UNKNOWN_COL_EMPTY_RAW_IDX{i}"
                unmapped_count += 1
            # Ensure uniqueness
            count = header_name_counts.get(base_name, 0)
            if count > 0:
                unique_name = f"{base_name}_DUP{count+1}"
            else:
                unique_name = base_name
            header_name_counts[base_name] = count + 1
            mapped_headers.append(unique_name)
    # ... existing code ...

    # Create DataFrame with standardized column names
    # Use processed_data_rows which are already length-normalized
    try:
        df_mapped_region = pd.DataFrame(processed_data_rows, columns=mapped_headers if mapped_headers else None)
    except ValueError as e:
        logging.error(f"{log_prefix} Error creating mapped DataFrame with mapped_headers: {e}. Mapped header length: {len(mapped_headers) if mapped_headers else 'None'}. Data shape: ({len(processed_data_rows)}, {len(processed_data_rows[0]) if processed_data_rows else 0})")
        # Fallback: create with default integer columns if mapped_headers are problematic
        df_mapped_region = pd.DataFrame(processed_data_rows)
        # Ensure raw_headers are still returned for consistency, even if df_mapped_region is compromised

    if unmapped_count > 0:
        logging.info(f"{log_prefix} {unmapped_count}/{len(raw_headers)} headers were not mapped to standard names for region.")

    # Basic cleaning of the mapped DataFrame values (e.g., stripping whitespace from strings)
    logging.debug(f"{log_prefix} Mapped headers before cleaning loop: {list(df_mapped_region.columns)}")
    for col in df_mapped_region.columns:
        logging.debug(f"{log_prefix} Processing column '{col}' for cleaning. Type of df_mapped_region[col]: {type(df_mapped_region[col])}")
        # Check if the column selector returns a Series, which should have a dtype
        if not isinstance(df_mapped_region[col], pd.Series):
            logging.error(f"{log_prefix} Column '{col}' is NOT a Series (type: {type(df_mapped_region[col])}). This is unexpected. Columns list: {list(df_mapped_region.columns)}")
            # Potentially skip this column or handle differently if it's a DataFrame due to duplicate names
            if isinstance(df_mapped_region[col], pd.DataFrame):
                logging.warning(f"{log_prefix} Column '{col}' selected as a DataFrame, likely due to duplicate column names. Skipping detailed cleaning for this one.")
                continue # Skip to next column if it's a DataFrame
            # If it's something else entirely, this is a bigger issue.

        if df_mapped_region[col].dtype == 'object':
            try:
                df_mapped_region[col] = df_mapped_region[col].astype(str).str.strip()
                # Replace various NA-like strings with pd.NA for consistent handling later
                na_like_values = ['None', '', 'nan', 'N/A', 'n/a', '#N/A', 'NA'] # Add more as needed
                df_mapped_region[col] = df_mapped_region[col].replace(na_like_values, pd.NA)
            except Exception as e:
                logging.warning(f"{log_prefix} Error during string cleaning for column {col}: {e}")

    logging.debug(f"{log_prefix} Extracted mapped DataFrame shape: {df_mapped_region.shape}, Raw DataFrame shape: {df_raw_region.shape}")
    return df_mapped_region, df_raw_region, raw_headers, region_format_flag

def process_workbook(input_file: str, output_file: str = None, source_type: str = None) -> Dict[str, Any]:
    """
    Main function to process a workbook file and extract data from different sheets.
    
    Args:
        input_file: Path to the input Excel file.
        output_file: Optional path to save the output Excel file.
        source_type: Optional classification of the source (e.g., "PLANNED", "DELIVERED")
        
    Returns:
        Dictionary with status and results information
    """
    # Structure for storing results from processing
    results = {
        'workbook_path': input_file,
        'sheet_data': {},
        'combined_data': None,
        'success': False,
        'error': None,
        'start_time': time.time(),
        'file_format': 'unknown',
        'errors': []  # Initialize errors list
    }
    
    combined_df: pd.DataFrame | None = None   # ensures name exists
    final_df_for_output: pd.DataFrame | None = None
    
    try:
        # Load workbook
        logging.info(f"Processing file: {input_file}")
        workbook = load_workbook(filename=input_file, data_only=True)
        sheet_names = workbook.sheetnames
        logging.info(f"Loaded workbook '{Path(input_file).name}' with sheets: {sheet_names}")
        
        # Detect file format
        file_format = detect_file_format(input_file, workbook)
        logging.info(f"Detected file format: {file_format}")
        results['file_format'] = file_format
        
        # Dictionary to store sheet DataFrames
        sheet_dfs = {}
        sheet_row_counts = {}
        
        # Process each sheet
        for sheet_name in sheet_names:
            try:
                # Check if this sheet corresponds to a known platform type
                normalized_name = get_normalized_sheet_name(sheet_name)
            
                # Skip sheets that don't match any known platform
                if normalized_name not in SHEETS_TO_PROCESS:
                    logging.info(f"Skipping sheet '{sheet_name}' as it doesn't match any known platform type")
                    continue
            
                logging.info(f"Processing sheet '{sheet_name}' (normalized: '{normalized_name}')")
            
                # Get the worksheet
                sheet = workbook[sheet_name]
            
                # ---- Optimised region scanning ----
                last_row_used, last_col_used = _determine_effective_bounds(sheet)
                logging.debug(f"[{sheet_name}] Effective bounds – last_row={last_row_used}, last_col={last_col_used}")

                # Pass input_file path, effective bounds and detected format
                table_regions = find_table_regions(
                    sheet,
                    sheet_name,
                    source_workbook_path=input_file,
                    effective_max_row=last_row_used,
                    effective_max_col=last_col_used,
                )
            
                # Process regions and construct DataFrame
                if table_regions:
                    all_region_dfs = []
                    region_dfs_info = []
                
                    for i, region in enumerate(table_regions):
                        logging.info(f"[{sheet_name}] Processing region {i}...")
                        try:
                            # Extract data from the region
                            detection_method = region.get('detection_method', 'unknown')
                            header_row = region.get('header_row', -1)
                        
                            # Check if region has specific format information
                            region_format = file_format  # Default to overall file format
                            if 'is_delivered_format' in region:
                                region_format = "delivered" if region['is_delivered_format'] else "planned"
                                logging.debug(f"[{sheet_name}] Region {i} has specific format flag: {region_format}")
                        
                            # Check if this region is a potential R&F table candidate
                            rf_table_candidate = region.get('rf_table_candidate', False)
                        
                            # Add region index to help with debugging
                            region['region_index'] = i
                        
                            # UPDATED: extract_data_to_dataframe now returns raw_df_region as the second item
                            df_region_data, raw_df_region, raw_headers_for_region, region_specific_format_flag = extract_data_to_dataframe(
                                sheet, normalized_name, region # Use normalized_name for sheet_name arg
                            )
                        
                            if df_region_data is not None and not df_region_data.empty:
                                # Add metadata columns to track source
                                # For R&F data, add suffix to Source_Sheet
                                if file_format == 'delivered' and 'PLATFORM' in df_region_data.columns:
                                    # Check if this is R&F data
                                    rf_mask = df_region_data['PLATFORM'].astype(str).str.contains(r'(Reach|Freq)', na=False, regex=True)
                                    df_region_data['Source_Sheet'] = normalized_name
                                    # Update Source_Sheet for R&F rows
                                    df_region_data.loc[rf_mask, 'Source_Sheet'] = f"{normalized_name}_RF"
                                    # Non-R&F rows get _MEDIA suffix
                                    df_region_data.loc[~rf_mask, 'Source_Sheet'] = f"{normalized_name}_MEDIA"
                                else:
                                    df_region_data['Source_Sheet'] = normalized_name # Store source sheet for PLANNED
                                    
                                df_region_data['REGION_INDEX'] = i
                                df_region_data['DETECTION_METHOD'] = detection_method
                                df_region_data['Source_File'] = os.path.basename(input_file)
                                
                                # Set Source_Type based on file format
                                if file_format == 'planned':
                                    df_region_data['Source_Type'] = "PLANNED"
                                elif file_format == 'delivered':
                                    # Default to DELIVERED MEDIA, will be updated later for R&F rows
                                    df_region_data['Source_Type'] = "DELIVERED MEDIA"
                            
                                # Always set/overwrite PLATFORM based on the normalized sheet name,
                                # as the sheet name is the ground truth for the platform.
                                df_region_data['PLATFORM'] = normalized_name
                            
                                # Track if R&F normalization was applied
                                rf_normalized = False
                            
                                # Apply Reach & Frequency table normalization if needed
                                # Determine the effective file format for R&F logic for this specific region
                                # This considers overall file_format and any region_specific_format_flag
                                effective_rf_logic_format = region_specific_format_flag if region_specific_format_flag and region_specific_format_flag.endswith('_rf') else file_format
                            
                                # Define when to attempt R&F normalization more clearly
                                should_attempt_rf = (
                                    effective_rf_logic_format == "delivered" or 
                                    rf_table_candidate
                                )
                            
                                if should_attempt_rf:
                                    logging.debug(f"[{normalized_name} Reg:{i}] Attempting R&F normalization. Effective R&F Logic Format: {effective_rf_logic_format}, RF candidate: {rf_table_candidate}")
                                    source_file_name = os.path.basename(input_file)
                                    original_first_col_name = str(raw_headers_for_region[0]) if raw_headers_for_region and raw_headers_for_region[0] else "UnknownFirstCol"
                                
                                    # Add targeted diagnostic logging for Region 0 specifically
                                    if i == 0:
                                        _count_campaign(df_region_data, f"{normalized_name}-Reg0-preRF")
                                
                                    try:
                                        # Call normalize_rf_table with both df_region_data (mapped) and raw_df_region (raw)
                                        df_processed_after_rf = normalize_rf_table(
                                            df_mapped=df_region_data, 
                                            df_raw=raw_df_region, 
                                            raw_first_column_name=original_first_col_name, 
                                            sheet_name=normalized_name, 
                                            source_file=source_file_name, 
                                            source_region_idx=i, 
                                            detection_method=detection_method, 
                                            file_format_for_rf_logic=effective_rf_logic_format
                                        )
                                    
                                        # Add targeted diagnostic logging for Region 0 specifically  
                                        if i == 0:
                                            _count_campaign(df_processed_after_rf, f"{normalized_name}-Reg0-postRF")
                                    
                                        # normalize_rf_table now returns the df_to_use (either processed R&F or original mapped)
                                        if df_processed_after_rf is not None and not df_processed_after_rf.equals(df_region_data):
                                            # This means R&F processing likely occurred and changed the DataFrame
                                            logging.info(f"[{normalized_name} Reg:{i}] R&F normalization applied. New shape: {df_processed_after_rf.shape}")
                                            df_region_data = df_processed_after_rf # Update df_region_data
                                            rf_normalized = True
                                            
                                            # Update Source_Type for R&F rows (they should already have it, but ensure consistency)
                                            if file_format == 'delivered' and 'PLATFORM' in df_region_data.columns:
                                                rf_mask = df_region_data['PLATFORM'].astype(str).str.contains(r'(Reach|Freq)', na=False, regex=True)
                                                df_region_data.loc[rf_mask, 'Source_Type'] = "DELIVERED R&F"
                                        elif df_processed_after_rf is df_region_data:
                                            logging.info(f"[{normalized_name} Reg:{i}] R&F normalization attempted but no changes made (returned original mapped df). Original shape: {df_region_data.shape}")
                                        elif df_processed_after_rf is None:
                                             logging.warning(f"[{normalized_name} Reg:{i}] R&F normalization returned None. Original mapped data (shape {df_region_data.shape}) will be used.")
                                        # No specific 'else' needed if it equals, df_region_data is already correct.
                                    except Exception as rf_error:
                                        # Log error but continue processing without normalization
                                        logging.error(f"[{normalized_name} Reg:{i}] Error in R&F normalization: {str(rf_error)}")
                                        logging.debug(traceback.format_exc())
                            
                                # Store DataFrame and information for later
                                all_region_dfs.append(df_region_data)
                                region_dfs_info.append({
                                    'region_index': i,
                                    'detection_method': detection_method,
                                    'header_row': header_row,
                                    'row_count': len(df_region_data),
                                    'rf_normalized': rf_normalized
                                })
                            else:
                                logging.warning(f"[{sheet_name} Reg:{i} M:{detection_method}] No valid data extracted.")
                        except Exception as e:
                            logging.error(f"Error processing region {i} in sheet '{sheet_name}': {str(e)}")
                            logging.error(traceback.format_exc())
                            results['errors'].append(f"Region {i} in '{sheet_name}': {str(e)}")
                
                    # Combine all region DataFrames for this sheet
                    if all_region_dfs:
                        try:
                            # Concatenate all region DataFrames
                            sheet_df = pd.concat(all_region_dfs, ignore_index=True)
                            logging.info(f"[{sheet_name}] Concatenated {len(all_region_dfs)} region(s) into sheet DataFrame with shape {sheet_df.shape}")
                        
                            # Add campaign count after region combination
                            _count_campaign(sheet_df, f"{normalized_name}-post_regions")
                        
                            # Tag every row that looks like Reach/Freq so later steps can respect it
                            if 'PLATFORM' in sheet_df.columns:
                                sheet_df["IS_RF"] = sheet_df["PLATFORM"].astype(str).str.contains(
                                    r"(Reach|Freq)", case=False, na=False
                                )
                                # Debug: log R&F detection results
                                rf_count = sheet_df["IS_RF"].sum()
                                logging.info(f"[{normalized_name}] RF Detection: Found {rf_count} R&F rows out of {len(sheet_df)} total rows")
                                if rf_count > 0:
                                    rf_platforms = sheet_df[sheet_df["IS_RF"]]["PLATFORM"].unique()
                                    logging.info(f"[{normalized_name}] RF Platforms detected: {rf_platforms}")
                            else:
                                sheet_df["IS_RF"] = False
                        
                            # Convert numeric columns to appropriate types
                            for col in NUMERIC_COLUMNS:
                                if col in sheet_df.columns:
                                    sheet_df[col] = pd.to_numeric(sheet_df[col], errors='coerce')
                        
                            # Standardize platform values
                            if 'PLATFORM' in sheet_df.columns:
                                sheet_df['PLATFORM'] = sheet_df['PLATFORM'].apply(standardize_platform_name)
                        
                            # Remove duplicates more carefully to preserve R&F data
                            # First, identify columns to use for deduplication (exclude value columns that might differ)
                            # dedup_cols = [col for col in sheet_df.columns if col not in ['UNIQUES_REACH', 'FREQUENCY', 'IMPRESSIONS', 'VIDEO_VIEWS', 'CLICKS_ACTIONS', 'BUDGET_LOCAL']]
                        
                            # Keep track of original row count for logging
                            original_count = len(sheet_df)
                        
                            # ------------------------------------------------------------------
                            # R&F rows must **never** deduplicate: Reach and Freq are distinct
                            # rows that share all identity columns except the numeric fields.
                            # We split them out first, dedup MEDIA rows only, then concat.
                            # ------------------------------------------------------------------
                        
                            # NEW split-dedup logic that preserves all RF rows
                            rf_df = sheet_df[sheet_df["IS_RF"]].copy()       # keep every Reach/Freq
                            non_rf_df = sheet_df[~sheet_df["IS_RF"]].copy() # Non-R&F rows for deduplication
                        
                            # *** CRITICAL FIX: Preserve DELIVERED MEDIA granularity at sheet level ***
                            if len(non_rf_df) > 0:
                                # Split non-R&F data into DELIVERED MEDIA and other data
                                delivered_media_mask = non_rf_df['Source_Type'] == 'DELIVERED MEDIA' if 'Source_Type' in non_rf_df.columns else pd.Series([False] * len(non_rf_df))
                                delivered_media_df = non_rf_df[delivered_media_mask].copy()  # Keep all DELIVERED MEDIA entries
                                other_data_df = non_rf_df[~delivered_media_mask].copy()      # Other data for deduplication
                                
                                logging.info(f"[{sheet_name}] SHEET-DEDUP: DeliveredMedia={len(delivered_media_df)}, Other={len(other_data_df)}")
                                
                                # CRITICAL FIX: Don't deduplicate PLANNED data at sheet level
                                # PLANNED data legitimately has multiple rows per market with different campaign details
                                if len(other_data_df) > 0:
                                    # Check if this is PLANNED data by looking at Source_Type
                                    is_planned_data = (
                                        (other_data_df['Source_Type'] == 'PLANNED').any() if 'Source_Type' in other_data_df.columns else False
                                    )
                                    
                                    if is_planned_data:
                                        # Preserve ALL PLANNED data - no deduplication
                                        logging.info(f"[{sheet_name}] PLANNED data detected: Preserving all {len(other_data_df)} rows (no deduplication)")
                                    else:
                                        # Apply deduplication for non-PLANNED data
                                        dedup_key = ["Source_File", "Source_Sheet", "MARKET", "PLATFORM"]
                                        other_data_df = other_data_df.drop_duplicates(subset=dedup_key, keep="first")
                                        logging.info(f"[{sheet_name}] After other data dedup: {len(other_data_df)} rows")
                                
                                # Recombine: All DELIVERED MEDIA + Deduplicated other data
                                non_rf_df = pd.concat([delivered_media_df, other_data_df], ignore_index=True)
                        
                            # Recombine R&F and deduplicated non-R&F data
                            sheet_df = pd.concat([rf_df, non_rf_df], ignore_index=True)
                        
                            # Optional: drop helper column to avoid polluting output
                            sheet_df.drop(columns=["IS_RF"], inplace=True)
                        
                            deduped_count = len(sheet_df)
                            if deduped_count < original_count:
                                logging.info(f"[{sheet_name}] Removed {original_count - deduped_count} duplicate non-R&F rows (preserved all R&F data)")
                        
                            # Add campaign count after deduplication
                            _count_campaign(sheet_df, f"{normalized_name}-post_dedup")
                        
                            # Defensive assertions to catch any future R&F row loss
                            if _rf_expected_for_sheet(sheet_df, file_format, normalized_name):
                                assert sheet_df["PLATFORM"].str.contains("Campaign Reach",
                                                                         case=False,
                                                                         na=False).any(), \
                                    f"{normalized_name}: Campaign Reach rows lost after sheet dedup"
                                assert sheet_df["PLATFORM"].str.contains("Campaign Freq",
                                                                         case=False,
                                                                         na=False).any(), \
                                    f"{normalized_name}: Campaign Freq rows lost after sheet dedup"
                            else:
                                logging.debug(f"[{normalized_name}] R&F assertions skipped "
                                              f"(not expected for {file_format} files)")
                        
                            # Store the DataFrame
                            sheet_dfs[sheet_name] = sheet_df
                            sheet_row_counts[sheet_name] = len(sheet_df)
                        
                            # Count RF-normalized regions
                            rf_normalized_count = sum(1 for info in region_dfs_info if info.get('rf_normalized', False))
                        
                            # Store results
                            results['sheet_data'][sheet_name] = {
                                'regions_found': len(table_regions),
                                'regions_processed': len(region_dfs_info),
                                'row_count': sheet_row_counts[sheet_name],
                                'rf_normalized_regions': rf_normalized_count,
                                'region_info': region_dfs_info
                            }
                        except Exception as e:
                            logging.error(f"Error combining region DataFrames for sheet '{sheet_name}': {str(e)}")
                            logging.error(traceback.format_exc())
                            results['errors'].append(f"Combining regions in '{sheet_name}': {str(e)}")
                else:
                    logging.warning(f"No valid table regions found in sheet '{sheet_name}'.")
                    results['sheet_data'][sheet_name] = {
                        'regions_found': 0,
                        'regions_processed': 0,
                        'row_count': 0,
                        'region_info': []
                    }
        
            except Exception as sheet_err:
                logging.error(f"[{sheet_name}] Fatal sheet error: {sheet_err}")
                logging.debug(traceback.format_exc())
                results['errors'].append(f"{sheet_name}: {sheet_err}")
                continue
        # Combine all sheet DataFrames into a single DataFrame
        if sheet_dfs:
            try:
                # List of DataFrames to combine
                dfs_to_combine = list(sheet_dfs.values())
                logging.info(f"Attempting to concatenate {len(dfs_to_combine)} sheet DataFrames.")
                
                # Add data flow tracing before concatenation
                for i, df in enumerate(dfs_to_combine):
                    media_count = len(df[df['Source_Type'] == 'DELIVERED MEDIA']) if 'Source_Type' in df.columns else 0
                    rf_count = len(df[df['Source_Type'] == 'DELIVERED R&F']) if 'Source_Type' in df.columns else 0
                    logging.info(f"[TRACE-PRE-CONCAT] Sheet {i}: Total={len(df)}, Media={media_count}, R&F={rf_count}")
                
                # Concatenate all sheets
                combined_df = pd.concat(dfs_to_combine, ignore_index=True)
                
                # Add data flow tracing after concatenation
                media_count = len(combined_df[combined_df['Source_Type'] == 'DELIVERED MEDIA']) if 'Source_Type' in combined_df.columns else 0
                rf_count = len(combined_df[combined_df['Source_Type'] == 'DELIVERED R&F']) if 'Source_Type' in combined_df.columns else 0
                logging.info(f"[TRACE-POST-CONCAT] Total={len(combined_df)}, Media={media_count}, R&F={rf_count}")
                
                # Fix Source_Sheet for R&F data to add _RF suffix
                if 'Source_Type' in combined_df.columns and 'Source_Sheet' in combined_df.columns:
                    rf_mask = combined_df['Source_Type'] == 'DELIVERED R&F'
                    for sheet_name in ['DV360', 'META', 'TIKTOK']:
                        sheet_rf_mask = rf_mask & (combined_df['Source_Sheet'] == sheet_name)
                        if sheet_rf_mask.any():
                            combined_df.loc[sheet_rf_mask, 'Source_Sheet'] = f"{sheet_name}_RF"
                            logging.info(f"Updated Source_Sheet to {sheet_name}_RF for {sheet_rf_mask.sum()} R&F rows")
                
                logging.info(f"Successfully concatenated all sheet data. Final shape: {combined_df.shape}")
                
                # Ensure all standard columns exist (adding NA for missing ones)
                for std_col in STANDARDIZED_COLUMNS:
                    if std_col not in combined_df.columns:
                        combined_df[std_col] = pd.NA
                
                # Drop legacy SHEET_NAME column if still present (all rows now use Source_Sheet)
                if 'SHEET_NAME' in combined_df.columns:
                    combined_df.drop(columns=['SHEET_NAME'], inplace=True)
                
                # Standardize platform values using our enhanced standardization function
                if 'PLATFORM' in combined_df.columns:
                    # Add diagnostic logging before and after the PLATFORM overwrite logic
                    _count_campaign(combined_df, "before_platform_standardization")
                    campaign_platform_before = combined_df[combined_df['PLATFORM'].astype(str).str.contains('Campaign', case=False, na=False)]['PLATFORM'].unique()
                    logging.info(f"[PLATFORM DEBUG] Campaign platforms before standardization: {campaign_platform_before}")
                    
                    combined_df['PLATFORM'] = combined_df['PLATFORM'].apply(standardize_platform_name)
                    
                    # Add diagnostic logging after the PLATFORM overwrite logic
                    _count_campaign(combined_df, "after_platform_standardization")
                    campaign_platform_after = combined_df[combined_df['PLATFORM'].astype(str).str.contains('Campaign', case=False, na=False)]['PLATFORM'].unique()
                    logging.info(f"[PLATFORM DEBUG] Campaign platforms after standardization: {campaign_platform_after}")
                    
                    # Check for rows with empty platform and try to infer from sheet name
                    if combined_df['PLATFORM'].isna().any() and 'Source_Sheet' in combined_df.columns:
                        # Add diagnostic logging for the overwrite operation
                        na_mask = combined_df['PLATFORM'].isna()
                        na_count = na_mask.sum()
                        logging.info(f"[PLATFORM DEBUG] Found {na_count} rows with NaN PLATFORM values")
                        
                        # Check if any Campaign metrics are about to be overwritten
                        campaign_na_rows = combined_df[na_mask & combined_df.index.isin(
                            combined_df[combined_df.astype(str).apply(lambda row: 'Campaign' in str(row).upper(), axis=1)].index
                        )]
                        if len(campaign_na_rows) > 0:
                            logging.warning(f"[PLATFORM DEBUG] WARNING: {len(campaign_na_rows)} Campaign metric rows have NaN PLATFORM and will be overwritten!")
                            for idx, row in campaign_na_rows.iterrows():
                                logging.warning(f"[PLATFORM DEBUG] Row {idx}: Will overwrite NaN with '{row.get('Source_Sheet', 'UNKNOWN')}'")
                        
                        mask = combined_df['PLATFORM'].isna()
                        combined_df.loc[mask, 'PLATFORM'] = combined_df.loc[mask, 'Source_Sheet'].apply(get_normalized_sheet_name)
                        
                        # Add diagnostic logging after the overwrite
                        _count_campaign(combined_df, "after_platform_overwrite")
                
                # Convert numeric columns with enhanced error handling for budget values
                for col in NUMERIC_COLUMNS:
                    if col in combined_df.columns:
                        if col == 'BUDGET_LOCAL':
                            # Special handling for budget - clean string values first
                            # Convert to string and clean currency symbols
                            combined_df[col] = combined_df[col].astype(str).str.replace(',', '').str.replace('$', '').str.strip()
                            # Replace dash-only values with NaN (these represent missing budget data)
                            combined_df[col] = combined_df[col].replace(['-', 'nan', 'None', ''], pd.NA)
                            # Convert to numeric, replacing any remaining non-numeric values with NaN
                            combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce')
                        elif col in ['UNIQUES_REACH', 'FREQUENCY']:
                            # Special handling for R&F values - preserve actual numbers, clean invalid values
                            before_count = combined_df[col].notna().sum()
                            logging.info(f"[NUMERIC DEBUG] {col} before conversion: {before_count} non-null values, dtype={combined_df[col].dtype}")
                            # Convert to string first to handle mixed types, then clean
                            combined_df[col] = combined_df[col].astype(str).str.replace(',', '').str.strip()
                            # Replace only clearly invalid values with NaN, preserve actual numbers
                            combined_df[col] = combined_df[col].replace(['-', 'nan', 'None', '', 'N/A', 'n/a'], pd.NA)
                            # Convert to numeric
                            combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce')
                            after_count = combined_df[col].notna().sum()
                            logging.info(f"[NUMERIC DEBUG] {col} after conversion: {after_count} non-null values, dtype={combined_df[col].dtype}")
                        else:
                            combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce')
                
                # Calculate derived metrics to ensure accuracy
                logging.info("Calculating derived metrics (CPM, CPC, CPV, CTR, VTR)...")
                
                # Fix CTR calculations - convert from decimal to percentage
                if 'CTR_PERCENT' in combined_df.columns:
                    # If CTR is in decimal format (0.0003), convert to percentage (0.03%)
                    ctr_mask = combined_df['CTR_PERCENT'].notna() & (combined_df['CTR_PERCENT'] < 1)
                    combined_df.loc[ctr_mask, 'CTR_PERCENT'] = combined_df.loc[ctr_mask, 'CTR_PERCENT'] * 100
                    logging.info(f"Converted {ctr_mask.sum()} CTR values from decimal to percentage format")
                
                # Fix VTR calculations - convert from decimal to percentage  
                if 'VTR_PERCENT' in combined_df.columns:
                    # Convert VTR column to numeric first
                    combined_df['VTR_PERCENT'] = pd.to_numeric(combined_df['VTR_PERCENT'], errors='coerce')
                    # If VTR is in decimal format (0.0225), convert to percentage (2.25%)
                    vtr_mask = combined_df['VTR_PERCENT'].notna() & (combined_df['VTR_PERCENT'] < 1)
                    combined_df.loc[vtr_mask, 'VTR_PERCENT'] = combined_df.loc[vtr_mask, 'VTR_PERCENT'] * 100
                    logging.info(f"Converted {vtr_mask.sum()} VTR values from decimal to percentage format")
                
                # Recalculate CPC to ensure accuracy: CPC = Budget / Clicks
                if all(col in combined_df.columns for col in ['BUDGET_LOCAL', 'CLICKS_ACTIONS', 'CPC_LOCAL']):
                    # Calculate CPC where we have both budget and clicks
                    calc_mask = (combined_df['BUDGET_LOCAL'].notna() & 
                               combined_df['CLICKS_ACTIONS'].notna() & 
                               (combined_df['CLICKS_ACTIONS'] > 0))
                    
                    if calc_mask.any():
                        calculated_cpc = combined_df.loc[calc_mask, 'BUDGET_LOCAL'] / combined_df.loc[calc_mask, 'CLICKS_ACTIONS']
                        combined_df.loc[calc_mask, 'CPC_LOCAL'] = calculated_cpc
                        logging.info(f"Recalculated CPC for {calc_mask.sum()} rows using Budget/Clicks formula")
                
                # Recalculate CPM: CPM = (Budget / Impressions) * 1000
                if all(col in combined_df.columns for col in ['BUDGET_LOCAL', 'IMPRESSIONS', 'CPM_LOCAL']):
                    calc_mask = (combined_df['BUDGET_LOCAL'].notna() & 
                               combined_df['IMPRESSIONS'].notna() & 
                               (combined_df['IMPRESSIONS'] > 0))
                    
                    if calc_mask.any():
                        calculated_cpm = (combined_df.loc[calc_mask, 'BUDGET_LOCAL'] / combined_df.loc[calc_mask, 'IMPRESSIONS']) * 1000
                        combined_df.loc[calc_mask, 'CPM_LOCAL'] = calculated_cpm
                        logging.info(f"Recalculated CPM for {calc_mask.sum()} rows using (Budget/Impressions)*1000 formula")
                
                # Recalculate CPV: CPV = Budget / Video Views
                if all(col in combined_df.columns for col in ['BUDGET_LOCAL', 'VIDEO_VIEWS', 'CPV_LOCAL']):
                    calc_mask = (combined_df['BUDGET_LOCAL'].notna() & 
                               combined_df['VIDEO_VIEWS'].notna() & 
                               (combined_df['VIDEO_VIEWS'] > 0))
                    
                    if calc_mask.any():
                        calculated_cpv = combined_df.loc[calc_mask, 'BUDGET_LOCAL'] / combined_df.loc[calc_mask, 'VIDEO_VIEWS']
                        combined_df.loc[calc_mask, 'CPV_LOCAL'] = calculated_cpv
                        logging.info(f"Recalculated CPV for {calc_mask.sum()} rows using Budget/VideoViews formula")
                
                # Format date columns to DD/MM/YY format
                date_columns = ["START_DATE", "END_DATE"]
                for date_col in date_columns:
                    if date_col in combined_df.columns:
                        try:
                            # Convert to datetime if not already
                            combined_df[date_col] = pd.to_datetime(combined_df[date_col], errors='coerce')
                            
                            # Format to DD/MM/YY, only for non-NaT values
                            valid_dates_mask = combined_df[date_col].notna()
                            combined_df.loc[valid_dates_mask, date_col] = combined_df.loc[valid_dates_mask, date_col].dt.strftime('%d/%m/%y')
                            
                            # Replace NaT values with pd.NA
                            combined_df.loc[~valid_dates_mask, date_col] = pd.NA
                            
                            # Log the date formatting operation
                            logging.info(f"Formatted {date_col} to DD/MM/YY format")
                        except Exception as e:
                            logging.warning(f"Error formatting {date_col}: {e}")
                            # Don't fail on date formatting errors - just keep the original values
                
                # Remove duplicate rows - but preserve R&F data and DELIVERED MEDIA granularity
                original_row_count = len(combined_df)
                
                # Add tracing before deduplication
                media_count_before = len(combined_df[combined_df['Source_Type'] == 'DELIVERED MEDIA']) if 'Source_Type' in combined_df.columns else 0
                rf_count_before = len(combined_df[combined_df['Source_Type'] == 'DELIVERED R&F']) if 'Source_Type' in combined_df.columns else 0
                logging.info(f"[TRACE-PRE-DEDUP] Total={len(combined_df)}, Media={media_count_before}, R&F={rf_count_before}")
                
                # Identify R&F rows by PLATFORM column containing Reach or Freq
                if 'PLATFORM' in combined_df.columns:
                    rf_mask = combined_df['PLATFORM'].str.contains(
                        r"(Reach|Freq)", case=False, na=False
                    )
                    
                    rf_df = combined_df[rf_mask].copy()      # Keep all R&F rows
                    non_rf_df = combined_df[~rf_mask].copy() # Non-R&F rows for deduplication
                    
                    # *** CRITICAL FIX: Preserve DELIVERED MEDIA granularity ***
                    # Only deduplicate PLANNED data, preserve all DELIVERED MEDIA entries
                    if len(non_rf_df) > 0:
                        delivered_media_mask = non_rf_df['Source_Type'] == 'DELIVERED MEDIA' if 'Source_Type' in non_rf_df.columns else pd.Series([False] * len(non_rf_df))
                        delivered_media_df = non_rf_df[delivered_media_mask].copy()  # Keep all DELIVERED MEDIA
                        planned_df = non_rf_df[~delivered_media_mask].copy()         # Deduplicate PLANNED only
                        
                        logging.info(f"[TRACE-DEDUP-SPLIT] DeliveredMedia={len(delivered_media_df)}, Planned={len(planned_df)}")
                        
                        # Deduplicate only PLANNED data using specific key columns
                        if len(planned_df) > 0:
                            logging.info(f"[TRACE-DEDUP-PLANNED] Before dedup: {len(planned_df)} planned rows")
                            
                            # Show sample data to understand duplication patterns
                            if 'Source_Sheet' in planned_df.columns and 'MARKET' in planned_df.columns:
                                sheet_counts = planned_df['Source_Sheet'].value_counts()
                                logging.info(f"[TRACE-DEDUP-PLANNED] Pre-dedup sheet counts: {dict(sheet_counts)}")
                                
                                # Show first few rows per sheet
                                for sheet in sheet_counts.index:
                                    sheet_data = planned_df[planned_df['Source_Sheet'] == sheet]
                                    logging.info(f"[TRACE-DEDUP-PLANNED] {sheet} sample data:")
                                    for i, (idx, row) in enumerate(sheet_data.head(3).iterrows()):
                                        market = row.get('MARKET', 'N/A')
                                        platform = row.get('PLATFORM', 'N/A')
                                        logging.info(f"[TRACE-DEDUP-PLANNED]   Row {i+1}: Market='{market}', Platform='{platform}'")
                            
                            # CRITICAL FIX: For PLANNED data, don't deduplicate across sheets - only within sheets
                            # PLANNED data legitimately has same markets across different sheets (DV360, META, TIKTOK)
                            # so we should NOT treat them as duplicates
                            logging.info(f"[TRACE-DEDUP-PLANNED] PRESERVING ALL PLANNED DATA - No cross-sheet deduplication")
                            logging.info(f"[TRACE-DEDUP-PLANNED] Final count: {len(planned_df)} planned rows")
                            
                            # Show final sheet counts
                            if 'Source_Sheet' in planned_df.columns:
                                final_sheet_counts = planned_df['Source_Sheet'].value_counts()
                                logging.info(f"[TRACE-DEDUP-PLANNED] Final sheet counts: {dict(final_sheet_counts)}")
                        
                        # Recombine: R&F + All DELIVERED MEDIA + Deduplicated PLANNED
                        non_rf_df = pd.concat([delivered_media_df, planned_df], ignore_index=True)
                    
                    # Recombine R&F and processed non-R&F data
                    combined_df = pd.concat([rf_df, non_rf_df], ignore_index=True)
                else:
                    # Fallback: if no PLATFORM column, use original approach
                    combined_df = combined_df.drop_duplicates()
                
                deduped_row_count = len(combined_df)
                
                # Add tracing after deduplication
                media_count_after = len(combined_df[combined_df['Source_Type'] == 'DELIVERED MEDIA']) if 'Source_Type' in combined_df.columns else 0
                rf_count_after = len(combined_df[combined_df['Source_Type'] == 'DELIVERED R&F']) if 'Source_Type' in combined_df.columns else 0
                logging.info(f"[TRACE-POST-DEDUP] Total={len(combined_df)}, Media={media_count_after}, R&F={rf_count_after}")
                
                if deduped_row_count < original_row_count:
                    logging.info(f"Removed {original_row_count - deduped_row_count} duplicate rows")
                
                # Add campaign count after final deduplication
                _count_campaign(combined_df, f"Final-Combined")
                
                # Save to Excel if output file specified
                if output_file:
                    try:
                        # Create output directory if it doesn't exist
                        output_dir = os.path.dirname(output_file)
                        if output_dir and not os.path.exists(output_dir):
                            os.makedirs(output_dir)
                        
                        # Save to Excel using the strictly column-controlled DataFrame
                        final_df_for_output = pd.DataFrame()
                        for col in FINAL_OUTPUT_COLUMNS:
                            if col in combined_df.columns:
                                final_df_for_output[col] = combined_df[col]
                            else:
                                final_df_for_output[col] = pd.NA
                        
                        # Final numeric conversion - ensure all numeric columns are truly numeric before Excel save
                        for col in NUMERIC_COLUMNS:
                            if col in final_df_for_output.columns:
                                # Replace any remaining "-" strings with NaN, then convert to numeric
                                final_df_for_output[col] = final_df_for_output[col].replace(['-', 'nan', 'None', '', 'N/A', 'n/a'], pd.NA)
                                final_df_for_output[col] = pd.to_numeric(final_df_for_output[col], errors='coerce')
                                logging.info(f"[FINAL NUMERIC] {col}: {final_df_for_output[col].notna().sum()} non-null values, dtype={final_df_for_output[col].dtype}")
                        
                        # Apply consistent date formatting to final output (match individual file format)
                        date_columns = ["START_DATE", "END_DATE"]
                        for date_col in date_columns:
                            if date_col in final_df_for_output.columns:
                                try:
                                    def _fmt_date_consistent(val):
                                        if isinstance(val, (pd.Timestamp, pd._libs.tslibs.timestamps.Timestamp)):
                                            return "'" + val.strftime('%d/%m/%y')
                                        if isinstance(val, str):
                                            # If looks like ISO datetime, trim time part
                                            if '00:00:00' in val or '-' in val:
                                                try:
                                                    parsed = pd.to_datetime(val)
                                                    return "'" + parsed.strftime('%d/%m/%y')
                                                except Exception:
                                                    pass
                                            # If already has quote prefix, don't add another
                                            if val.startswith("'"):
                                                return val
                                            # If it's already in dd/mm/yy format, add quote prefix
                                            if len(val) == 8 and val[2] == '/' and val[5] == '/':
                                                return "'" + val
                                            return val
                                        return val
                                    
                                    final_df_for_output[date_col] = final_df_for_output[date_col].apply(_fmt_date_consistent)
                                    logging.info(f"Applied consistent date formatting to {date_col} in final output")
                                except Exception as e:
                                    logging.warning(f"Error formatting {date_col} in final output: {e}")
                        
                        # Add campaign count before final save
                        _count_campaign(final_df_for_output, "Pre-Save-Final")
                        
                        # Save with explicit date column formatting as text
                        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                            final_df_for_output.to_excel(writer, index=False, sheet_name='Sheet1')
                            worksheet = writer.sheets['Sheet1']
                            
                            # Format date columns as text to prevent Excel auto-conversion
                            from openpyxl.utils import get_column_letter
                            date_columns = ["START_DATE", "END_DATE"]
                            for date_col in date_columns:
                                if date_col in final_df_for_output.columns:
                                    col_idx = final_df_for_output.columns.get_loc(date_col) + 1  # Excel columns are 1-based
                                    col_letter = get_column_letter(col_idx)
                                    for row in range(2, len(final_df_for_output) + 2):  # Start from row 2 (after header)
                                        cell = worksheet[f'{col_letter}{row}']
                                        cell.number_format = '@'  # Text format
                        
                        logging.info(f"Saved combined data to '{output_file}' with {len(final_df_for_output)} rows and {len(final_df_for_output.columns)} columns.")
                        results['output_file'] = output_file
                    except Exception as e:
                        logging.error(f"Error saving results to '{output_file}': {str(e)}")
                        logging.error(traceback.format_exc())
                        results['errors'].append(f"Error saving to '{output_file}': {str(e)}")
                
                # Post-format numeric cells to "#,##0.00"
                try:
                    from openpyxl import load_workbook as _load_wb
                    _wb = _load_wb(output_file)
                    _ws = _wb.active
                    headers = [c.value for c in _ws[1]]
                    numeric_cols_idx = [idx+1 for idx, h in enumerate(headers) if h in NUMERIC_COLUMNS]
                    for c in numeric_cols_idx:
                        for r in range(2, _ws.max_row+1):
                            cell = _ws.cell(row=r, column=c)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                    _wb.save(output_file)
                except Exception as _fmt_err:
                    logging.warning(f"Numeric formatting skipped: {_fmt_err}")
                
            except Exception as e:
                logging.error(f"Error combining sheet DataFrames: {str(e)}")
                logging.error(traceback.format_exc())
                results['errors'].append(f"Error combining sheets: {str(e)}")
        
        # Store the combined DataFrame
        if combined_df is None:
            logging.warning("No sheet data collected – skipping final combine/save.")
            return results
            
        results['combined_data'] = combined_df

        # ------------------------------------------------------------------
        # Apply specific market name mapping for TikTok Delivered Media
        # ------------------------------------------------------------------
        if 'MARKET' in combined_df.columns and 'PLATFORM' in combined_df.columns and 'Source_Type' in combined_df.columns:
            market_mapping_tiktok_delivered = TIKTOK_MARKET_MAPPING
            
            # Ensure PLATFORM and Source_Type are strings for reliable comparison, handling potential NAs
            platform_str = combined_df['PLATFORM'].astype(str).str.lower()
            source_type_str = combined_df['Source_Type'].astype(str).str.lower()

            tiktok_delivered_mask = (
                platform_str.str.contains('tiktok', na=False) &
                source_type_str.str.contains('delivered', na=False)
            )
            
            if tiktok_delivered_mask.any():
                # Apply mapping only to relevant rows and existing market values
                # Store original market values for logging
                original_markets = combined_df.loc[tiktok_delivered_mask, 'MARKET'].copy()
                
                # Map values, keeping original if not in mapping
                combined_df.loc[tiktok_delivered_mask, 'MARKET'] = combined_df.loc[tiktok_delivered_mask, 'MARKET'].replace(market_mapping_tiktok_delivered)
                
                # Log changes
                changed_markets_mask = original_markets != combined_df.loc[tiktok_delivered_mask, 'MARKET']
                num_changed = changed_markets_mask.sum()
                if num_changed > 0:
                    logging.info(f"Applied market name mapping for {num_changed} TikTok 'DELIVERED MEDIA' rows.")
                    # For detailed logging, uncomment below:
                    # for original_val, new_val in zip(original_markets[changed_markets_mask], combined_df.loc[tiktok_delivered_mask, 'MARKET'][changed_markets_mask]):
                    #     logging.debug(f"Market changed: '{original_val}' -> '{new_val}'")
                else:
                    logging.info("TikTok 'DELIVERED MEDIA' market mapping: No applicable market names found for replacement or already standard.")
            else:
                logging.info("No TikTok 'DELIVERED MEDIA' rows found for market name mapping.")
        else:
            logging.warning("Market, Platform, or Source_Type column not found. Skipping TikTok market name mapping.")

        # ------------------------------------------------------------------
        # Final cleanup: replace all remaining pd.NA/NaN or empty strings with '-'
        # This is applied just before selecting FINAL_OUTPUT_COLUMNS
        # ------------------------------------------------------------------
        placeholder_map = {pd.NA: '-', np.nan: '-', '': '-', 'N/A': '-', 'n/a': '-', 'N/a': '-', 'n.a': '-', 'N.A': '-'}
        combined_df = combined_df.replace(placeholder_map).infer_objects(copy=False)

        # Ensure all FINAL_OUTPUT_COLUMNS are present, adding NA if missing
        # And select only these columns in the specified order for the final output file.
        # Also ensure the Source_Type column is included in the output
        final_df_for_output = pd.DataFrame()

        # Create list of columns including Source_Type for output
        output_columns = list(FINAL_OUTPUT_COLUMNS)
        if 'Source_Type' not in output_columns:
            output_columns.append('Source_Type')
        
        for col in output_columns:
            if col in combined_df.columns:
                final_df_for_output[col] = combined_df[col]
            else:
                final_df_for_output[col] = pd.NA # Add as NA if truly missing
        
        # Apply same placeholder replacement to the final output DataFrame
        final_df_for_output = final_df_for_output.replace(placeholder_map).infer_objects(copy=False)

        # Ensure date columns are stored as plain strings (DD/MM/YY) so Excel will not add time component
        from datetime import datetime as _dt_cls
        date_columns_present = [c for c in ("START_DATE", "END_DATE") if c in final_df_for_output.columns]
        for _dc in date_columns_present:
            def _fmt_date(val):
                if isinstance(val, (pd.Timestamp, _dt_cls)):
                    return "'" + val.strftime('%d/%m/%y')
                if isinstance(val, str):
                    # If looks like ISO datetime, trim time part
                    if '00:00:00' in val or '-' in val:
                        # Attempt to parse
                        try:
                            dt_obj = pd.to_datetime(val, errors='coerce')
                            if pd.notna(dt_obj):
                                return "'" + dt_obj.strftime('%d/%m/%y')
                        except Exception:
                            pass
                    # If already has quote prefix, don't add another
                    if val.startswith("'"):
                        return val
                    # If it's already in dd/mm/yy format, add quote prefix
                    if len(val) == 8 and val[2] == '/' and val[5] == '/':
                        return "'" + val
                    return val
                return val

            final_df_for_output[_dc] = final_df_for_output[_dc].apply(_fmt_date)

        # Save to Excel if output file specified
        if output_file:
            # Save with explicit date column formatting as text
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                final_df_for_output.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                
                # Format date columns as text to prevent Excel auto-conversion
                from openpyxl.utils import get_column_letter
                date_columns = ["START_DATE", "END_DATE"]
                for date_col in date_columns:
                    if date_col in final_df_for_output.columns:
                        col_idx = final_df_for_output.columns.get_loc(date_col) + 1  # Excel columns are 1-based
                        col_letter = get_column_letter(col_idx)
                        for row in range(2, len(final_df_for_output) + 2):  # Start from row 2 (after header)
                            cell = worksheet[f'{col_letter}{row}']
                            cell.number_format = '@'  # Text format
            
            logging.info(f"Saved combined data to '{output_file}' with {len(final_df_for_output)} rows and {len(final_df_for_output.columns)} columns.")
            results['output_file'] = output_file
        
        # Mark as successful if we have no errors
        results['success'] = len(results['errors']) == 0
        
    except Exception as e:
        logging.error(f"Error processing workbook '{input_file}': {str(e)}")
        logging.error(traceback.format_exc())
        results['success'] = False
        results['error'] = str(e)
        results['errors'].append(str(e))
    
    # Calculate processing time
    results['processing_time'] = time.time() - results['start_time']
    
    # Combine log entries for sheets
    sheet_summaries = []
    total_rows = 0
    total_rf_normalized = 0
    
    for sheet_name, sheet_data in results.get('sheet_data', {}).items():
        row_count = sheet_data.get('row_count', 0)
        total_rows += row_count
        
        # Get R&F normalized region count if available
        rf_normalized = sheet_data.get('rf_normalized_regions', 0)
        total_rf_normalized += rf_normalized
        
        # Include R&F info in summary if applicable
        rf_info = ""
        if rf_normalized > 0:
            rf_info = f"\n    - R&F normalized regions: {rf_normalized}"
            
        sheet_summaries.append(f"  Sheet: {sheet_name}\n    - Regions found: {sheet_data.get('regions_found', 0)}\n"
                             f"    - Regions processed: {sheet_data.get('regions_processed', 0)}{rf_info}\n"
                              f"    - Rows processed: {row_count}")
    
    # Log summary
    summary = "Processing summary:\n" + "\n".join(sheet_summaries) + f"\nTotal rows processed: {total_rows}"
    if total_rf_normalized > 0:
        summary += f"\nTotal R&F normalized regions: {total_rf_normalized}"
    summary += f"\nProcessing time: {results['processing_time']:.2f} seconds"
    if output_file:
        summary += f"\nOutput saved to: {output_file}"
    logging.info(summary)
    
    return results


# --- Argument Parsing & Main Execution ---
def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Extract table data from Excel media plan files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""
        Examples:
          # Process both PLANNED and DELIVERED files with combination
          %(prog)s --planned input/planned.xlsx --delivered input/delivered.xlsx --combine
          
          # Process only PLANNED file
          %(prog)s --planned input/planned.xlsx --output output/
          
          # Process with debug logging
          %(prog)s --planned input/planned.xlsx --log-level DEBUG
          
        For more information, see USER_GUIDE.md
        """))
    
    parser.add_argument("-i", "--input-file", 
                       help="Path to the input Excel file (legacy mode).")
    parser.add_argument("-p", "--planned", "--planned-input", 
                       dest="planned_input",
                       help="Path to the PLANNED input Excel file.")
    parser.add_argument("-d", "--delivered", "--delivered-input", 
                       dest="delivered_input",
                       help="Path to the DELIVERED input Excel file.")
    parser.add_argument("-o", "--output", "--output-dir", 
                       dest="output_dir",
                       default="output", 
                       help="Directory to save processed files (default: output).")
    parser.add_argument("-c", "--combine", 
                       action="store_true", 
                       help="Combine PLANNED and DELIVERED data into a single output file.")
    parser.add_argument("-l", "--log-level",
                       choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
                       default=None,
                       help="Set logging level (default: INFO or EXCEL_EXTRACTOR_LOG_LEVEL env var)")
    parser.add_argument("-v", "--version",
                       action="version",
                       version="Excel Data Extractor v1.0.0")
    
    return parser.parse_args()

def main():
    """Main entry point for the script."""
    try:
        # Parse command line arguments first to get log level
        args = parse_arguments()
        
        # Set up logging with optional command-line log level
        setup_logging(args.log_level)
        
        # Check required libraries
        if not check_required_libraries():
            logging.error("Required libraries check failed")
            sys.exit(1)
        
        # Validate that at least one input is provided
        if not args.input_file and not args.planned_input and not args.delivered_input:
            print("Error: Either --input-file or one/both of --planned-input and --delivered-input must be specified.")
            print("\nUsage examples:")
            print("  python excel_data_extractor.py --input-file input.xlsx")
            print("  python excel_data_extractor.py --planned-input planned.xlsx --delivered-input delivered.xlsx --combine")
            sys.exit(1)
        
        # Load configuration from config.json
        try:
            load_config()
        except Exception as e:
            logging.error(f"Failed to load configuration: {str(e)}")
            print(f"\nError: Failed to load configuration from config.json: {str(e)}")
            print("Please ensure config.json exists and is valid JSON.")
            sys.exit(1)
    
        # Create output directory if it doesn't exist
        output_dir = Path(args.output_dir)
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError:
            logging.error(f"Permission denied creating output directory: {output_dir}")
            print(f"\nError: Cannot create output directory '{output_dir}'. Permission denied.")
            sys.exit(1)
        except Exception as e:
            logging.error(f"Failed to create output directory: {str(e)}")
            print(f"\nError: Failed to create output directory '{output_dir}': {str(e)}")
            sys.exit(1)
    
        # Check if we're doing a single file or combined processing
        if args.input_file:
            # Process single file (legacy mode)
            input_path = Path(args.input_file)
            
            # Validate input file
            if not validate_file_path(str(input_path)):
                print(f"\nError: Input file '{input_path}' does not exist or is not readable.")
                sys.exit(1)
            
            # Generate timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{input_path.stem}_{timestamp}_processed{input_path.suffix}"
            output_path = output_dir / output_filename
            
            try:
                results = process_workbook(str(input_path), str(output_path))
                
                # Display results
                if results['errors']:
                    print(f"\nProcessing completed with errors. Check log file: {LOG_FILE}")
                    for error in results['errors'][:5]:  # Show first 5 errors
                        print(f"  - {error}")
                    if len(results['errors']) > 5:
                        print(f"  ... and {len(results['errors']) - 5} more errors")
                elif not results.get('output_file'):
                    print(f"\nProcessing completed, but no data was extracted. Check log file: {LOG_FILE}")
                else:
                    print(f"\nProcessing completed successfully. Output saved to {results['output_file']}")
                    print(f"Detailed logs available in: {LOG_FILE}")
            
            except PermissionError as e:
                logging.error(f"Permission error: {str(e)}")
                print(f"\nError: Cannot access file '{input_path}'.")
                print("The file may be open in Excel or another application. Please close it and try again.")
                sys.exit(1)
            except FileProcessingError as e:
                logging.error(f"File processing error: {str(e)}")
                print(f"\nError processing file: {str(e)}")
                sys.exit(1)
            except Exception as e:
                logging.error(f"Unexpected error: {str(e)}\n{traceback.format_exc()}")
                print(f"\nUnexpected error: {str(e)}")
                print(f"Check log file for details: {LOG_FILE}")
                sys.exit(1)
    
        elif args.planned_input or args.delivered_input:
            # Process PLANNED and/or DELIVERED files
            planned_df = None
            delivered_df = None
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            errors_encountered = False
            
            if args.planned_input:
                planned_path = Path(args.planned_input)
                
                # Validate planned input file
                if not validate_file_path(str(planned_path)):
                    print(f"\nError: PLANNED input file '{planned_path}' does not exist or is not readable.")
                    errors_encountered = True
                else:
                    planned_output = output_dir / f"PLANNED_{timestamp}.xlsx"
                    
                    try:
                        planned_results = process_workbook(str(planned_path), str(planned_output), source_type="PLANNED")
                        
                        if 'combined_data' in planned_results and planned_results['combined_data'] is not None:
                            planned_df = planned_results['combined_data']
                            print(f"✓ PLANNED data processed: {len(planned_df)} rows")
                        else:
                            print(f"✗ Failed to process PLANNED input file. Check log file: {LOG_FILE}")
                            errors_encountered = True
                    
                    except PermissionError:
                        print(f"\nError: Cannot access PLANNED file '{planned_path}'.")
                        print("The file may be open in Excel. Please close it and try again.")
                        errors_encountered = True
                    except FileProcessingError as e:
                        print(f"\nError processing PLANNED file: {str(e)}")
                        errors_encountered = True
                    except Exception as e:
                        logging.error(f"Unexpected error processing PLANNED: {str(e)}\n{traceback.format_exc()}")
                        print(f"\nUnexpected error processing PLANNED file: {str(e)}")
                        errors_encountered = True
            
            if args.delivered_input:
                delivered_path = Path(args.delivered_input)
                
                # Validate delivered input file
                if not validate_file_path(str(delivered_path)):
                    print(f"\nError: DELIVERED input file '{delivered_path}' does not exist or is not readable.")
                    errors_encountered = True
                else:
                    delivered_output = output_dir / f"DELIVERED_{timestamp}.xlsx"
                    
                    try:
                        delivered_results = process_workbook(str(delivered_path), str(delivered_output), source_type="DELIVERED")
                        
                        if 'combined_data' in delivered_results and delivered_results['combined_data'] is not None:
                            delivered_df = delivered_results['combined_data']
                            print(f"✓ DELIVERED data processed: {len(delivered_df)} rows")
                        else:
                            print(f"✗ Failed to process DELIVERED input file. Check log file: {LOG_FILE}")
                            errors_encountered = True
                    
                    except PermissionError:
                        print(f"\nError: Cannot access DELIVERED file '{delivered_path}'.")
                        print("The file may be open in Excel. Please close it and try again.")
                        errors_encountered = True
                    except FileProcessingError as e:
                        print(f"\nError processing DELIVERED file: {str(e)}")
                        errors_encountered = True
                    except Exception as e:
                        logging.error(f"Unexpected error processing DELIVERED: {str(e)}\n{traceback.format_exc()}")
                        print(f"\nUnexpected error processing DELIVERED file: {str(e)}")
                        errors_encountered = True
        
        # Combine the dataframes if requested and both are available
        if args.combine and (planned_df is not None or delivered_df is not None):
            combined_df = None
            
            if planned_df is not None and delivered_df is not None:
                # Ensure Source_Type column exists in both dataframes
                # For planned data, always use "PLANNED"
                planned_df['Source_Type'] = "PLANNED"
                
                # For delivered data, we need to examine the rows to identify R&F metrics
                # Since Source_Type isn't in the file, we'll detect R&F based on the data pattern
                # Typical pattern: R&F data has PLATFORM as the metric name, and populated UNIQUES_REACH or FREQUENCY
                delivered_df['Source_Type'] = "DELIVERED MEDIA"  # Default value
                
                # Flag rows from R&F tables based on PLATFORM and UNIQUES_REACH/FREQUENCY values
                rf_platform_pattern = (
                    delivered_df['PLATFORM'].fillna('').str.contains('Reach', case=False, regex=True) |
                    delivered_df['PLATFORM'].fillna('').str.contains('Freq', case=False, regex=True)
                )
                
                rf_data_pattern = (
                    delivered_df['UNIQUES_REACH'].notna() | 
                    delivered_df['FREQUENCY'].notna()
                )
                
                # Special handling for Campaign-level metrics that are processed but may not have 
                # UNIQUES_REACH/FREQUENCY initially populated
                campaign_metrics_pattern = (
                    delivered_df['PLATFORM'].fillna('').str.contains('Campaign Reach', case=False) |
                    delivered_df['PLATFORM'].fillna('').str.contains('Campaign Freq', case=False)
                )
                
                # Combine the patterns: standard R&F detection OR known Campaign metrics
                rf_pattern = (rf_platform_pattern & rf_data_pattern) | campaign_metrics_pattern
                
                # Set the identified R&F rows
                delivered_df.loc[rf_pattern, 'Source_Type'] = "DELIVERED R&F"
                
                # Use robust market mapper for intelligent combination
                try:
                    from robust_market_mapper import RobustMarketMapper
                    mapper = RobustMarketMapper()
                    combined_df = mapper.map_campaigns(planned_df, delivered_df)
                    print(f"Robust mapping completed: {len(combined_df)} total rows from {len(planned_df)} PLANNED and {len(delivered_df)} DELIVERED.")
                    
                    # Add source type information back
                    if 'match_status' in combined_df.columns:
                        combined_df['Source_Type'] = combined_df['match_status'].map({
                            'planned_only': 'PLANNED',
                            'delivered_only': 'DELIVERED',
                            'matched': 'COMBINED'
                        })
                except ImportError:
                    logging.warning("Robust market mapper not available, using simple concatenation")
                    # Fallback to simple concatenation
                    combined_df = pd.concat([planned_df, delivered_df], ignore_index=True)
                    print(f"Combined {len(planned_df)} PLANNED rows and {len(delivered_df)} DELIVERED rows.")
            elif planned_df is not None:
                combined_df = planned_df
                print(f"Only PLANNED data available for combination: {len(planned_df)} rows.")
            elif delivered_df is not None:
                combined_df = delivered_df
                print(f"Only DELIVERED data available for combination: {len(delivered_df)} rows.")
            
            if combined_df is not None:
                # Ensure all FINAL_OUTPUT_COLUMNS plus Source_Type are present
                output_columns = list(FINAL_OUTPUT_COLUMNS)
                if 'Source_Type' not in output_columns:
                    output_columns.append('Source_Type')
                
                # Initialize final_df_for_output DataFrame
                final_df_for_output = pd.DataFrame()
                
                for col in output_columns:
                    if col in combined_df.columns:
                        final_df_for_output[col] = combined_df[col]
                    else:
                        final_df_for_output[col] = pd.NA
                
                # Save the combined output (moved outside the loop)
                combined_output_filename = f"COMBINED_{timestamp}.xlsx"
                combined_output_path = output_dir / combined_output_filename
                
                try:
                    final_df_for_output.to_excel(str(combined_output_path), index=False)
                    print(f"\n✓ Combined output saved to {combined_output_path} with {len(final_df_for_output)} total rows.")
                    print(f"Detailed logs available in: {LOG_FILE}")
                except PermissionError:
                    print(f"\nError: Cannot write to output file '{combined_output_path}'.")
                    print("The file may be open in Excel. Please close it and try again.")
                    sys.exit(1)
                except Exception as e:
                    logging.error(f"Failed to save combined output: {str(e)}")
                    print(f"\nError: Failed to save combined output: {str(e)}")
                    sys.exit(1)
            
            # Exit with error code if any errors were encountered
            if errors_encountered:
                print(f"\nProcessing completed with errors. Check log file: {LOG_FILE}")
                sys.exit(1)
    
    except KeyboardInterrupt:
        print("\n\nProcessing interrupted by user.")
        sys.exit(130)  # Standard exit code for Ctrl+C
    except Exception as e:
        logging.error(f"Unhandled exception in main: {str(e)}\n{traceback.format_exc()}")
        print(f"\nFatal error: {str(e)}")
        print(f"Check log file for details: {LOG_FILE}")
        sys.exit(1)

if __name__ == "__main__":
    main() 